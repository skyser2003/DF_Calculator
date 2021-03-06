from __future__ import annotations

ver_time='200812'

#-*- coding: utf-8 -*-
## 코드를 무단으로 복제하여 개조 및 배포하지 말 것##
## Copyright ⓒ 2020 Dawnclass(새벽반) dawnclass16@naver.com
import itertools
import os
import random
import threading
import time
import tkinter.font
import tkinter.messagebox
import tkinter.ttk
import urllib.request
import webbrowser
from enum import IntEnum
from functools import partial, reduce
from collections import Counter
from json import loads
from tkinter import *
from typing import List, Dict, Tuple, Callable
from urllib import parse
from concurrent.futures import ProcessPoolExecutor, Future, as_completed
from multiprocessing import Value, Manager

import numpy as np
import openpyxl
from PIL import ImageGrab
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

import calc_fullset
import calc_gif
import calc_list_job
import calc_list_wep
import calc_profile
import calc_result
import calc_setlist
import calc_update
from calc_calc import make_setopt_num, make_set_list, hard_coding_dealer, inv_auto_dealer, make_all_equ_list

'''
상의 - 11 + 01~15, 28~31, [0, 1]
하의 - 12 + 01~15, 24~27, [0]
어깨 - 13 + 01~15, [0]
벨트 - 14 + 01~15, [0]
신발 - 15 + 01~15, 32~35 [0]

팔찌 - 21 + 16~19, 24~27 [0, 1]
목걸이 - 22 + 16~19, 28~31 [0]
반지 - 23 + 16~19, 32~35

보조장비 - 31 + 20~23, 28~31 [0]
마법석 - 32 + 20~23, 24~27, [0]
귀걸이 - 32 + 20~23, 32~35, [0, 1]
'''


class EquipPartCode(IntEnum):
    TOP = 11
    BOTTOM = 12
    SHOULDER = 13
    BELT = 14
    SHOES = 15

    BRACELET = 21
    NECKLACE = 22
    RING = 23

    SUB_EQUIPMENT = 31
    STONE = 32
    EARRING = 33


class ArmorEquipPartCode(IntEnum):
    TOP = EquipPartCode.TOP
    BOTTOM = EquipPartCode.BOTTOM
    SHOULDER = EquipPartCode.SHOULDER
    BELT = EquipPartCode.BELT
    SHOES = EquipPartCode.SHOES


class AccessoryEquipPartCode(IntEnum):
    BRACELET = EquipPartCode.BRACELET
    NECKLACE = EquipPartCode.NECKLACE
    RING = EquipPartCode.RING


class SpecialEquipPartCode(IntEnum):
    SUB_EQUIPMENT = EquipPartCode.SUB_EQUIPMENT
    STONE = EquipPartCode.STONE
    EARRING = EquipPartCode.EARRING


class TopNeckSubEquipPartCode(IntEnum):
    TOP = EquipPartCode.TOP
    NECKLACE = EquipPartCode.NECKLACE
    SUB_EQUIPMENT = EquipPartCode.SUB_EQUIPMENT


class BottomBraceStoneEquipPartCode(IntEnum):
    BOTTOM = EquipPartCode.BOTTOM
    BRACELET = EquipPartCode.BRACELET
    STONE = EquipPartCode.STONE


class ShoesRingEarringEquipPartCode(IntEnum):
    SHOES = EquipPartCode.SHOES
    RING = EquipPartCode.RING
    EARRING = EquipPartCode.EARRING


class SetCode(IntEnum):
    ARMOR_BEGIN = 1
    ARMOR_END = 15

    ACCESSORY_BEGIN = 16
    ACCESSORY_END = 19

    SPECIAL_BEGIN = 20
    SPECIAL_END = 23

    TOP_NECK_SUB_BEGIN = 28
    TOP_NECK_SUB_END = 31

    BOTTOM_BRACE_STONE_BEGIN = 24
    BOTTOM_BRACE_STONE_END = 27

    SHOES_RING_EARRING_BEGIN = 32
    SHOES_RING_EARRING_END = 35


class GodEquipPartCode(IntEnum):
    TOP = EquipPartCode.TOP
    BRACELET = EquipPartCode.BRACELET
    EARRING = EquipPartCode.EARRING


class GodCheckCode(IntEnum):
    NON_GOD = 0
    GOD = 1


exit_calc: Value = None
count_num: Value = None
count_all: Value = None


def init_global(calc_exit_calc: Value, calc_count_num: Value, calc_count_all: Value):
    global exit_calc
    global count_num
    global count_all

    exit_calc = calc_exit_calc
    count_num = calc_count_num
    count_all = calc_count_all


def calc_per_thread(calc_list: List[Tuple[List[str], bool]], thread_id: int, calc_wep_num: int,
                    job_name: str, opt_one: Dict[str, str], opt_buf: Dict[str, str], opt_buflvl: Dict[str, str],
                    inv_tg: int, inv_type_list: List[str], set_perfect: int, wep_num: List, extra_dam: int, extra_cri: int, extra_bon: int, extra_all: int, extra_att: int, extra_sta: int,
                    ele_in: int, fixed_dam: int, fixed_cri: int, extra_pas2: int, cool_eff_dictnum: int, betterang: int,
                    active_eff_one: int,
                    job_lv1,
                    job_lv2,
                    job_lv3,
                    job_lv4,
                    job_lv5,
                    job_lv6,
                    silmari: int,
                    job_pas0,
                    job_pas1,
                    job_pas2,
                    job_pas3,
                    inv2_on_tg: int,
                    job_ult1,
                    job_ult2,
                    job_ult3,
                    ele_skill: int,
                    wep_pre_calced: List[float],
                    cool_eff: float,
                    cool_eff2: float,
                    cool_on: int,
                    cool_pre_calced: List[float],
                    cool_pre_calced2: List[float],
                    wep_name_list_temp: List[str],
                    db_preset,
                    extra_clvl: int,
                    extra_stat: int,
                    inv1_opt: str,
                    inv1_val: str,
                    inv2_opt: str,
                    inv2_val: str,
                    inv3_opt: str,
                    inv3_val: str,
                    inv4_opt: str,
                    inv4_val: str,
                    extra_blvl: int,
                    extra_cstat: int,
                    extra_bstat: float,
                    extra_batt: float,
                    extra_cper: float):
    max_setopt = 0

    thread_save_list = {}  # 딜러1번
    thread_save_list0 = {}  # 딜러2번
    thread_save_list1 = {}  # 버퍼1번
    thread_save_list2 = {}  # 버퍼2번
    thread_save_list3 = {}  # 버퍼3번

    loop_counter = 0

    all_list_god = []
    all_list = []

    for now_all_list, has_god in calc_list:
        loop_counter += 1

        if has_god:
            all_list_god.append(now_all_list)
        else:
            all_list.append(now_all_list)

    print(f"{thread_id} 쓰레드 {loop_counter} 회차 : {len(all_list_god) + len(all_list)}")

    if job_name != "세인트" and job_name != "세라핌" and job_name != "헤카테":
        getone = opt_one.get

        for calc_now in all_list_god:
            if exit_calc.value == 1:
                print(f"Thread {thread_id} stopped")
                return

            set_list, setopt_num = make_setopt_num(calc_now, 1)

            if setopt_num >= max_setopt - set_perfect:
                if setopt_num >= max_setopt: max_setopt = setopt_num
                base_array = np.array(
                    [0, 0, extra_dam, extra_cri, extra_bon, 0, extra_all, extra_att, extra_sta, ele_in,
                     0, 1, 0, 0, 0, 0, 0, 0, extra_pas2, 0, 0, 0, 0, 0, 0, 0, 0, 0])
                max_damper = fixed_dam
                max_criper = fixed_cri
                skiper = 0
                damage = 0
                coolper = 0
                ult_1 = 0
                ult_2 = 0
                ult_3 = 0
                ult_skiper = 0
                calc_wep = wep_num[calc_wep_num] + calc_now
                for_calc = wep_num[calc_wep_num] + make_set_list(calc_now, set_list)
                hard_coding = for_calc.count
                oneone = len(for_calc)
                oneonelist = []
                for i in range(oneone):
                    no_cut = getone(for_calc[i])  ## 11번 스증 ## 20번 쿨감

                    cut = np.array(no_cut[0:20] + no_cut[22:23] + no_cut[34:35] + no_cut[38:44])
                    skiper = (skiper / 100 + 1) * (cut[11] / 100 + 1) * 100 - 100
                    coolper = (1 - (100 - coolper) / 100 * (
                            100 - no_cut[22 + cool_eff_dictnum]) / 100) * 100
                    max_damper = max([no_cut[44], max_damper])
                    max_criper = max([no_cut[45], max_criper])
                    ult_1 = (no_cut[46] / 100 + 1) * (ult_1 / 100 + 1) * 100 - 100
                    ult_2 = (no_cut[47] / 100 + 1) * (ult_2 / 100 + 1) * 100 - 100
                    ult_3 = (no_cut[48] / 100 + 1) * (ult_3 / 100 + 1) * 100 - 100
                    oneonelist.append(cut)
                for i in range(oneone):
                    base_array = base_array + oneonelist[i]

                hard_code_result = hard_coding_dealer(base_array, betterang, for_calc, coolper, skiper)
                base_array = hard_code_result[0]
                coolper = hard_code_result[1]
                skiper = hard_code_result[2]

                base_array[11] = skiper
                base_array[2] = max_damper + base_array[2]
                base_array[3] = max_criper + base_array[3]
                only_bon = base_array[4]
                base_array[4] = base_array[4] + base_array[5] * (base_array[9] * 0.0045 + 1.05)
                actlvl = ((base_array[active_eff_one] + base_array[22] * job_lv1 + base_array[
                    23] * job_lv2 + base_array[24] * job_lv3 +
                           base_array[25] * job_lv4 + base_array[26] * job_lv5 + base_array[
                               27] * job_lv6) / 100 + 1)
                actlvl2 = base_array[22] * (0.5 - silmari * 0.06) * 0.0213 + base_array[24] * (
                        0.5 - silmari * 0.06) * 0.04 + base_array[24] * (
                                  0.1484 - silmari * 0.0284) * 0.0674 + 1
                paslvl = ((100 + base_array[16] * job_pas0) / 100) * (
                        (100 + base_array[17] * job_pas1) / 100) * (
                                 (100 + base_array[18] * job_pas2) / 100) * (
                                 (100 + base_array[19] * job_pas3) / 100)
                if inv_tg == 2:
                    inv_auto = inv_auto_dealer(base_array, only_bon, inv2_on_tg, inv_type_list)
                    base_array = inv_auto[0]
                    only_bon = inv_auto[1]
                    inv1_opt = inv_auto[2]
                    inv2_opt = inv_auto[3]
                    inv1_val = inv_auto[4]
                    inv2_val = inv_auto[5]
                if ult_2 != 0:
                    ult1_per = job_ult1 * (1 + base_array[23] * 0.0653) / actlvl * (ult_1 / 100)
                    ult2_per = job_ult2 * (1 + (base_array[25] * 0.1203 + 0.04348 * base_array[
                        27] * silmari)) / actlvl * (ult_2 / 100)
                    ult3_per = job_ult3 * (1 + base_array[27] * 0.1883) / actlvl * (ult_3 / 100)
                    ult_skiper = (ult1_per + ult2_per + ult3_per) * 100
                real_bon_not_ele = only_bon + base_array[5] * (
                        (base_array[9] - int(ele_skill)) * 0.0045 + 1.05)
                damage = ((base_array[2] / 100 + 1) * (base_array[3] / 100 + 1) * (
                        base_array[4] / 100 + 1) * (base_array[6] / 100 + 1) * (
                                  base_array[7] / 100 + 1) *
                          (base_array[8] / 100 + 1) * (base_array[9] * 0.0045 + 1.05) * (
                                  base_array[10] / 100 + 1) * (skiper / 100 + 1) *
                          paslvl * ((54500 + 3.31 * base_array[0]) / 54500) * (
                                  (4800 + base_array[1]) / 4800) / (
                                  1.05 + 0.0045 * int(ele_skill))) * wep_pre_calced[calc_wep_num]
                final_damage = damage * ((100 / (100 - coolper) - 1) * cool_eff * cool_on + 1) * (
                        (base_array[12] + (actlvl - 1) * 100 + ult_skiper) / 100 + 1) * \
                               cool_pre_calced[calc_wep_num]
                final_damage2 = damage * ((100 / (100 - coolper) - 1) * cool_eff2 + 1) * (
                        (base_array[12] + (actlvl2 - 1) * 100) / 100 + 1) * cool_pre_calced2[
                                    calc_wep_num]
                damage_not_ele = final_damage * (1.05 + 0.0045 * int(ele_skill)) / (
                        base_array[9] * 0.0045 + 1.05) * ((base_array[9] - int(
                    ele_skill)) * 0.0045 + 1.05) / 1.05 * (real_bon_not_ele / 100 + 1) / (
                                         base_array[4] / 100 + 1)
                damage_not_ele2 = final_damage2 * (1.05 + 0.0045 * int(ele_skill)) / (
                        base_array[9] * 0.0045 + 1.05) * ((base_array[9] - int(
                    ele_skill)) * 0.0045 + 1.05) / 1.05 * (real_bon_not_ele / 100 + 1) / (
                                          base_array[4] / 100 + 1)
                damage_only_equ = damage / paslvl / wep_pre_calced[calc_wep_num]

                inv_string = "잔향부여= " + inv1_opt + "(" + str(inv1_val) + "%) / " + inv2_opt + "(" + str(
                    inv2_val) + "%)"
                thread_save_list[final_damage] = [calc_wep, base_array, damage, damage_not_ele, inv_string,
                                                  [ult_1, ult_2, ult_3, ult_skiper], damage_only_equ,
                                                  final_damage2, wep_name_list_temp[calc_wep_num]]
                thread_save_list0[final_damage2] = [calc_wep, base_array, damage, damage_not_ele2, inv_string,
                                                    [ult_1, ult_2, ult_3, ult_skiper], damage_only_equ,
                                                    final_damage, wep_name_list_temp[calc_wep_num]]
                with count_num.get_lock():
                    count_num.value += 1
            else:
                with count_all.get_lock():
                    count_all.value += 1
        # 코드 이름
        # 0추스탯 1추공 2증 3크 4추 5속추
        # 6모 7공 8스탯 9속강 10지속 11스증 12특수
        # 13공속 14크확 / 15 특수액티브 / 16~19 패시브 /20 그로기포함/21 2각캐특수액티브 /22~27 액티브레벨링/
        if max_setopt != 8 or set_perfect == 1:
            for calc_now in all_list:
                if exit_calc.value == 1:
                    print(f"Thread {thread_id} stopped")
                    return

                set_list, setopt_num = make_setopt_num(calc_now, 0)

                if setopt_num >= max_setopt - set_perfect:
                    if setopt_num >= max_setopt: max_setopt = setopt_num
                    base_array = np.array(
                        [0, 0, extra_dam, extra_cri, extra_bon, 0, extra_all, extra_att, extra_sta, ele_in,
                         0, 1, 0, 0, 0, 0, 0, 0, extra_pas2, 0, 0, 0, 0, 0, 0, 0, 0, 0])
                    ult_1 = 0
                    ult_2 = 0
                    ult_3 = 0
                    ult_skiper = 0
                    skiper = 0
                    damage = 0
                    coolper = 0
                    calc_wep = wep_num[calc_wep_num] + calc_now
                    for_calc = wep_num[calc_wep_num] + make_set_list(calc_now, set_list)
                    hard_coding = for_calc.count
                    oneone = len(for_calc)
                    oneonelist = []
                    max_damper = fixed_dam
                    max_criper = fixed_cri
                    for i in range(oneone):
                        no_cut = getone(for_calc[i])  ## 11번 스증
                        cut = np.array(no_cut[0:20] + no_cut[22:23] + no_cut[34:35] + no_cut[38:44])
                        skiper = (skiper / 100 + 1) * (cut[11] / 100 + 1) * 100 - 100
                        coolper = (1 - (100 - coolper) / 100 * (
                                100 - no_cut[22 + cool_eff_dictnum]) / 100) * 100
                        max_damper = max([no_cut[44], max_damper])
                        max_criper = max([no_cut[45], max_criper])
                        ult_1 = (no_cut[46] / 100 + 1) * (ult_1 / 100 + 1) * 100 - 100
                        ult_2 = (no_cut[47] / 100 + 1) * (ult_2 / 100 + 1) * 100 - 100
                        ult_3 = (no_cut[48] / 100 + 1) * (ult_3 / 100 + 1) * 100 - 100
                        oneonelist.append(cut)
                    for i in range(oneone):
                        base_array = base_array + oneonelist[i]

                    hard_code_result = hard_coding_dealer(base_array, betterang, for_calc, coolper, skiper)
                    base_array = hard_code_result[0]
                    coolper = hard_code_result[1]
                    skiper = hard_code_result[2]

                    base_array[11] = skiper
                    base_array[2] = max_damper + base_array[2]
                    base_array[3] = max_criper + base_array[3]
                    only_bon = base_array[4]
                    base_array[4] = base_array[4] + base_array[5] * (base_array[9] * 0.0045 + 1.05)
                    actlvl = ((base_array[active_eff_one] + base_array[22] * job_lv1 + base_array[
                        23] * job_lv2 + base_array[24] * job_lv3 +
                               base_array[25] * job_lv4 + base_array[26] * job_lv5 + base_array[
                                   27] * job_lv6) / 100 + 1)
                    actlvl2 = base_array[22] * (0.5 - silmari * 0.06) * 0.0213 + base_array[24] * (
                            0.5 - silmari * 0.06) * 0.04 + base_array[24] * (
                                      0.1484 - silmari * 0.0284) * 0.0674 + 1
                    paslvl = ((100 + base_array[16] * job_pas0) / 100) * (
                            (100 + base_array[17] * job_pas1) / 100) * (
                                     (100 + base_array[18] * job_pas2) / 100) * (
                                     (100 + base_array[19] * job_pas3) / 100)
                    if inv_tg == 2:
                        inv_auto = inv_auto_dealer(base_array, only_bon, inv2_on_tg, inv_type_list)
                        base_array = inv_auto[0]
                        only_bon = inv_auto[1]
                        inv1_opt = inv_auto[2]
                        inv2_opt = inv_auto[3]
                        inv1_val = inv_auto[4]
                        inv2_val = inv_auto[5]
                    if ult_2 != 0:
                        ult1_per = job_ult1 * (1 + base_array[23] * 0.0653) / actlvl * (ult_1 / 100)
                        ult2_per = job_ult2 * (1 + (base_array[25] * 0.1203 + 0.04348 * base_array[
                            27] * silmari)) / actlvl * (ult_2 / 100)
                        ult3_per = job_ult3 * (1 + base_array[27] * 0.1883) / actlvl * (ult_3 / 100)
                        ult_skiper = (ult1_per + ult2_per + ult3_per) * 100
                    real_bon_not_ele = only_bon + base_array[5] * (
                            (base_array[9] - int(ele_skill)) * 0.0045 + 1.05)
                    damage = ((base_array[2] / 100 + 1) * (base_array[3] / 100 + 1) * (
                            base_array[4] / 100 + 1) * (base_array[6] / 100 + 1) * (
                                      base_array[7] / 100 + 1) *
                              (base_array[8] / 100 + 1) * (base_array[9] * 0.0045 + 1.05) * (
                                      base_array[10] / 100 + 1) * (skiper / 100 + 1) *
                              paslvl * ((54500 + 3.31 * base_array[0]) / 54500) * (
                                      (4800 + base_array[1]) / 4800) / (
                                      1.05 + 0.0045 * int(ele_skill))) * wep_pre_calced[calc_wep_num]
                    final_damage = damage * ((100 / (100 - coolper) - 1) * cool_eff * cool_on + 1) * (
                            (base_array[12] + (actlvl - 1) * 100 + ult_skiper) / 100 + 1) * \
                                   cool_pre_calced[calc_wep_num]
                    final_damage2 = damage * ((100 / (100 - coolper) - 1) * cool_eff2 + 1) * (
                            (base_array[12] + (actlvl2 - 1) * 100) / 100 + 1) * cool_pre_calced2[
                                        calc_wep_num]
                    damage_not_ele = final_damage * (1.05 + 0.0045 * int(ele_skill)) / (
                            base_array[9] * 0.0045 + 1.05) * ((base_array[9] - int(
                        ele_skill)) * 0.0045 + 1.05) / 1.05 * (real_bon_not_ele / 100 + 1) / (
                                             base_array[4] / 100 + 1)
                    damage_not_ele2 = final_damage2 * (1.05 + 0.0045 * int(ele_skill)) / (
                            base_array[9] * 0.0045 + 1.05) * ((base_array[9] - int(
                        ele_skill)) * 0.0045 + 1.05) / 1.05 * (real_bon_not_ele / 100 + 1) / (
                                              base_array[4] / 100 + 1)
                    damage_only_equ = damage / paslvl / wep_pre_calced[calc_wep_num]

                    inv_string = "잔향부여= " + inv1_opt + "(" + str(inv1_val) + "%) / " + inv2_opt + "(" + str(
                        inv2_val) + "%)"
                    thread_save_list[final_damage] = [calc_wep, base_array, damage, damage_not_ele, inv_string,
                                                      [ult_1, ult_2, ult_3, ult_skiper], damage_only_equ,
                                                      final_damage2, wep_name_list_temp[calc_wep_num]]
                    thread_save_list0[final_damage2] = [calc_wep, base_array, damage, damage_not_ele2, inv_string,
                                                        [ult_1, ult_2, ult_3, ult_skiper], damage_only_equ,
                                                        final_damage, wep_name_list_temp[calc_wep_num]]
                    with count_num.get_lock():
                        count_num.value += 1
                else:
                    with count_all.get_lock():
                        count_all.value += 1
        else:
            print('스킵됨')
            with count_all.get_lock():
                count_all.value += len(all_list)
    else:  ##버퍼
        base_b = 10 + int(db_preset['H2'].value) + int(db_preset['H4'].value) + int(
            db_preset['H5'].value) + 1 + extra_blvl
        base_c = 12 + int(db_preset['H3'].value) + 1 + extra_clvl
        base_pas0 = 0
        base_pas0_c = 3
        base_pas0_b = 0
        base_stat_s = 4339 + int(db_preset['H1'].value) + extra_stat - 40  ##2각 꺼지면 -528, 진각 추가로 40 제거
        base_stat_d = int(db_preset['H6'].value) - int(db_preset['H1'].value)
        base_stat_h = 4405 + int(db_preset['H1'].value) + extra_stat - 40  ##2각 꺼지면 -528, 진각 추가로 40 제거
        base_pas0_1 = 0
        lvlget = opt_buflvl.get
        inv_string = "1옵션= " + inv3_opt + " [" + str(inv3_val) + "]\n2옵션= " + inv4_opt + " [" + str(
            inv4_val) + "]"
        # 코드 이름
        # 0 체정 1 지능
        # 축복 2 스탯% 3 물공% 4 마공% 5 독공%
        # 아포 6 고정 7 스탯%
        # 8 축렙 9 포렙
        # 10 아리아/보징증폭
        # 11 전직패 12 보징/크크 13 각패1 14 각패2 15 2각 16 각패3
        # 17 깡신념 18 깡신실 19 아리아쿨 20 하베쿨 21 1각시특수피증(시로코옵션) 22 진각렙

        setget = opt_buf.get
        if len(all_list_god) != 0:
            for calc_now in all_list_god:
                if exit_calc.value == 1:
                    print(f"Thread {thread_id} stopped")
                    return

                set_list, setopt_num = make_setopt_num(calc_now, 1)

                if setopt_num >= max_setopt - set_perfect:
                    base_array = np.array(
                        [base_stat_h, base_stat_s, 0, 0, 0, 0, extra_cstat, 0, base_b, base_c, 0, base_pas0,
                         base_pas0_1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

                    if setopt_num >= max_setopt: max_setopt = setopt_num
                    b_stat = (extra_bstat / 100 + 1) * 1.236384 * 100 - 100  ##탈리스만 8%/8%/6%
                    b_phy = extra_batt
                    b_mag = extra_batt
                    b_ind = extra_batt
                    c_per = extra_cper
                    calc_wep = wep_num[calc_wep_num] + calc_now
                    for_calc = wep_num[calc_wep_num] + make_set_list(calc_now, set_list)
                    hard_coding = for_calc.count
                    oneone = len(for_calc)
                    oneonelist = []
                    for i in range(oneone):
                        no_cut = np.array(setget(for_calc[i]))  ## 2 3 4 5 7
                        base_array = base_array + no_cut
                        b_stat = (b_stat / 100 + 1) * (no_cut[2] / 100 + 1) * 100 - 100
                        b_phy = (b_phy / 100 + 1) * (no_cut[3] / 100 + 1) * 100 - 100
                        b_mag = (b_mag / 100 + 1) * (no_cut[4] / 100 + 1) * 100 - 100
                        b_ind = (b_ind / 100 + 1) * (no_cut[5] / 100 + 1) * 100 - 100
                        c_per = (c_per / 100 + 1) * (no_cut[7] / 100 + 1) * 100 - 100
                        oneonelist.append(no_cut)

                    if hard_coding('1441') == 1:
                        if hard_coding('11440') != 1:  ##3셋 스탯160, 영축힘지8%, 물마독3% 감소
                            base_array[0] = base_array[0] - 160
                            base_array[1] = base_array[1] - 160
                            b_stat = (b_stat / 100 + 1) / 1.08 * 100 - 100
                            b_phy = (b_phy / 100 + 1) / 1.03 * 100 - 100
                            b_mag = (b_mag / 100 + 1) / 1.03 * 100 - 100
                            b_ind = (b_ind / 100 + 1) / 1.03 * 100 - 100

                    if job_name == "세인트":
                        b_base_att = lvlget('hol_b_atta')[int(base_array[8])]
                        stat_pas0lvl_b = lvlget('pas0')[int(base_array[11]) + base_pas0_b] + \
                                         lvlget('hol_pas0_1')[int(base_array[12])]
                        stat_pas0lvl_c = lvlget('pas0')[int(base_array[11]) + base_pas0_c] + \
                                         lvlget('hol_pas0_1')[int(base_array[12])]
                        stat_pas1lvl = lvlget('hol_pas1')[int(base_array[13])] + base_array[17]
                        stat_pas2lvl = lvlget('hol_act2')[int(base_array[15])]
                        stat_pas3lvl = lvlget('pas3')[int(base_array[16])]
                        stat_b = base_array[
                                     0] + stat_pas0lvl_b + stat_pas1lvl + stat_pas2lvl + stat_pas3lvl + 19 * \
                                 base_array[10] + base_stat_d
                        stat_c = base_array[
                                     0] + stat_pas0lvl_c + stat_pas1lvl + stat_pas2lvl + stat_pas3lvl + 19 * \
                                 base_array[10]
                        b_stat_calc = int(
                            int(lvlget('hol_b_stat')[int(base_array[8])] * (b_stat / 100 + 1)) * (
                                    stat_b / 620 + 1))
                        b_phy_calc = int(int(b_base_att * (b_phy / 100 + 1)) * (stat_b / 620 + 1))
                        b_mag_calc = int(int(b_base_att * (b_mag / 100 + 1)) * (stat_b / 620 + 1))
                        b_ind_calc = int(int(b_base_att * (b_ind / 100 + 1)) * (stat_b / 620 + 1))
                        b_average = int((b_phy_calc + b_mag_calc + b_ind_calc) / 3) + 42 + int(
                            lvlget('hol_act0')[int(base_array[12])])
                        c_calc2 = int(int((lvlget('c_stat')[int(base_array[9])] + base_array[6]) * (
                                1.1 + 0.01 * int(base_array[22])) * (c_per / 100 + 1)) * (
                                              stat_c / 750 + 1))
                        c_calc = int(int((lvlget('c_stat')[int(base_array[9])] + base_array[6]) * (
                                1.25 + 0.01 * int(base_array[22])) * (c_per / 100 + 1)) * (
                                             stat_c / 750 + 1))
                        pas1_calc = int(lvlget('hol_pas1_out')[int(base_array[13])] + 273)
                        pas1_out = str(int(lvlget('hol_pas1_out')[int(base_array[13])] + 273)) + "(" + str(
                            int(20 + base_array[13])) + "렙)"
                        save1 = '스탯=' + str(b_stat_calc) + "\n앞뎀=" + str(b_average) + "\n\n적용스탯= " + str(
                            int(stat_b)) + "\n적용레벨= " + str(int(base_array[8])) + "렙"
                        save2 = '스탯= ' + str(c_calc) + '(' + str(c_calc2) + ')' + "\n\n적용스탯= " + str(
                            int(stat_c)) + "\n적용레벨= " + str(int(base_array[9])) + "렙"

                    else:
                        if job_name == "세라핌":
                            b_value = 665
                            aria = 1.3
                            amuguna = 1.15
                            amuguna2 = 1.15
                            amuguna_stat = 40
                            crux = 0
                        if job_name == "헤카테":
                            b_value = 665
                            aria = 1.25 * 1.15
                            amuguna = 1.25 + 0.01 * int(base_array[22])
                            amuguna2 = 1.1 + 0.01 * int(base_array[22])
                            amuguna_stat = 0
                            crux = 42 + int(base_array[13])

                        b_base_att = lvlget('se_b_atta')[int(base_array[8])]
                        stat_pas0lvl_b = lvlget('pas0')[int(base_array[11]) + int(base_pas0_b)]
                        stat_pas0lvl_c = lvlget('pas0')[int(base_array[11]) + int(base_pas0_c)]
                        stat_pas1lvl = lvlget('se_pas1')[int(base_array[13])] + base_array[18]
                        stat_pas2lvl = lvlget('se_pas2')[int(base_array[14])]
                        stat_pas3lvl = lvlget('pas3')[int(base_array[16])]
                        stat_b = base_array[
                                     1] + stat_pas0lvl_b + stat_pas1lvl + stat_pas2lvl + stat_pas3lvl + base_stat_d + amuguna_stat
                        stat_c = base_array[
                                     1] + stat_pas0lvl_c + stat_pas1lvl + stat_pas2lvl + stat_pas3lvl + amuguna_stat
                        b_stat_calc = int(
                            int(lvlget('se_b_stat')[int(base_array[8])] * (b_stat / 100 + 1)) * (
                                    stat_b / b_value + 1) * aria)
                        b_phy_calc = int(
                            int(b_base_att * (b_phy / 100 + 1) * (stat_b / b_value + 1)) * aria)
                        b_mag_calc = int(
                            int(b_base_att * (b_mag / 100 + 1) * (stat_b / b_value + 1)) * aria)
                        b_ind_calc = int(
                            int(b_base_att * (b_ind / 100 + 1) * (stat_b / b_value + 1)) * aria)
                        b_average = int((b_phy_calc + b_mag_calc + b_ind_calc) / 3) + crux
                        c_calc = int(int(
                            (lvlget('c_stat')[int(base_array[9])] + base_array[6]) * amuguna * (
                                    c_per / 100 + 1)) * (stat_c / 750 + 1))
                        c_calc2 = 0
                        pas1_calc = int(stat_pas1lvl + 442)
                        pas1_out = str(int(stat_pas1lvl + 442)) + "(" + str(int(20 + base_array[13])) + "렙)"
                        save1 = '스탯=' + str(b_stat_calc) + "(" + str(
                            int(b_stat_calc / aria)) + ")\n앞뎀=" + str(b_average) + "(" + str(
                            int(b_average / aria)) + ")\n\n적용스탯= " + str(int(stat_b)) + "\n적용레벨= " + str(
                            int(base_array[8])) + "렙"
                        save2 = '스탯= ' + str(c_calc) + "\n\n적용스탯= " + str(int(stat_c)) + "\n적용레벨= " + str(
                            int(base_array[9])) + "렙"
                        if job_name == "헤카테":
                            c_calc2 = int(int(
                                (lvlget('c_stat')[int(base_array[9])] + base_array[6]) * amuguna2 * (
                                        c_per / 100 + 1)) * (stat_c / 750 + 1))
                            save2 = '스탯= ' + str(c_calc) + '(' + str(c_calc2) + ')' + "\n\n적용스탯= " + str(
                                int(stat_c)) + "\n적용레벨= " + str(int(base_array[9])) + "렙"
                    ##1축 2포 3합
                    final_buf1 = ((15000 + b_stat_calc) / 250 + 1) * (2650 + b_average)
                    final_buf2 = ((15000 + c_calc) / 250 + 1) * 2650 * (base_array[21] / 100 + 1)
                    final_buf2_2 = ((15000 + c_calc2) / 250 + 1) * 2650 * (base_array[21] / 100 + 1)
                    final_buf3 = ((15000 + pas1_calc + c_calc + b_stat_calc) / 250 + 1) * (
                            2650 + b_average) * (base_array[21] / 100 + 1)
                    final_buf3_2 = ((15000 + pas1_calc + c_calc2 + b_stat_calc) / 250 + 1) * (
                            2650 + b_average) * (base_array[21] / 100 + 1)
                    thread_save_list1[final_buf1] = [list(calc_wep), [save1, save2, pas1_out], inv_string,
                                                     base_array, final_buf1, wep_name_list_temp[calc_wep_num]]
                    thread_save_list2[final_buf2] = [list(calc_wep), [save1, save2, pas1_out], inv_string,
                                                     base_array, final_buf2_2, wep_name_list_temp[calc_wep_num]]
                    thread_save_list3[final_buf3] = [list(calc_wep), [save1, save2, pas1_out], inv_string,
                                                     base_array, final_buf3_2, wep_name_list_temp[calc_wep_num]]

                    with count_num.get_lock():
                        count_num.value += 1
                else:
                    with count_all.get_lock():
                        count_all.value += 1

        if max_setopt != 8 or set_perfect == 1:
            for calc_now in all_list:
                if exit_calc.value == 1:
                    print(f"Thread {thread_id} stopped")
                    return

                set_list, setopt_num = make_setopt_num(calc_now, 0)

                if setopt_num >= max_setopt - set_perfect:
                    base_array = np.array(
                        [base_stat_h, base_stat_s, 0, 0, 0, 0, extra_cstat, 0, base_b, base_c, 0, base_pas0,
                         base_pas0_1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

                    if setopt_num >= max_setopt: max_setopt = setopt_num
                    b_stat = (extra_bstat / 100 + 1) * 1.236384 * 100 - 100  ##탈리스만 8%/8%/6%
                    b_phy = extra_batt
                    b_mag = extra_batt
                    b_ind = extra_batt
                    c_per = extra_cper
                    calc_wep = wep_num[calc_wep_num] + calc_now
                    for_calc = wep_num[calc_wep_num] + make_set_list(calc_now, set_list)
                    hard_coding = for_calc.count
                    oneone = len(for_calc)
                    oneonelist = []
                    for i in range(oneone):
                        no_cut = np.array(setget(for_calc[i]))  ## 2 3 4 5 7
                        base_array = base_array + no_cut
                        b_stat = (b_stat / 100 + 1) * (no_cut[2] / 100 + 1) * 100 - 100
                        b_phy = (b_phy / 100 + 1) * (no_cut[3] / 100 + 1) * 100 - 100
                        b_mag = (b_mag / 100 + 1) * (no_cut[4] / 100 + 1) * 100 - 100
                        b_ind = (b_ind / 100 + 1) * (no_cut[5] / 100 + 1) * 100 - 100
                        c_per = (c_per / 100 + 1) * (no_cut[7] / 100 + 1) * 100 - 100
                        oneonelist.append(no_cut)

                    if hard_coding('1441') == 1:
                        if hard_coding('11440') != 1:  ##3셋 스탯160, 영축힘지8%, 물마독3% 감소
                            base_array[0] = base_array[0] - 160
                            base_array[1] = base_array[1] - 160
                            b_stat = (b_stat / 100 + 1) / 1.08 * 100 - 100
                            b_phy = (b_phy / 100 + 1) / 1.03 * 100 - 100
                            b_mag = (b_mag / 100 + 1) / 1.03 * 100 - 100
                            b_ind = (b_ind / 100 + 1) / 1.03 * 100 - 100

                    if job_name == "세인트":
                        b_base_att = lvlget('hol_b_atta')[int(base_array[8])]
                        stat_pas0lvl_b = lvlget('pas0')[int(base_array[11]) + base_pas0_b] + \
                                         lvlget('hol_pas0_1')[int(base_array[12])]
                        stat_pas0lvl_c = lvlget('pas0')[int(base_array[11]) + base_pas0_c] + \
                                         lvlget('hol_pas0_1')[int(base_array[12])]
                        stat_pas1lvl = lvlget('hol_pas1')[int(base_array[13])] + base_array[17]
                        stat_pas2lvl = lvlget('hol_act2')[int(base_array[15])]
                        stat_pas3lvl = lvlget('pas3')[int(base_array[16])]
                        stat_b = base_array[
                                     0] + stat_pas0lvl_b + stat_pas1lvl + stat_pas2lvl + stat_pas3lvl + 19 * \
                                 base_array[10] + base_stat_d
                        stat_c = base_array[
                                     0] + stat_pas0lvl_c + stat_pas1lvl + stat_pas2lvl + stat_pas3lvl + 19 * \
                                 base_array[10]
                        b_stat_calc = int(
                            int(lvlget('hol_b_stat')[int(base_array[8])] * (b_stat / 100 + 1)) * (
                                    stat_b / 630 + 1))
                        b_phy_calc = int(
                            int(b_base_att * (b_phy / 100 + 1)) * (stat_b / 620 + 1) + lvlget('hol_act0')[
                                int(base_array[12])]) + 42
                        b_mag_calc = int(
                            int(b_base_att * (b_mag / 100 + 1)) * (stat_b / 620 + 1) + lvlget('hol_act0')[
                                int(base_array[12])]) + 42
                        b_ind_calc = int(
                            int(b_base_att * (b_ind / 100 + 1)) * (stat_b / 620 + 1) + lvlget('hol_act0')[
                                int(base_array[12])]) + 42
                        b_average = int((b_phy_calc + b_mag_calc + b_ind_calc) / 3)
                        c_calc2 = int(int((lvlget('c_stat')[int(base_array[9])] + base_array[6]) * (
                                1.1 + 0.01 * int(base_array[22])) * (c_per / 100 + 1)) * (
                                              stat_c / 750 + 1))
                        c_calc = int(int((lvlget('c_stat')[int(base_array[9])] + base_array[6]) * (
                                1.25 + 0.01 * int(base_array[22])) * (c_per / 100 + 1)) * (
                                             stat_c / 750 + 1))
                        pas1_calc = int(lvlget('hol_pas1_out')[int(base_array[13])] + 273)
                        pas1_out = str(int(lvlget('hol_pas1_out')[int(base_array[13])] + 273)) + "(" + str(
                            int(20 + base_array[13])) + "렙)"
                        save1 = '스탯=' + str(b_stat_calc) + "\n앞뎀=" + str(b_average) + "\n\n적용스탯= " + str(
                            int(stat_b)) + "\n적용레벨= " + str(int(base_array[8])) + "렙"
                        save2 = '스탯= ' + str(c_calc) + '(' + str(c_calc2) + ')' + "\n\n적용스탯= " + str(
                            int(stat_c)) + "\n적용레벨= " + str(int(base_array[9])) + "렙"

                    else:
                        if job_name == "세라핌":
                            b_value = 665
                            aria = 1.3
                            amuguna = 1.15
                            amuguna2 = 1.15
                            amuguna_stat = 40
                            crux = 0
                        if job_name == "헤카테":
                            b_value = 665
                            aria = 1.25 * 1.15
                            amuguna = 1.25 + 0.01 * int(base_array[22])
                            amuguna2 = 1.1 + 0.01 * int(base_array[22])
                            amuguna_stat = 0
                            crux = 42 + int(base_array[13])

                        b_base_att = lvlget('se_b_atta')[int(base_array[8])]
                        stat_pas0lvl_b = lvlget('pas0')[int(base_array[11]) + int(base_pas0_b)]
                        stat_pas0lvl_c = lvlget('pas0')[int(base_array[11]) + int(base_pas0_c)]
                        stat_pas1lvl = lvlget('se_pas1')[int(base_array[13])] + base_array[18]
                        stat_pas2lvl = lvlget('se_pas2')[int(base_array[14])]
                        stat_pas3lvl = lvlget('pas3')[int(base_array[16])]
                        stat_b = base_array[
                                     1] + stat_pas0lvl_b + stat_pas1lvl + stat_pas2lvl + stat_pas3lvl + base_stat_d + amuguna_stat
                        stat_c = base_array[
                                     1] + stat_pas0lvl_c + stat_pas1lvl + stat_pas2lvl + stat_pas3lvl + amuguna_stat
                        b_stat_calc = int(
                            int(lvlget('se_b_stat')[int(base_array[8])] * (b_stat / 100 + 1)) * (
                                    stat_b / b_value + 1) * aria)
                        b_phy_calc = int(
                            int(b_base_att * (b_phy / 100 + 1) * (stat_b / b_value + 1)) * aria)
                        b_mag_calc = int(
                            int(b_base_att * (b_mag / 100 + 1) * (stat_b / b_value + 1)) * aria)
                        b_ind_calc = int(
                            int(b_base_att * (b_ind / 100 + 1) * (stat_b / b_value + 1)) * aria)
                        b_average = int((b_phy_calc + b_mag_calc + b_ind_calc) / 3) + crux
                        c_calc = int(int(
                            (lvlget('c_stat')[int(base_array[9])] + base_array[6]) * amuguna * (
                                    c_per / 100 + 1)) * (stat_c / 750 + 1))
                        c_calc2 = 0
                        pas1_calc = int(stat_pas1lvl + 442)
                        pas1_out = str(int(stat_pas1lvl + 442)) + "(" + str(int(20 + base_array[13])) + "렙)"
                        save1 = '스탯=' + str(b_stat_calc) + "(" + str(
                            int(b_stat_calc / aria)) + ")\n앞뎀=" + str(b_average) + "(" + str(
                            int(b_average / aria)) + ")\n\n적용스탯= " + str(int(stat_b)) + "\n적용레벨= " + str(
                            int(base_array[8])) + "렙"
                        save2 = '스탯= ' + str(c_calc) + "\n\n적용스탯= " + str(int(stat_c)) + "\n적용레벨= " + str(
                            int(base_array[9])) + "렙"
                        if job_name == "헤카테":
                            c_calc2 = int(int(
                                (lvlget('c_stat')[int(base_array[9])] + base_array[6]) * amuguna2 * (
                                        c_per / 100 + 1)) * (stat_c / 750 + 1))
                            save2 = '스탯= ' + str(c_calc) + '(' + str(c_calc2) + ')' + "\n\n적용스탯= " + str(
                                int(stat_c)) + "\n적용레벨= " + str(int(base_array[9])) + "렙"
                    ##1축 2포 3합
                    final_buf1 = ((15000 + b_stat_calc) / 250 + 1) * (2650 + b_average)
                    final_buf2 = ((15000 + c_calc) / 250 + 1) * 2650 * (base_array[21] / 100 + 1)
                    final_buf2_2 = ((15000 + c_calc2) / 250 + 1) * 2650 * (base_array[21] / 100 + 1)
                    final_buf3 = ((15000 + pas1_calc + c_calc + b_stat_calc) / 250 + 1) * (
                            2650 + b_average) * (base_array[21] / 100 + 1)
                    final_buf3_2 = ((15000 + pas1_calc + c_calc2 + b_stat_calc) / 250 + 1) * (
                            2650 + b_average) * (base_array[21] / 100 + 1)
                    thread_save_list1[final_buf1] = [list(calc_wep), [save1, save2, pas1_out], inv_string,
                                                     base_array, final_buf1, wep_name_list_temp[calc_wep_num]]
                    thread_save_list2[final_buf2] = [list(calc_wep), [save1, save2, pas1_out], inv_string,
                                                     base_array, final_buf2_2, wep_name_list_temp[calc_wep_num]]
                    thread_save_list3[final_buf3] = [list(calc_wep), [save1, save2, pas1_out], inv_string,
                                                     base_array, final_buf3_2, wep_name_list_temp[calc_wep_num]]

                    with count_num.get_lock():
                        count_num.value += 1
                else:
                    with count_all.get_lock():
                        count_all.value += 1
        else:
            print('스킵됨')
            with count_all.get_lock():
                count_all.value += len(all_list_god)

    return thread_save_list, thread_save_list0, thread_save_list1, thread_save_list2, thread_save_list3


class Calculator:
    def __init__(self):
        self.result_window: Toplevel = None
        self.canvas_res: Canvas = None
        self.photo_images = []
        self.all_list_list_num = 0  # 계산 전체 경우의 수
        self.a_num_all = 0
        self.equip_list = {}

        self.exit_calc = Value("i", 0)  # 계산 종료 판정
        self.count_num = Value("i", 0)  # 유효 계산 카운터
        self.count_all = Value("i", 0)  # 전체 계산 카운터

        self.jobtype_select: tkinter.ttk.Combobox = None
        self.jobup_select: tkinter.ttk.Combobox = None
        self.wep_job_select: tkinter.ttk.Combobox = None
        self.wep_type_select: tkinter.ttk.Combobox = None
        self.wep_select: tkinter.ttk.Combobox = None
        self.select_perfect: tkinter.ttk.Combobox = None
        self.style_select: tkinter.ttk.Combobox = None
        self.creature_select: tkinter.ttk.Combobox = None
        self.req_cool: tkinter.ttk.Combobox = None
        self.inv_mod: tkinter.ttk.Combobox = None
        self.inv_select1_1: tkinter.ttk.Combobox = None
        self.inv_select1_2: tkinter.ttk.Combobox = None
        self.inv_select2_1: tkinter.ttk.Combobox = None
        self.inv_select2_2: tkinter.ttk.Combobox = None
        self.inv_select3_1: tkinter.ttk.Combobox = None
        self.inv_select3_2: tkinter.ttk.Combobox = None
        self.inv_select4_1: tkinter.ttk.Combobox = None
        self.inv_select4_2: tkinter.ttk.Combobox = None

        self.preset_values: Dict[str, Callable[[], tkinter.ttk.Combobox]] = {
            "jobtype_select": lambda: self.jobtype_select,
            "jobup_select": lambda: self.jobup_select,
            "wep_job_select": lambda: self.wep_job_select,
            "wep_type_select": lambda: self.wep_type_select,
            "wep_select": lambda: self.wep_select,
            "select_perfect": lambda: self.select_perfect,
            "style_select": lambda: self.style_select,
            "creature_select": lambda: self.creature_select,
            "req_cool": lambda: self.req_cool,
            "inv_mod": lambda: self.inv_mod,
            "inv_select1_1": lambda: self.inv_select1_1,
            "inv_select1_2": lambda: self.inv_select1_2,
            "inv_select2_1": lambda: self.inv_select2_1,
            "inv_select2_2": lambda: self.inv_select2_2,
            "inv_select3_1": lambda: self.inv_select3_1,
            "inv_select3_2": lambda: self.inv_select3_2,
            "inv_select4_1": lambda: self.inv_select4_1,
            "inv_select4_2": lambda: self.inv_select4_2
        }

        self.show_number = 0  # 숫자 갱신 여부
        self.inv_tg = 0  # 잔향 부여 선택(0:미부여,1:선택부여,2:최적부여)
        self.wep_name_list: List[str] = []
        self.style_calced = ""
        self.creature_calced = ""
        self.default_base_equip = 0  # 레전더리(0), 차원 레전더리(1), 초오광(2)
        self.image_list_wep: Dict[str, PhotoImage] = {}
        self.image_list: Dict[str, PhotoImage] = {}
        self.image_list2: Dict[str, PhotoImage] = {}
        self.image_list_tag: Dict[str, PhotoImage] = {}
        self.image_list_set: Dict[str, PhotoImage] = {}
        self.image_list_set2: Dict[str, PhotoImage] = {}
        self.set_name_toggle = 0
        self.now_version = "4.0.3"
        self.pause_gif = 0
        self.stop_gif = 0
        self.stop_gif2 = 0
        self.gif_images = {}
        self.now_rank_num = 0
        self.res_wep = None
        self.res_wep_img = []
        self.rank_setting: List[int] = []
        self.rank0_setting: List[int] = []
        self.deal_rank_wep_name: List[List[int]] = [[], []]
        self.buff_rank_wep_name: List[List[int]] = [[], [], []]
        self.deal_rank_wep_img: List[List[int]] = [[], []]
        self.buff_rank_wep_img: List[List[int]] = [[], [], []]
        self.deal_result_image_on: List[List[Dict[str, PhotoImage]]] = [[], []]
        self.buff_result_image_on: List[List[Dict[str, PhotoImage]]] = [[], [], []]
        self.deal_result_image_tag: List[List[Dict[str, str]]] = [[], []]
        self.rank_dam: List[List[int]] = [[], []]
        self.deal_rank_stat1: List[List[int]] = [[], []]
        self.deal_rank_stat2: List[List[int]] = [[], []]
        self.deal_rank_stat3: List[List[int]] = [[], []]
        self.res_dam: Text = None
        self.res_stat: Text = None
        self.res_stat2: Text = None
        self.res_stat3: Text = None
        self.rank_dam_tagk: List[List[int]] = [[], []]
        self.rank_dam_nolv: List[List[int]] = [[], []]
        self.rank_dam_noele: List[List[int]] = [[], []]
        self.rank_dam_tagk_nolv: List[List[int]] = [[], []]
        self.rank_dam_tagk_noele: List[List[int]] = [[], []]
        self.res_ele: Text = None
        self.deal_rank_inv: List[List[int]] = [[], []]
        self.buff_rank_inv: List[List[int]] = [[], [], []]
        self.res_inv: Text = None
        self.deal_result_image_gif: List[List[List[int]]] = [[], []]
        self.deal_result_image_gif_tg: List[List[List[int]]] = [[], []]
        self.deal_result_siroco_gif: List[List[List[int]]] = [[], []]
        self.deal_result_siroco_gif_tg: List[List[List[int]]] = [[], []]
        self.res_cool_what: Text = None
        self.cool_eff_text = ""
        self.rank_list: List[List[Tuple[float, List[str]]]] = [[], []]
        self.tagkgum_exist = 0
        self.tagk_tg = 0
        self.tagkgum: Button = None
        self.tagkgum_img: PhotoImage = None
        self.rank_stat_tagk_um: List[List[int]] = [[], []]
        self.rank_stat_tagk_yang: List[List[int]] = [[], []]
        self.res_dam_list: List[int] = []
        self.res_item_list: List[Dict[str, Image]] = []
        self.tg_groggy = 0
        self.groggy_bt: Button = None
        self.tg_groggy_img1: PhotoImage = None
        self.tg_groggy_img2: PhotoImage = None
        self.result_cool_canvas_list: List[int] = []
        self.rank_buf: List[List[int]] = [[], [], []]
        self.rank_type_buf = 3
        self.res_buf: Text = None
        self.res_img_list: Dict[str, Image] = {}
        self.res_buf_list: Dict[str, Text] = {}
        self.res_buf_ex: List[Text] = []
        self.rank_buf_ex: List[List[int]] = [[], [], []]
        self.res_buf_type_what: Text = None
        self.result_image_on_tag: List[List[Dict[str, str]]] = [[], [], []]
        self.buff_result_image_gif: List[List[List[int]]] = [[], [], []]
        self.buff_result_image_gif_tg: List[List[List[int]]] = [[], [], []]
        self.buff_result_siroco_gif: List[List[List[int]]] = [[], [], []]
        self.buff_result_siroco_gif_tg: List[List[List[int]]] = [[], [], []]
        self.rank_neo_buf = [[], [], []]
        self.buf_jingak_exist = 0
        self.buf_jingak_tg = 0
        self.buf_jingak: Button = None
        self.auto_saved = 0  # 클라이언트 업데이트 시 preset 업데이트 변수
        self.save_name_list: List[str] = []  # 프리셋 이름 리스트
        self.selected_weapon_img_list: List[PhotoImage] = []
        self.selected_weapon_label_list: List[Label] = []
        self.owned_equipments: Dict[str, int] = {} # 장비 선택 시 점등
        self.auto_custom = 0  # 클라이언트 업데이트 시 preset 업데이트 여부
        self.opt_one: Dict[str, str] = {}
        self.opt_job: Dict[str, str] = {}
        self.opt_job_ele: Dict[str, str] = {}
        self.opt_buf: Dict[str, str] = {}
        self.opt_buflvl: Dict[str, str] = {}
        self.opt_leveling: Dict[str, str] = {}
        self.server_list = ["카인", "디레지에", "바칼", "힐더", "안톤", "카시야스", "프레이", "시로코"]
        self.wep_list: List[str] = []
        self.inv_type_list = ["증뎀", "크증", "추뎀", "모공", "공%", "스탯"]

        # 에픽 장비
        self.normal_epic_list: List[str] = []

        # 지혜의 산물
        self.know_list: List[str] = []
        self.know_set_list: List[str] = []
        self.know_jin_list: List[str] = []

        # 시로코 에픽
        self.siroco_equip_list: List[str] = []

        ## GUI 메인
        self.window = tkinter.Tk()

        window = self.window
        window.title("에픽 조합 자동 계산기")
        window.geometry("910x720+0+0")
        window.resizable(False, False)
        self.place_center(window, 0)
        window.iconbitmap(r'ext_img/icon.ico')

        self.dark_main = self._from_rgb((32, 34, 37))
        self.dark_sub = self._from_rgb((50, 46, 52))
        self.dark_blue = self._from_rgb((29, 30, 36))
        self.result_sub = self._from_rgb((31, 28, 31))

        self.guide_font = tkinter.font.Font(family="맑은 고딕", size=10, weight='bold')
        self.small_font = tkinter.font.Font(family="맑은 고딕", size=8, weight='bold')
        self.mid_font = tkinter.font.Font(family="맑은 고딕", size=14, weight='bold')

        bg_img = self.get_photo_image("ext_img/bg_img.png")
        bg_wall = tkinter.Label(window, image=bg_img)
        bg_wall.place(x=-2, y=0)
        window.configure(bg=self.dark_main)

        self.buf_jingak_img: List[PhotoImage] = [
            self.get_photo_image("ext_img/buf_jin_1.png"),
            self.get_photo_image("ext_img/buf_jin_2.png")
        ]
        self.save_select: tkinter.ttk.Combobox = None

        self.case_count_label: Label = None
        self.total_count_label: Label = None
        self.calc_state_label: Label = None
        self.weapon_list_num_label: Label = None
        self.void_weapon_img: PhotoImage = None
        self.equip_buttons: Dict[str, Button] = {}
        self.set_buttons: Dict[str, Button] = {}

        self.refresh_db()
        self.load_preset_name()

        self.init_images()
        self.init_equipments()
        self.init_ui()

        self.update_preset_version()

    def get_photo_image(self, file: str):
        photo_image = PhotoImage(file=file)
        self.photo_images.append(photo_image)

        return photo_image

    # 통합 커스텀 설정창
    def create_custom_window(self, auto: int):
        try:
            self.custom_window.destroy()
        except:
            pass

        self.custom_window = tkinter.Toplevel(self.window)
        custom_window = self.custom_window
        custom_window.attributes("-topmost", True)
        custom_window.geometry("620x400+750+20")

        load_preset = load_workbook("preset.xlsx", data_only=True)
        db_preset = load_preset["custom"]

        tkinter.Label(custom_window, text="<쿨감보정>", font=self.mid_font).place(x=100, y=10)
        tkinter.Label(custom_window, text="그로기   =          %", font=self.guide_font).place(x=160,
                                                                                                  y=50)  ##Y11/Z11
        cool_con = tkinter.Entry(custom_window, width=5)
        cool_con.place(x=230, y=52)
        cool_con.insert(END, db_preset['B2'].value)
        tkinter.Label(custom_window, text="지속딜   =          %", font=self.guide_font).place(x=10, y=50)  ##Y11/Z11
        cool_con2 = tkinter.Entry(custom_window, width=5)
        cool_con2.place(x=80, y=52)
        cool_con2.insert(END, db_preset['B20'].value)

        tkinter.Label(custom_window, text="<딜러장비>", font=self.mid_font).place(x=100, y=85)
        tkinter.Label(custom_window, text="% 입력창은 그만큼 %딜증가로 환산한다는 뜻", fg="Red").place(x=30, y=120)
        tkinter.Label(custom_window, text="선택벨트=  자동", font=self.guide_font).place(x=160, y=155)  ##O164
        cus1 = tkinter.Entry(custom_window, width=5)
        cus1.insert(END, 0)
        tkinter.Label(custom_window, text="선택신발=          %", font=self.guide_font).place(x=160, y=185)  ##O180
        cus2 = tkinter.Entry(custom_window, width=5)
        cus2.place(x=230, y=187)
        cus2.insert(END, db_preset['B4'].value)
        tkinter.Label(custom_window, text="베테랑=", font=self.guide_font).place(x=160, y=215)  ##G276
        lvl_list = ['전설↓', '영웅↑']
        cus3 = tkinter.ttk.Combobox(custom_window, width=5, values=lvl_list)
        cus3.place(x=230, y=217)
        cus3.set(db_preset['B12'].value)
        tkinter.Label(custom_window, text="먼동강화=          강", font=self.guide_font).place(x=160, y=245)
        lvl_list = [10, 11, 12, 13]
        cus4 = tkinter.ttk.Combobox(custom_window, width=2, values=lvl_list)
        cus4.place(x=230, y=247)
        cus4.set(db_preset['B13'].value)

        tkinter.Label(custom_window, text="흐름상의=          %", font=self.guide_font).place(x=10, y=155)  ##O100
        cus6 = tkinter.Entry(custom_window, width=5)
        cus6.place(x=80, y=157)
        cus6.insert(END, db_preset['B5'].value)
        tkinter.Label(custom_window, text="흐름하의=          %", font=self.guide_font).place(x=10, y=185)  ##O127
        cus7 = tkinter.Entry(custom_window, width=5)
        cus7.place(x=80, y=187)
        cus7.insert(END, db_preset['B6'].value)
        tkinter.Label(custom_window, text="흐름어깨=          %", font=self.guide_font).place(x=10, y=215)  ##O147
        cus8 = tkinter.Entry(custom_window, width=5)
        cus8.place(x=80, y=217)
        cus8.insert(END, db_preset['B7'].value)
        tkinter.Label(custom_window, text="흐름벨트=          %", font=self.guide_font).place(x=10, y=245)  ##O163
        cus9 = tkinter.Entry(custom_window, width=5)
        cus9.place(x=80, y=247)
        cus9.insert(END, db_preset['B8'].value)
        tkinter.Label(custom_window, text="흐름신발=          %", font=self.guide_font).place(x=10, y=275)  ##O179
        cus10 = tkinter.Entry(custom_window, width=5)
        cus10.place(x=80, y=277)
        cus10.insert(END, db_preset['B9'].value)
        tkinter.Label(custom_window, text="흐름2셋=           %", font=self.guide_font).place(x=10, y=305)  ##O295
        cus11 = tkinter.Entry(custom_window, width=5)
        cus11.place(x=80, y=307)
        cus11.insert(END, db_preset['B10'].value)
        tkinter.Label(custom_window, text="흐름3셋=           %", font=self.guide_font).place(x=10,
                                                                                                 y=335)  ##O296,O297
        cus12 = tkinter.Entry(custom_window, width=5)
        cus12.place(x=80, y=337)
        cus12.insert(END, db_preset['B11'].value)

        tkinter.Label(custom_window, text="<버퍼설정>", font=self.mid_font, fg='blue').place(x=410, y=5)
        tkinter.Label(custom_window, text="노증폭/극마부/극찬작 기준 스탯에서\n자신의 스탯이 얼마나 가감되는지 기입", fg="Red").place(x=350, y=33)
        tkinter.Label(custom_window, text="1각스탯+          ", font=self.guide_font).place(x=320, y=80)  ##
        c_stat = tkinter.Entry(custom_window, width=7)
        c_stat.place(x=390, y=82)
        c_stat.insert(END, db_preset['H1'].value)
        tkinter.Label(custom_window, text="축복스탯+          ", font=self.guide_font).place(x=470, y=80)  ##
        b_stat = tkinter.Entry(custom_window, width=7)
        b_stat.place(x=540, y=82)
        b_stat.insert(END, db_preset['H6'].value)
        three = [0, 1, 2, 3]
        two = [0, 1, 2]
        tkinter.Label(custom_window, text="축복칭호=", font=self.guide_font).place(x=320, y=110)
        b_style_lvl = tkinter.ttk.Combobox(custom_window, width=5, values=three)
        b_style_lvl.place(x=390, y=112)  ##
        b_style_lvl.set(db_preset['H2'].value)
        tkinter.Label(custom_window, text="1각칭호=", font=self.guide_font).place(x=470, y=110)
        c_style_lvl = tkinter.ttk.Combobox(custom_window, width=5, values=two)
        c_style_lvl.place(x=540, y=112)  ##
        c_style_lvl.set(db_preset['H3'].value)
        tkinter.Label(custom_window, text="축복플티=", font=self.guide_font).place(x=320, y=140)
        b_plt = tkinter.ttk.Combobox(custom_window, width=5, values=two)
        b_plt.place(x=390, y=142)  ##
        b_plt.set(db_preset['H4'].value)
        tkinter.Label(custom_window, text="축복클쳐=", font=self.guide_font).place(x=470, y=140)
        b_cri = tkinter.ttk.Combobox(custom_window, width=5, values=[0, 1])
        b_cri.place(x=540, y=142)  ##
        b_cri.set(db_preset['H5'].value)
        aria_value = ['항상증폭', '템에따라', '항상미증폭']
        # tkinter.Label(custom_window, text="아리아/퍼펫 증폭율 여부=", font=self.guide_font).place(x=320, y=170)
        aria_up = tkinter.ttk.Combobox(custom_window, values=aria_value, width=10)
        # aria_up.place(x=540 - 55, y=170)
        aria_up.set(db_preset['H7'].value)

        tkinter.Label(custom_window, text="<딜러속강>", font=self.mid_font).place(x=410, y=175 + 70)
        tkinter.Label(custom_window, text="주속성=", font=self.guide_font).place(x=320, y=210 + 70)
        ele_list = ['화', '수', '명', '암']
        ele_type = tkinter.ttk.Combobox(custom_window, width=5, values=ele_list)
        ele_type.place(x=390, y=212 + 70)  ##
        ele_type.set(db_preset['B1'].value)
        tkinter.Label(custom_window, text="마부총합=", font=self.guide_font).place(x=470, y=210 + 70)
        ele1 = tkinter.Entry(custom_window, width=7)
        ele1.place(x=540, y=212 + 70)  ##
        ele1.insert(END, db_preset['B14'].value)
        tkinter.Label(custom_window, text="오라속강=", font=self.guide_font).place(x=470, y=240 + 70)
        ele2 = tkinter.Entry(custom_window, width=7)
        ele2.place(x=540, y=242 + 70)  ##
        ele2.insert(END, db_preset['B15'].value)
        tkinter.Label(custom_window, text=" 젬 속강=", font=self.guide_font).place(x=470, y=270 + 70)
        ele3 = tkinter.Entry(custom_window, width=7)
        ele3.place(x=540, y=272 + 70)  ##
        ele3.insert(END, db_preset['B16'].value)
        # tkinter.Label(custom_window,text="스킬속강= 자동",font=self.guide_font).place(x=320,y=210+70)
        ele4 = tkinter.Entry(custom_window, width=7)  ##ele4.place(x=390,y=212) ## 자속강 비활성화
        ele4.insert(END, db_preset['B17'].value)
        tkinter.Label(custom_window, text=" 몹 속저=", font=self.guide_font).place(x=320, y=240 + 70)
        ele5 = tkinter.Entry(custom_window, width=7)
        ele5.place(x=390, y=242 + 70)  ##
        ele5.insert(END, db_preset['B18'].value)
        tkinter.Label(custom_window, text="버퍼속깎=", font=self.guide_font).place(x=320, y=270 + 70)
        ele6 = tkinter.Entry(custom_window, width=7)
        ele6.place(x=390, y=272 + 70)  ##
        ele6.insert(END, db_preset['B19'].value)
        tkinter.Label(custom_window, font=self.guide_font, fg="red",
                      text="반드시 메인창 SAVE 버튼을 눌러야 세이브 슬롯에 저장됩니다").place(x=143, y=370)

        load_preset.close()
        save_command = lambda: self.save_custom(ele_type.get(), cool_con.get(), cus1.get(), cus2.get(), cus3.get(),
                                           cus4.get(),
                                           cus6.get(), cus7.get(), cus8.get(), cus9.get(), cus10.get(), cus11.get(),
                                           cus12.get(),
                                           c_stat.get(), b_stat.get(), b_style_lvl.get(), c_style_lvl.get(),
                                           b_plt.get(), b_cri.get(),
                                           ele1.get(), ele2.get(), ele3.get(), ele4.get(), ele5.get(), ele6.get(),
                                           aria_up.get(), cool_con2.get())
        tkinter.Button(custom_window, text="저장하기", font=self.mid_font, command=save_command,
                       bg="lightyellow").place(x=190, y=295)
        if auto == 1:
            self.auto_saved = 1
            self.save_custom(ele_type.get(), cool_con.get(), cus1.get(), cus2.get(), cus3.get(), cus4.get(),
                        cus6.get(), cus7.get(), cus8.get(), cus9.get(), cus10.get(), cus11.get(), cus12.get(),
                        c_stat.get(), b_stat.get(), b_style_lvl.get(), c_style_lvl.get(), b_plt.get(), b_cri.get(),
                        ele1.get(), ele2.get(), ele3.get(), ele4.get(), ele5.get(), ele6.get(), aria_up.get(),
                        cool_con2.get())
            print('자동저장')
            self.auto_saved = 0

    # 통합 커스텀 저장
    def save_custom(self, ele_type, cool_con, cus1, cus2, cus3, cus4, cus6, cus7, cus8, cus9, cus10, cus11, cus12, c_stat,
                    b_stat, b_style_lvl, c_style_lvl, b_plt, b_cri, ele1, ele2, ele3, ele4, ele5, ele6, aria_up,
                    cool_con2):
        try:
            load_excel3 = load_workbook("DATA.xlsx")
            load_preset1 = load_workbook("preset.xlsx")
            db_custom1 = load_preset1["custom"]
            db_save_one = load_excel3["one"]
            db_save_set = load_excel3["set"]

            db_custom1['B1'] = ele_type
            if ele_type == '화':
                db_save_one['L181'] = 0
                db_save_one['L165'] = 0
                db_save_one['L149'] = 24
                db_save_one['L129'] = 0
                db_save_one['L429'] = 0
                db_save_one['L430'] = 0
                db_save_one['L431'] = 20
                db_save_one['L433'] = 0
            elif ele_type == '수':
                db_save_one['L181'] = 0
                db_save_one['L165'] = 24
                db_save_one['L149'] = 0
                db_save_one['L129'] = 0
                db_save_one['L429'] = 20
                db_save_one['L430'] = 0
                db_save_one['L431'] = 0
                db_save_one['L433'] = 0
            elif ele_type == '명':
                db_save_one['L181'] = 24
                db_save_one['L165'] = 0
                db_save_one['L149'] = 0
                db_save_one['L129'] = 0
                db_save_one['L429'] = 0
                db_save_one['L430'] = 20
                db_save_one['L431'] = 0
                db_save_one['L433'] = 0
            elif ele_type == '암':
                db_save_one['L181'] = 0
                db_save_one['L165'] = 0
                db_save_one['L149'] = 0
                db_save_one['L129'] = 24
                db_save_one['L429'] = 0
                db_save_one['L430'] = 0
                db_save_one['L431'] = 0
                db_save_one['L433'] = 20

            db_custom1['B3'] = float(cus1)
            db_save_one['O164'] = float(cus1)
            db_custom1['B4'] = float(cus2)
            db_save_one['O180'] = float(cus2)
            db_custom1['B5'] = float(cus6)
            db_save_one['O100'] = float(cus6)
            db_save_one['O101'] = float(cus6)
            db_custom1['B6'] = float(cus7)
            db_save_one['O127'] = float(cus7)
            db_custom1['B7'] = float(cus8)
            db_save_one['O147'] = float(cus8)
            db_custom1['B8'] = float(cus9)
            db_save_one['O163'] = float(cus9)
            db_custom1['B9'] = float(cus10)
            db_save_one['O179'] = float(cus10)
            db_custom1['B10'] = float(cus11)
            db_save_one['O295'] = float(cus11)
            db_custom1['B11'] = float(cus12)
            db_save_one['O296'] = float(cus12)
            db_save_one['O297'] = float(cus12)
            db_custom1['B12'] = cus3
            db_custom1['B2'] = cool_con
            db_custom1['B20'] = cool_con2
            if cus3 == '전설↓':
                db_save_one['J86'] = 34
                db_save_one['F120'] = 34
                db_save_one['N140'] = 34
                db_save_one['L156'] = 68
                db_save_one['K172'] = 34
                db_save_one['G276'] = 40
            else:
                db_save_one['J86'] = 35
                db_save_one['F120'] = 35
                db_save_one['N140'] = 35
                db_save_one['L156'] = 72
                db_save_one['K172'] = 35
                db_save_one['G276'] = 41
            db_custom1['B13'] = cus4
            db_save_one['N189'] = int(cus4) + 4
            db_save_one['N190'] = int(cus4) + 4
            db_save_one['K205'] = int(cus4) + 4
            db_save_one['E214'] = int(cus4) + 4

            db_custom1['H1'] = c_stat
            db_custom1['H6'] = b_stat
            db_custom1['H2'] = b_style_lvl
            db_custom1['H3'] = c_style_lvl
            db_custom1['H4'] = b_plt
            db_custom1['H5'] = b_cri
            db_custom1['H7'] = aria_up

            db_custom1['B14'] = ele1
            db_custom1['B15'] = ele2
            db_custom1['B16'] = ele3
            db_custom1['B17'] = ele4
            db_custom1['B18'] = ele5
            db_custom1['B19'] = ele6

            load_preset1.save("preset.xlsx")
            load_preset1.close()
            load_excel3.save("DATA.xlsx")
            load_excel3.close()
            self.custom_window.destroy()
            if self.auto_saved != 1:
                tkinter.messagebox.showinfo("알림", "저장 완료")
        except PermissionError as error:
            tkinter.messagebox.showerror("에러", "엑셀을 닫고 다시 시도해주세요.")

    # 지혜의 산물 선택창
    def create_knowledge_window(self):
        image_list_set = self.image_list_set
        equip_buttons = self.equip_buttons

        try:
            self.know_window.destroy()
        except:
            pass

        self.know_window = tkinter.Toplevel(self.window)
        know_window = self.know_window
        know_window.attributes("-topmost", True)
        know_window.geometry("545x405+750+20")
        know_window.resizable(False, False)
        know_window.configure(bg=self.dark_main)

        for code in self.get_all_knowledge_equipment_list():
            if self.owned_equipments[f"tg{code}"] == 0:
                know_button_image = self.image_list2[code]
            else:
                know_button_image = self.image_list[code]

            equip_buttons[code] = self.create_equip_button(know_window, know_button_image,
                                                           partial(self.click_equipment, code))

        equip_buttons["13390150"].place(x=303 - 290, y=20)
        equip_buttons["22390240"].place(x=333 - 290, y=20)
        equip_buttons["23390450"].place(x=363 - 290, y=20)
        equip_buttons["33390750"].place(x=393 - 290, y=20)
        equip_buttons["21390340"].place(x=424 - 290, y=20)
        equip_buttons["31390540"].place(x=454 - 290, y=20)
        equip_buttons["32390650"].place(x=484 - 290, y=20)
        equip_buttons["11390850"].place(x=484 - 290 + 80, y=20)
        equip_buttons["12390950"].place(x=484 - 290 + 110, y=20)
        equip_buttons["13391050"].place(x=484 - 290 + 140, y=20)
        equip_buttons["14391150"].place(x=484 - 290 + 170, y=20)
        equip_buttons["15391250"].place(x=484 - 290 + 200, y=20)

        tkinter.Label(know_window, bg=self.dark_main, image=image_list_set['201']).place(x=303 - 290, y=70)
        equip_buttons["22400150"].place(x=303 - 290 + 63, y=70)
        tkinter.Label(know_window, bg=self.dark_main, image=image_list_set['202']).place(x=303 - 290, y=70 + 40)
        equip_buttons["22400250"].place(x=303 - 290 + 63, y=70 + 40)
        tkinter.Label(know_window, bg=self.dark_main, image=image_list_set['203']).place(x=303 - 290, y=70 + 80)
        equip_buttons["22400350"].place(x=303 - 290 + 63, y=70 + 80)
        tkinter.Label(know_window, bg=self.dark_main, image=image_list_set['204']).place(x=303 - 290, y=70 + 120)
        equip_buttons["22400450"].place(x=303 - 290 + 63, y=70 + 120)
        tkinter.Label(know_window, bg=self.dark_main, image=image_list_set['205']).place(x=303 - 290, y=70 + 160)
        equip_buttons["22400550"].place(x=303 - 290 + 63, y=70 + 160)
        tkinter.Label(know_window, bg=self.dark_main, image=image_list_set['206']).place(x=303 - 290, y=70 + 200)
        equip_buttons["21400640"].place(x=303 - 290 + 63, y=70 + 200)
        tkinter.Label(know_window, bg=self.dark_main, image=image_list_set['207']).place(x=303 - 290, y=70 + 240)
        equip_buttons["31400750"].place(x=303 - 290 + 63, y=70 + 240)
        tkinter.Label(know_window, bg=self.dark_main, image=image_list_set['208']).place(x=120, y=70)
        equip_buttons["31400850"].place(x=120 + 63, y=70)
        tkinter.Label(know_window, bg=self.dark_main, image=image_list_set['209']).place(x=120, y=70 + 40)
        equip_buttons["31400950"].place(x=120 + 63, y=70 + 40)
        tkinter.Label(know_window, bg=self.dark_main, image=image_list_set['210']).place(x=120, y=70 + 80)
        equip_buttons["31401050"].place(x=120 + 63, y=70 + 80)
        tkinter.Label(know_window, bg=self.dark_main, image=image_list_set['211']).place(x=120, y=70 + 120)
        equip_buttons["31401150"].place(x=120 + 63, y=70 + 120)
        tkinter.Label(know_window, bg=self.dark_main, image=image_list_set['212']).place(x=120, y=70 + 160)
        equip_buttons["32401240"].place(x=120 + 63, y=70 + 160)
        tkinter.Label(know_window, bg=self.dark_main, image=image_list_set['213']).place(x=120, y=70 + 200)
        equip_buttons["32401340"].place(x=120 + 63, y=70 + 200)
        tkinter.Label(know_window, bg=self.dark_main, image=image_list_set['214']).place(x=120, y=70 + 240)
        equip_buttons["32401440"].place(x=120 + 63, y=70 + 240)

        tkinter.Label(know_window, bg=self.dark_main, image=image_list_set['215']).place(x=250, y=69 + 200)
        equip_buttons["11410100"].place(x=280 + 45, y=70 + 200)
        equip_buttons["11410110"].place(x=280 + 75, y=70 + 200)
        equip_buttons["11410120"].place(x=280 + 105, y=70 + 200)
        equip_buttons["11410130"].place(x=280 + 135, y=70 + 200)
        equip_buttons["11410140"].place(x=280 + 165, y=70 + 200)
        equip_buttons["11410150"].place(x=280 + 195, y=70 + 200)

        tkinter.Label(know_window, bg=self.dark_main, image=image_list_set['216']).place(x=250, y=69 + 240)
        equip_buttons["21420100"].place(x=280 + 45, y=70 + 240)
        equip_buttons["21420110"].place(x=280 + 75, y=70 + 240)
        equip_buttons["21420120"].place(x=280 + 105, y=70 + 240)
        equip_buttons["21420130"].place(x=280 + 135, y=70 + 240)
        equip_buttons["21420140"].place(x=280 + 165, y=70 + 240)
        equip_buttons["21420150"].place(x=280 + 195, y=70 + 240)

        tkinter.Label(know_window, bg=self.dark_main, image=image_list_set['217']).place(x=250, y=69 + 280)
        equip_buttons["33430100"].place(x=280 + 45, y=70 + 280)
        equip_buttons["33430110"].place(x=280 + 75, y=70 + 280)
        equip_buttons["33430120"].place(x=280 + 105, y=70 + 280)
        equip_buttons["33430130"].place(x=280 + 135, y=70 + 280)
        equip_buttons["33430140"].place(x=280 + 165, y=70 + 280)
        equip_buttons["33430150"].place(x=280 + 195, y=70 + 280)

        tkinter.Label(know_window, bg=self.dark_main, fg='white',
                      text=("세트 산물은 증/크증이 겹치는 경우가 있습니다.\n칭호/크리쳐 선택에 유의하세요.\n(중복안되게 계산식 처리 해놨음)\n\n"
                            + "불마/엘드 셋의 '마딜 전용'옵션은\n따로 구분되어 계산되지 않습니다.\n알아서 빼주세요.\n\n"
                            + "스탯 옵션은 버프+가호 받은 기준입니다.\n가호 미적용 스탯이 많은 산물 특성상,\n수련방 솔플 효율과 굉장히 다를수 있습니다.")).place(x=250,
                                                                                                                y=70)

    def create_equip_button(self, window: Wm, image: PhotoImage, callback: Callable):
        return tkinter.Button(window, relief="flat", borderwidth=0,activebackground=self.dark_main, bg=self.dark_main,
                              image=image, command=callback)

    # preset 리스트 이름 변경
    def change_list_name(self):
        try:
            self.change_window.destroy()
        except:
            pass

        self.change_window = tkinter.Toplevel(self.window)

        change_window = self.change_window
        change_window.geometry("390x320+750+200")
        tkinter.Label(change_window, text="1번슬롯").place(x=20, y=10)
        tkinter.Label(change_window, text="2번슬롯").place(x=20, y=35)
        tkinter.Label(change_window, text="3번슬롯").place(x=20, y=60)
        tkinter.Label(change_window, text="4번슬롯").place(x=20, y=85)
        tkinter.Label(change_window, text="5번슬롯").place(x=20, y=110)
        tkinter.Label(change_window, text="6번슬롯").place(x=20, y=135)
        tkinter.Label(change_window, text="7번슬롯").place(x=20, y=160)
        tkinter.Label(change_window, text="8번슬롯").place(x=20, y=185)
        tkinter.Label(change_window, text="9번슬롯").place(x=20, y=210)
        tkinter.Label(change_window, text="10번슬롯").place(x=20, y=235)
        tkinter.Label(change_window, text="11번슬롯").place(x=220, y=10)
        tkinter.Label(change_window, text="12번슬롯").place(x=220, y=35)
        tkinter.Label(change_window, text="13번슬롯").place(x=220, y=60)
        tkinter.Label(change_window, text="14번슬롯").place(x=220, y=85)
        tkinter.Label(change_window, text="15번슬롯").place(x=220, y=110)
        tkinter.Label(change_window, text="16번슬롯").place(x=220, y=135)
        tkinter.Label(change_window, text="17번슬롯").place(x=220, y=160)
        tkinter.Label(change_window, text="18번슬롯").place(x=220, y=185)
        tkinter.Label(change_window, text="19번슬롯").place(x=220, y=210)
        tkinter.Label(change_window, text="20번슬롯").place(x=220, y=235)
        entry1 = tkinter.Entry(change_window, width=10)
        entry1.place(x=95, y=12)
        entry1.insert(END, self.save_name_list[0])
        entry2 = tkinter.Entry(change_window, width=10)
        entry2.place(x=95, y=37)
        entry2.insert(END, self.save_name_list[1])
        entry3 = tkinter.Entry(change_window, width=10)
        entry3.place(x=95, y=62)
        entry3.insert(END, self.save_name_list[2])
        entry4 = tkinter.Entry(change_window, width=10)
        entry4.place(x=95, y=87)
        entry4.insert(END, self.save_name_list[3])
        entry5 = tkinter.Entry(change_window, width=10)
        entry5.place(x=95, y=112)
        entry5.insert(END, self.save_name_list[4])
        entry6 = tkinter.Entry(change_window, width=10)
        entry6.place(x=95, y=137)
        entry6.insert(END, self.save_name_list[5])
        entry7 = tkinter.Entry(change_window, width=10)
        entry7.place(x=95, y=162)
        entry7.insert(END, self.save_name_list[6])
        entry8 = tkinter.Entry(change_window, width=10)
        entry8.place(x=95, y=187)
        entry8.insert(END, self.save_name_list[7])
        entry9 = tkinter.Entry(change_window, width=10)
        entry9.place(x=95, y=212)
        entry9.insert(END, self.save_name_list[8])
        entry10 = tkinter.Entry(change_window, width=10)
        entry10.place(x=95, y=237)
        entry10.insert(END, self.save_name_list[9])

        entry11 = tkinter.Entry(change_window, width=10)
        entry11.place(x=295, y=12)
        entry11.insert(END, self.save_name_list[10])
        entry12 = tkinter.Entry(change_window, width=10)
        entry12.place(x=295, y=37)
        entry12.insert(END, self.save_name_list[11])
        entry13 = tkinter.Entry(change_window, width=10)
        entry13.place(x=295, y=62)
        entry13.insert(END, self.save_name_list[12])
        entry14 = tkinter.Entry(change_window, width=10)
        entry14.place(x=295, y=87)
        entry14.insert(END, self.save_name_list[13])
        entry15 = tkinter.Entry(change_window, width=10)
        entry15.place(x=295, y=112)
        entry15.insert(END, self.save_name_list[14])
        entry16 = tkinter.Entry(change_window, width=10)
        entry16.place(x=295, y=137)
        entry16.insert(END, self.save_name_list[15])
        entry17 = tkinter.Entry(change_window, width=10)
        entry17.place(x=295, y=162)
        entry17.insert(END, self.save_name_list[16])
        entry18 = tkinter.Entry(change_window, width=10)
        entry18.place(x=295, y=187)
        entry18.insert(END, self.save_name_list[17])
        entry19 = tkinter.Entry(change_window, width=10)
        entry19.place(x=295, y=212)
        entry19.insert(END, self.save_name_list[18])
        entry20 = tkinter.Entry(change_window, width=10)
        entry20.place(x=295, y=237)
        entry20.insert(END, self.save_name_list[19])

        tkinter.Button(change_window, text="저장", font=self.mid_font, command=lambda: self.change_savelist(
            [entry1.get(), entry2.get(), entry3.get(), entry4.get(), entry5.get(),
             entry6.get(), entry7.get(), entry8.get(), entry9.get(), entry10.get(),
             entry11.get(), entry12.get(), entry13.get(), entry14.get(), entry15.get(),
             entry16.get(), entry17.get(), entry18.get(), entry19.get(), entry20.get()])).place(x=170, y=270)

    def change_savelist(self, changed_savelist_name: List[str]):
        try:
            load_preset5 = load_workbook("preset.xlsx", data_only=True)
            db_custom2 = load_preset5["custom"]

            for i in range(1, 21):
                db_custom2.cell(i, 5).value = changed_savelist_name[i - 1]

            self.save_name_list.clear()
            self.save_name_list.extend(changed_savelist_name)
            load_preset5.save("preset.xlsx")
            load_preset5.close()
            self.save_select.set(self.save_name_list[0])
            self.save_select['values'] = self.save_name_list
            self.change_window.destroy()
            tkinter.messagebox.showinfo("알림", "저장 완료")
        except PermissionError as error:
            tkinter.messagebox.showerror("에러", "엑셀을 닫고 다시 시도해주세요.")

    # 타임라인 조회 창
    def timeline_select(self):
        try:
            self.timeline_window.destroy()
        except:
            pass

        self.timeline_window = tkinter.Toplevel(self.window)
        timeline_window = self.timeline_window
        timeline_window.attributes("-topmost", True)
        timeline_window.geometry("310x150+750+20")
        tkinter.Label(timeline_window, text="캐릭터명=\n(정확히)", font=self.guide_font).place(x=10, y=9)
        cha_name = tkinter.Entry(timeline_window, width=13)
        cha_name.place(x=80, y=12)
        tkinter.Label(timeline_window, text="서버명=", font=self.guide_font).place(x=10, y=59)

        serv_name = tkinter.ttk.Combobox(timeline_window, values=self.server_list, width=11)
        serv_name.place(x=80, y=62)
        serv_name.set('카인')
        load_timeline = tkinter.Button(timeline_window, command=lambda: self.show_timeline(cha_name.get(), serv_name.get()),
                                       text="불러오기", font=self.mid_font)
        load_timeline.place(x=200, y=25)
        tkinter.Label(timeline_window, text="타임라인에 있는 에픽만 불러옵니다(일부X)", fg="Red").place(x=10, y=100)
        tkinter.Label(timeline_window, text="서버 불안정때매 안되면 여러번 눌러보세요", fg="Red").place(x=10, y=120)

    # 타임라인 조회
    def show_timeline(self, name, server):
        apikey = self.get_api_key()

        server_dict = {'안톤': 'anton', '바칼': 'bakal', '카인': 'cain', '카시야스': 'casillas',
                       '디레지에': 'diregie', '힐더': 'hilder', '프레이': 'prey', '시로코': 'siroco'}
        try:
            sever_code = server_dict[server]
            cha_id_api = urllib.request.urlopen(
                'https://api.neople.co.kr/df/servers/' + sever_code + '/characters?characterName=' + parse.quote(
                    name) + '&apikey=' + apikey)
            cha_id_dic = loads(cha_id_api.read().decode("utf-8"))
            cha_id = cha_id_dic['rows'][0]['characterId']

            ##
            print(sever_code)
            print(cha_id)
            time.sleep(0.3)
            start_time = '20200101T0000'
            time_now = time.strftime('%Y%m%dT%H%M', time.localtime(time.time()))
            now = time_now
            now_1 = '20200101T0000'
            now_2 = '20200101T0000'
            now_3 = '20200101T0000'
            now2 = '20200101T0000'
            now3 = '20200101T0000'
            now4 = '20200101T0000'  ## 현재 11월 13일까지 조회 가능
            if int(time_now[0:8]) >= 20200316:
                now = '20200315T2359'
                now_1 = '20200316T0000'
                now2 = time.strftime('%Y%m%dT%H%M', time.localtime(time.time()))
            if int(time_now[0:8]) >= 20200601:
                now2 = '20200531T2359'
                now_2 = '20200601T0000'
                now3 = time.strftime('%Y%m%dT%H%M', time.localtime(time.time()))
            if int(time_now[0:8]) >= 20200816:
                now3 = '20200815T2359'
                now_3 = '20200816T0000'
                now4 = time.strftime('%Y%m%dT%H%M', time.localtime(time.time()))
            time_code = '504,505,506,507,508,510,511,512,513,514'
            timeline_list = []
            for nows in [[now, start_time], [now2, now_1], [now3, now_2], [now4, now_3]]:
                if nows[0] != '20200101T0000':
                    timeline = urllib.request.urlopen(
                        'https://api.neople.co.kr/df/servers/' + sever_code + '/characters/' + cha_id + '/timeline?limit=100&code=' + time_code + '&startDate=' +
                        nows[1] + '&endDate=' + nows[0] + '&apikey=' + apikey)
                    timeline2 = loads(timeline.read().decode("utf-8"))['timeline']
                    show_next = timeline2['next']
                    timeline_list = timeline_list + timeline2['rows']
                    time.sleep(0.3)
                    while show_next != None:
                        timeline_next = urllib.request.urlopen(
                            'https://api.neople.co.kr/df/servers/' + sever_code + '/characters/' + cha_id + '/timeline?next=' + show_next + '&apikey=' + apikey)
                        timeline_next2 = loads(timeline_next.read().decode("utf-8"))['timeline']
                        timeline_list = timeline_list + timeline_next2['rows']
                        time.sleep(0.3)
                        show_next = timeline_next2['next']

            all_item = []
            for now in timeline_list:
                item = now['data']['itemId']
                all_item.append(item)
            xl = openpyxl.load_workbook("DATA.xlsx", data_only=True)
            sh = xl['one']

            self.reset_equipments()

            for i in range(76, 257):
                api_cod = sh.cell(i, 40).value
                if all_item.count(api_cod) != 0:
                    self.owned_equipments['tg{}'.format(str(sh.cell(i, 1).value))] = 1
            xl.close()
            self.check_equipment()
            for i in range(101, 136):
                self.check_set(i)
            self.timeline_window.destroy()
            tkinter.messagebox.showinfo("주의", "과거 메타몽했던 에픽도 전부 불러와집니다.\n" +
                                        "알아서 빼주세요.\n\n초월한 에픽은 뜨지않습니다.\n알아서 넣으세요.\n\n" +
                                        "현재 무기와 시로코 에픽은 불러오지 않습니다")
        except urllib.error.HTTPError as error:
            tkinter.messagebox.showerror("에러", "API 접근 실패(네트워크 오류)")

    ########## 버전 최초 구동 프리셋 업데이트 ###########
    def create_update_window(self):
        def donotshow():
            load_show = load_workbook("preset.xlsx", data_only=True)
            load_show["custom"]['K2'] = 1
            load_show.save("preset.xlsx")
            load_show.close()
            update_window.destroy()

        update_window = tkinter.Toplevel(self.window)
        update_window.attributes("-topmost", True)
        update_window.geometry("300x320+0+0")
        update_window.resizable(False, False)
        self.place_center(update_window, 0)
        update_window.configure(bg=self.dark_main)

        def _on_mousewheel(event):  # 마우스 휠 스크롤링
            update_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        scrollbar = Scrollbar(update_window)
        scrollbar.pack(side=RIGHT, fill=Y)
        update_text = ''
        try:
            update_file = open(self.now_version + " 패치노트.txt", 'r', encoding='UTF8')
            lines = update_file.readlines()
            for line in lines:
                update_text = update_text + '\n' + line
            update_file.close()
        except:
            update_text = '업데이트 텍스트 파일 누락'

        update_canvas = tkinter.Canvas(update_window, width=276, height=250, bg=self.dark_main, bd=0)
        update_canvas.pack()
        update_canvas.create_text(5, 0, text=update_text, fill='white', font=self.guide_font, anchor='nw',
                                  width=280)
        update_canvas.bind("<MouseWheel>", _on_mousewheel)
        tkinter.Label(update_window, bg=self.dark_main, font=self.guide_font, fg='red',
                      text="2.X.X 버전이랑 계수 호환 X", anchor='w').pack()
        tkinter.Button(update_window, font=self.small_font, command=donotshow, text="업데이트 전까지 보지않기").pack()

        scrollbar.config(command=update_canvas.yview)

    def init_images(self):
        file_list = os.listdir("image")
        file_list_wep = os.listdir("image_wep")

        for i in file_list:
            if i[-3:] != 'gif':
                if i[-5] == 'n':
                    self.image_list[i[:-5]] = self.get_photo_image(f"image/{i}")
                elif i[-5] == 'f':
                    self.image_list2[i[:-5]] = self.get_photo_image(f"image/{i}")
                elif i[-5] == 't':
                    self.image_list_tag[i[:-5]] = self.get_photo_image(f"image/{i}")

                if i[0] == 'n' and i[-5] == 't':
                    self.image_list_tag[i[:-5]] = self.get_photo_image(f"image/{i}")
                elif i[0] == 'n':
                    self.image_list[i[:-4]] = self.get_photo_image(f"image/{i}")

        for i in range(1, 56):
            try:
                self.image_list_set[str(100 + i)] = self.get_photo_image(f"set_name/{i + 100}.png")
                self.image_list_set2[str(100 + i)] = self.get_photo_image(f"set_name/{i + 100}f.png")
            except:
                pass

        for i in range(1, 18):
            self.image_list_set[str(200 + i)] = self.get_photo_image(f"set_name/{i + 200}.png")

        self.image_list['99990'] = self.get_photo_image("image/99990.png")
        self.image_list2['99990'] = self.get_photo_image("image/99990.png")
        self.image_list_tag['99990'] = self.get_photo_image("image/99990.png")

        for i in file_list_wep:
            self.image_list_wep[calc_list_wep.wep_image_filename.get(i[:-4])] = self.get_photo_image(
                "image_wep/{}".format(i))

        self.default_img1n = self.get_photo_image("ext_img/default1n.png")
        self.default_img1f = self.get_photo_image("ext_img/default1f.png")
        self.default_img2n = self.get_photo_image("ext_img/default2n.png")
        self.default_img2f = self.get_photo_image("ext_img/default2f.png")
        self.default_img3n = self.get_photo_image("ext_img/default3n.png")
        self.default_img3f = self.get_photo_image("ext_img/default3f.png")

    def init_ui(self):
        self.case_count_label = tkinter.Label(self.window, font=self.guide_font, fg="white", bg=self.dark_sub)
        self.case_count_label.place(x=700, y=145 - 12)

        self.total_count_label = tkinter.Label(self.window, font=self.guide_font, fg="white", bg=self.dark_sub)
        self.total_count_label.place(x=430 + 259, y=480 - 287)

        self.calc_state_label = tkinter.Label(self.window, text="상태 표시 칸", font=self.guide_font, fg="white",
                                   bg=self.dark_sub)
        self.calc_state_label.place(x=700, y=125 - 12)

        self.weapon_list_num_label = tkinter.Label(self.window, font=self.guide_font, bg=self.dark_main, fg='white',
                                                   text="무기 수=" + "0 / 10", anchor='c')
        self.weapon_list_num_label.place(x=365, y=67)

        for i in range(0, 10):
            weapon_label = tkinter.Label(self.window, bg=self.dark_sub, bd=0)
            weapon_label.place(x=32 + 31 * i, y=65)

            self.selected_weapon_label_list.append(weapon_label)

        self.void_weapon_img = self.get_photo_image("ext_img/00000.png")

        # 장비, 세트 버튼
        equip_buttons = self.equip_buttons

        all_equip_list = self.normal_epic_list + self.siroco_equip_list

        for code in all_equip_list:
            equip_buttons[code] = self.create_equip_button(self.window, self.image_list2[code],
                                                           partial(self.click_equipment, code))

        # 일반 에픽 버튼 위치
        # 상의
        equip_buttons["11010"].place(x=100, y=100)
        equip_buttons["11011"].place(x=130, y=100)
        equip_buttons["11020"].place(x=100, y=130)
        equip_buttons["11021"].place(x=130, y=130)
        equip_buttons["11030"].place(x=100, y=160)
        equip_buttons["11031"].place(x=130, y=160)
        equip_buttons["11040"].place(x=100, y=190)
        equip_buttons["11041"].place(x=130, y=190)
        equip_buttons["11050"].place(x=100, y=220)
        equip_buttons["11051"].place(x=130, y=220)
        equip_buttons["11060"].place(x=100, y=250)
        equip_buttons["11061"].place(x=130, y=250)
        equip_buttons["11070"].place(x=100, y=280)
        equip_buttons["11071"].place(x=130, y=280)
        equip_buttons["11080"].place(x=100, y=310)
        equip_buttons["11081"].place(x=130, y=310)
        equip_buttons["11090"].place(x=100, y=340)
        equip_buttons["11091"].place(x=130, y=340)
        equip_buttons["11100"].place(x=100, y=370)
        equip_buttons["11101"].place(x=130, y=370)
        equip_buttons["11110"].place(x=100, y=400)
        equip_buttons["11111"].place(x=130, y=400)
        equip_buttons["11120"].place(x=100, y=430)
        equip_buttons["11121"].place(x=130, y=430)
        equip_buttons["11130"].place(x=100, y=460)
        equip_buttons["11131"].place(x=130, y=460)
        equip_buttons["11140"].place(x=100, y=490)
        equip_buttons["11141"].place(x=130, y=490)
        equip_buttons["11150"].place(x=100, y=520)
        equip_buttons["11151"].place(x=130, y=520)

        equip_buttons["11280"].place(x=100, y=570)
        equip_buttons["11281"].place(x=130, y=570)
        equip_buttons["11290"].place(x=100, y=600)
        equip_buttons["11291"].place(x=130, y=600)
        equip_buttons["11300"].place(x=100, y=630)
        equip_buttons["11301"].place(x=130, y=630)
        equip_buttons["11310"].place(x=100, y=660)
        equip_buttons["11311"].place(x=130, y=660)

        # 하의
        equip_buttons["12010"].place(x=161, y=100)
        equip_buttons["12020"].place(x=161, y=130)
        equip_buttons["12030"].place(x=161, y=160)
        equip_buttons["12040"].place(x=161, y=190)
        equip_buttons["12050"].place(x=161, y=220)
        equip_buttons["12060"].place(x=161, y=250)
        equip_buttons["12070"].place(x=161, y=280)
        equip_buttons["12080"].place(x=161, y=310)
        equip_buttons["12090"].place(x=161, y=340)
        equip_buttons["12100"].place(x=161, y=370)
        equip_buttons["12110"].place(x=161, y=400)
        equip_buttons["12120"].place(x=161, y=430)
        equip_buttons["12130"].place(x=161, y=460)
        equip_buttons["12140"].place(x=161, y=490)
        equip_buttons["12150"].place(x=161, y=520)
        equip_buttons["12240"].place(x=296, y=570)
        equip_buttons["12250"].place(x=296, y=600)
        equip_buttons["12260"].place(x=296, y=630)
        equip_buttons["12270"].place(x=296, y=660)

        # 어깨
        equip_buttons["13010"].place(x=192, y=100)
        equip_buttons["13020"].place(x=192, y=130)
        equip_buttons["13030"].place(x=192, y=160)
        equip_buttons["13040"].place(x=192, y=190)
        equip_buttons["13050"].place(x=192, y=220)
        equip_buttons["13060"].place(x=192, y=250)
        equip_buttons["13070"].place(x=192, y=280)
        equip_buttons["13080"].place(x=192, y=310)
        equip_buttons["13090"].place(x=192, y=340)
        equip_buttons["13100"].place(x=192, y=370)
        equip_buttons["13110"].place(x=192, y=400)
        equip_buttons["13120"].place(x=192, y=430)
        equip_buttons["13130"].place(x=192, y=460)
        equip_buttons["13140"].place(x=192, y=490)
        equip_buttons["13150"].place(x=192, y=520)

        # 벨트
        equip_buttons["14010"].place(x=223, y=100)
        equip_buttons["14020"].place(x=223, y=130)
        equip_buttons["14030"].place(x=223, y=160)
        equip_buttons["14040"].place(x=223, y=190)
        equip_buttons["14050"].place(x=223, y=220)
        equip_buttons["14060"].place(x=223, y=250)
        equip_buttons["14070"].place(x=223, y=280)
        equip_buttons["14080"].place(x=223, y=310)
        equip_buttons["14090"].place(x=223, y=340)
        equip_buttons["14100"].place(x=223, y=370)
        equip_buttons["14110"].place(x=223, y=400)
        equip_buttons["14120"].place(x=223, y=430)
        equip_buttons["14130"].place(x=223, y=460)
        equip_buttons["14140"].place(x=223, y=490)
        equip_buttons["14150"].place(x=223, y=520)

        # 신발
        equip_buttons["15010"].place(x=254, y=100)
        equip_buttons["15020"].place(x=254, y=130)
        equip_buttons["15030"].place(x=254, y=160)
        equip_buttons["15040"].place(x=254, y=190)
        equip_buttons["15050"].place(x=254, y=220)
        equip_buttons["15060"].place(x=254, y=250)
        equip_buttons["15070"].place(x=254, y=280)
        equip_buttons["15080"].place(x=254, y=310)
        equip_buttons["15090"].place(x=254, y=340)
        equip_buttons["15100"].place(x=254, y=370)
        equip_buttons["15110"].place(x=254, y=400)
        equip_buttons["15120"].place(x=254, y=430)
        equip_buttons["15130"].place(x=254, y=460)
        equip_buttons["15140"].place(x=254, y=490)
        equip_buttons["15150"].place(x=254, y=520)
        equip_buttons["15320"].place(x=492, y=570)
        equip_buttons["15330"].place(x=492, y=600)
        equip_buttons["15340"].place(x=492, y=630)
        equip_buttons["15350"].place(x=492, y=660)

        # 팔찌
        equip_buttons["21160"].place(x=370 - 12, y=100)
        equip_buttons["21161"].place(x=370 - 12 + 30, y=100)
        equip_buttons["21170"].place(x=370 - 12, y=130)
        equip_buttons["21171"].place(x=370 - 12 + 30, y=130)
        equip_buttons["21180"].place(x=370 - 12, y=160)
        equip_buttons["21181"].place(x=370 - 12 + 30, y=160)
        equip_buttons["21190"].place(x=370 - 12, y=190)
        equip_buttons["21191"].place(x=370 - 12 + 30, y=190)
        equip_buttons["21240"].place(x=327, y=570)
        equip_buttons["21241"].place(x=357, y=570)
        equip_buttons["21250"].place(x=327, y=600)
        equip_buttons["21251"].place(x=357, y=600)
        equip_buttons["21260"].place(x=327, y=630)
        equip_buttons["21261"].place(x=357, y=630)
        equip_buttons["21270"].place(x=327, y=660)
        equip_buttons["21271"].place(x=357, y=660)

        # 목걸이
        equip_buttons["22160"].place(x=419, y=100)
        equip_buttons["22170"].place(x=419, y=130)
        equip_buttons["22180"].place(x=419, y=160)
        equip_buttons["22190"].place(x=419, y=190)
        equip_buttons["22280"].place(x=161, y=570)
        equip_buttons["22290"].place(x=161, y=600)
        equip_buttons["22300"].place(x=161, y=630)
        equip_buttons["22310"].place(x=161, y=660)

        # 반지
        equip_buttons["23160"].place(x=450, y=100)
        equip_buttons["23170"].place(x=450, y=130)
        equip_buttons["23180"].place(x=450, y=160)
        equip_buttons["23190"].place(x=450, y=190)
        equip_buttons["23320"].place(x=523, y=570)
        equip_buttons["23330"].place(x=523, y=600)
        equip_buttons["23340"].place(x=523, y=630)
        equip_buttons["23350"].place(x=523, y=660)

        # 보조장비
        equip_buttons["31200"].place(x=554, y=100)
        equip_buttons["31210"].place(x=554, y=130)
        equip_buttons["31220"].place(x=554, y=160)
        equip_buttons["31230"].place(x=554, y=190)
        equip_buttons["31280"].place(x=192, y=570)
        equip_buttons["31290"].place(x=192, y=600)
        equip_buttons["31300"].place(x=192, y=630)
        equip_buttons["31310"].place(x=192, y=660)

        # 마법석
        equip_buttons["32200"].place(x=585, y=100)
        equip_buttons["32210"].place(x=585, y=130)
        equip_buttons["32220"].place(x=585, y=160)
        equip_buttons["32230"].place(x=585, y=190)
        equip_buttons["32240"].place(x=388, y=570)
        equip_buttons["32250"].place(x=388, y=600)
        equip_buttons["32260"].place(x=388, y=630)
        equip_buttons["32270"].place(x=388, y=660)

        # 귀걸이
        equip_buttons["33200"].place(x=616, y=100)
        equip_buttons["33201"].place(x=646, y=100)
        equip_buttons["33210"].place(x=616, y=130)
        equip_buttons["33211"].place(x=646, y=130)
        equip_buttons["33220"].place(x=616, y=160)
        equip_buttons["33221"].place(x=646, y=160)
        equip_buttons["33230"].place(x=616, y=190)
        equip_buttons["33231"].place(x=646, y=190)
        equip_buttons["33320"].place(x=554, y=570)
        equip_buttons["33321"].place(x=584, y=570)
        equip_buttons["33330"].place(x=554, y=600)
        equip_buttons["33331"].place(x=584, y=600)
        equip_buttons["33340"].place(x=554, y=630)
        equip_buttons["33341"].place(x=584, y=630)
        equip_buttons["33350"].place(x=554, y=660)
        equip_buttons["33351"].place(x=584, y=660)

        # 시로코 에픽 버튼 위치
        se = self.siroco_equip_list
        assert(len(se) == 15)

        equip_buttons[se[0]].place(x=710 + 10 + 71, y=445 + 95)
        equip_buttons[se[1]].place(x=710 + 10 + 71, y=445 + 30 + 95)
        equip_buttons[se[2]].place(x=710 + 10 + 71, y=445 + 60 + 95)
        equip_buttons[se[3]].place(x=710 + 10 + 71, y=445 + 90 + 95)
        equip_buttons[se[4]].place(x=710 + 10 + 71, y=445 + 120 + 95)
        equip_buttons[se[5]].place(x=710 + 10 + 71 + 31, y=445 + 95)
        equip_buttons[se[6]].place(x=710 + 10 + 71 + 31, y=445 + 30 + 95)
        equip_buttons[se[7]].place(x=710 + 10 + 71 + 31, y=445 + 60 + 95)
        equip_buttons[se[8]].place(x=710 + 10 + 71 + 31, y=445 + 90 + 95)
        equip_buttons[se[9]].place(x=710 + 10 + 71 + 31, y=445 + 120 + 95)
        equip_buttons[se[10]].place(x=710 + 10 + 71 + 62, y=445 + 95)
        equip_buttons[se[11]].place(x=710 + 10 + 71 + 62, y=445 + 30 + 95)
        equip_buttons[se[12]].place(x=710 + 10 + 71 + 62, y=445 + 60 + 95)
        equip_buttons[se[13]].place(x=710 + 10 + 71 + 62, y=445 + 90 + 95)
        equip_buttons[se[14]].place(x=710 + 10 + 71 + 62, y=445 + 120 + 95)

        # 세트 버튼
        set_buttons = self.set_buttons
        set_list: List[int] = []

        for i in range(101, 136):
            set_list.append(i)

        for i in range(151, 156):
            set_list.append(i)

        for i in set_list:
            str_i = str(i)
            set_buttons[str_i] = tkinter.Button(self.window, bg=self.dark_main, borderwidth=0,
                                                activebackground=self.dark_main, image=self.image_list_set2[str_i],
                                                command=partial(self.click_set, i))

        set_buttons["101"].place(x=29, y=100)
        set_buttons["102"].place(x=29, y=130)
        set_buttons["103"].place(x=29, y=160)
        set_buttons["104"].place(x=29, y=190)
        set_buttons["105"].place(x=29, y=220)
        set_buttons["106"].place(x=29, y=250)
        set_buttons["107"].place(x=29, y=280)
        set_buttons["108"].place(x=29, y=310)
        set_buttons["109"].place(x=29, y=340)
        set_buttons["110"].place(x=29, y=370)
        set_buttons["111"].place(x=29, y=400)
        set_buttons["112"].place(x=29, y=430)
        set_buttons["113"].place(x=29, y=460)
        set_buttons["114"].place(x=29, y=490)
        set_buttons["115"].place(x=29, y=520)
        set_buttons["116"].place(x=320 - 33, y=100)
        set_buttons["117"].place(x=320 - 33, y=130)
        set_buttons["118"].place(x=320 - 33, y=160)
        set_buttons["119"].place(x=320 - 33, y=190)
        set_buttons["120"].place(x=500 - 17, y=100)
        set_buttons["121"].place(x=500 - 17, y=130)
        set_buttons["122"].place(x=500 - 17, y=160)
        set_buttons["123"].place(x=500 - 17, y=190)
        set_buttons["124"].place(x=225, y=570)
        set_buttons["125"].place(x=225, y=600)
        set_buttons["126"].place(x=225, y=630)
        set_buttons["127"].place(x=225, y=660)
        set_buttons["128"].place(x=29, y=570)
        set_buttons["129"].place(x=29, y=600)
        set_buttons["130"].place(x=29, y=630)
        set_buttons["131"].place(x=29, y=660)
        set_buttons["132"].place(x=421, y=570)
        set_buttons["133"].place(x=421, y=600)
        set_buttons["134"].place(x=421, y=630)
        set_buttons["135"].place(x=421, y=660)

        set_buttons["151"].place(x=710 + 10, y=445 + 95)
        set_buttons["152"].place(x=710 + 10, y=475 + 95)
        set_buttons["153"].place(x=710 + 10, y=505 + 95)
        set_buttons["154"].place(x=710 + 10, y=535 + 95)
        set_buttons["155"].place(x=710 + 10, y=565 + 95)

        self.save_select = tkinter.ttk.Combobox(self.window, width=8, values=self.save_name_list)
        self.save_select.place(x=345 + 165, y=410 - 100)
        self.save_select.set(self.save_name_list[0])

        self.tg_groggy_img1 = self.get_photo_image("ext_img/groggy_swi1.png")
        self.tg_groggy_img2 = self.get_photo_image("ext_img/groggy_swi2.png")

        self.type1_img = self.get_photo_image('ext_img/type_bless.png')
        self.type2_img = self.get_photo_image('ext_img/type_crux.png')
        self.type3_img = self.get_photo_image('ext_img/type_all.png')
        self.show_detail_img = self.get_photo_image('ext_img/show_detail.png')
        self.show_tag_img = self.get_photo_image('ext_img/show_set_tag.png')
        self.capture_img = self.get_photo_image('ext_img/capture_img.png')
        self.style_compare_img = self.get_photo_image('ext_img/style_compare.png')

        self.result_upbox_img = self.get_photo_image('ext_img/bg_result_upbox.png')
        self.result_downbox_img = self.get_photo_image('ext_img/bg_result_downbox.png')
        self.result_sidebox_img = self.get_photo_image('ext_img/bg_result_sidebox.png')
        self.result_showbox_img = self.get_photo_image('ext_img/bg_result_showbox.png')

        self.result_gauge_bar_img = self.get_photo_image('ext_img/result_gauge_bar.png')
        self.result_gauge_img = self.get_photo_image('ext_img/result_gauge.png')
        self.result_checklist_img = self.get_photo_image('ext_img/result_show_checklist.png')

        def guide_speed():
            tkinter.messagebox.showinfo("정확도 선택",
                                        "매우빠름=세트옵션7개 풀적용 경우의 수만 계산. 중간세팅은 고려하지 않음\n빠름=단일 선택 부위를 전부 제거\n중간=단일은 포함하되, 신화에 우선권 부여\n느림=세트 수 우선권 완화, 신화 우선권 삭제")

        tkinter.Label(self.window, font=self.mid_font, fg="white", bg=self.dark_sub,
                      text="<딜러 프로필 생성기>").place(x=301, y=401)
        tkinter.Label(self.window, fg="white", bg=self.dark_sub, text="서버명=").place(x=296, y=433)
        tkinter.Label(self.window, fg="white", bg=self.dark_sub, text="캐릭명=").place(x=296, y=460)

        sever_in = tkinter.ttk.Combobox(self.window, width=9, values=self.server_list)
        sever_in.place(x=346, y=435)
        sever_in.set('카인')

        cha_Entry = tkinter.Entry(self.window, width=12)
        cha_Entry.place(x=346, y=462)
        sever_in.bind('<Return>', lambda e: self.show_profile(str(cha_Entry.get()), str(sever_in.get())))
        cha_Entry.bind('<Return>', lambda e: self.show_profile(str(cha_Entry.get()), str(sever_in.get())))

        generate_cha = self.get_photo_image("ext_img/generate_cha.png")
        tkinter.Button(self.window, image=generate_cha,
                       command=lambda: self.show_profile(str(cha_Entry.get()), str(sever_in.get())), borderwidth=0,
                       activebackground=self.dark_sub, bg=self.dark_sub).place(x=440, y=434)

        tkinter.Label(self.window, text='엔터로도 조회됩니다', font=self.guide_font, fg='white',
                      bg=self.dark_sub).place(x=332, y=482)
        cha_caution_text = """장비%:  12부위장비+칭호클쳐의
                   수준을 표현한 % (쿨감O)
                   (계산기 값과 유사)

        세팅%:  위를 제외한 나머지의
                   투자/세팅 실효율%
                   (노증극세팅이 100%)"""
        tkinter.Label(self.window, text=cha_caution_text, font=self.small_font, fg='white',
                      bg=self.dark_sub, anchor='nw', justify='left').place(x=512, y=405)

        select_perfect = self.select_perfect = tkinter.ttk.Combobox(self.window,
                                                                    values=['풀셋모드(매우빠름)',
                                                                            '메타몽풀셋모드',
                                                                            '단품제외(보통)',
                                                                            '단품포함(느림)',
                                                                            '세트필터↓(매우느림)'],
                                                                    width=15)
        select_perfect.place(x=145 + 605, y=11)
        select_perfect.set('단품포함(느림)')
        select_speed_img = self.get_photo_image("ext_img/select_speed.png")

        tkinter.Button(self.window, command=guide_speed, image=select_speed_img, borderwidth=0,
                       activebackground=self.dark_main, bg=self.dark_main).place(x=29 + 605, y=7)

        reset_img = self.get_photo_image("ext_img/reset.png")
        tkinter.Button(self.window, command=self.reset_equipments, image=reset_img, borderwidth=0,
                       activebackground=self.dark_main, bg=self.dark_main).place(x=302 + 180 + 17 + 135, y=476 - 435)

        wep_image = self.get_photo_image("ext_img/wep.png")
        tkinter.Label(self.window, image=wep_image, borderwidth=0, activebackground=self.dark_main,
                      bg=self.dark_main) \
            .place(x=29, y=10)

        wep_job_type = list(calc_list_wep.DNF_wep_list.keys())
        wep_job_select = self.wep_job_select = tkinter.ttk.Combobox(self.window, width=12,
                                                                                     values=wep_job_type)

        def wep_job_selected(event):
            wep_type_select["values"] = list(calc_list_wep.DNF_wep_list[str(wep_job_select.get())].keys())
            wep_type_select.set(list(calc_list_wep.DNF_wep_list[str(wep_job_select.get())].keys())[0])
            wep_select["values"] = list(calc_list_wep.DNF_wep_list[str(wep_job_select.get())][str(wep_type_select.get())])
            wep_select.set(list(calc_list_wep.DNF_wep_list[str(wep_job_select.get())][str(wep_type_select.get())])[0])

        def wep_job_selected2(event):
            wep_select["values"] = list(calc_list_wep.DNF_wep_list[str(wep_job_select.get())][str(wep_type_select.get())])
            wep_select.set(list(calc_list_wep.DNF_wep_list[str(wep_job_select.get())][str(wep_type_select.get())])[0])

        wep_job_select.place(x=110, y=10)
        wep_job_select.set('귀검/나이트')
        wep_job_select.bind("<<ComboboxSelected>>", wep_job_selected)

        wep_type = list(calc_list_wep.DNF_wep_list['귀검/나이트'].keys())
        wep_type_select = self.wep_type_select = tkinter.ttk.Combobox(self.window, width=12, values=wep_type)
        wep_type_select.place(x=236, y=10)
        wep_type_select.set('광검')
        wep_type_select.bind("<<ComboboxSelected>>", wep_job_selected2)

        wep_default = list(calc_list_wep.DNF_wep_list['귀검/나이트']['광검'])
        wep_select = self.wep_select = tkinter.ttk.Combobox(self.window, width=30, values=wep_default)
        wep_select.place(x=110, y=38)
        wep_select.set('(광검)별의 바다 : 바드나후')

        wep_select_image = self.get_photo_image("ext_img/wep_select.png")
        wep_reset_image = self.get_photo_image("ext_img/wep_reset.png")
        wep_select_bt = tkinter.Button(self.window, image=wep_select_image, fg="white", borderwidth=0,
                                       activebackground=self.dark_main, command=self.wep_list_select,
                                       bg=self.dark_main, font=self.mid_font)
        wep_select_bt.place(x=350, y=10)
        wep_select_bt = tkinter.Button(self.window, image=wep_reset_image, fg="white", borderwidth=0,
                                       activebackground=self.dark_main, command=self.wep_list_reset,
                                       bg=self.dark_main)
        wep_select_bt.place(x=440, y=10)

        def job_type_selected(event):
            jobup_select["values"] = list(calc_list_job.DNF_job_list[jobtype_select.get()])
            jobup_select.set(list(calc_list_job.DNF_job_list[jobtype_select.get()])[0])

        jobtype_select = self.jobtype_select = tkinter.ttk.Combobox(self.window, width=13,
                                                                    values=list(calc_list_job.DNF_job_list.keys()))
        jobtype_select.bind("<<ComboboxSelected>>", job_type_selected)
        jobtype_select.set('귀검사(남)')
        jobtype_select.place(x=390 - 17, y=190 + 52)

        jobup_select = self.jobup_select = tkinter.ttk.Combobox(self.window, width=13,
                                                                                       values=list(
                                                                                           calc_list_job.DNF_job_list[
                                                                                               '귀검사(남)']))
        jobup_select.set('검신(진각)')
        jobup_select.place(x=390 - 17, y=220 + 52)

        style_list = ['증뎀15%', '속강32', '증뎀10%', '추뎀10%', '크증10%', '기타(직접비교)']
        style_select = self.style_select = tkinter.ttk.Combobox(self.window, width=13,
                                                                                       values=style_list)
        style_select.set('증뎀15%')
        style_select.place(x=390 - 17, y=250 + 52)

        creature_list = ['증뎀10%', '크증10%', '모공15%', '크증18%', '물마독공18%', '기타(직접비교)']
        creature_select = self.creature_select = tkinter.ttk.Combobox(self.window,
                                                                                             width=13,
                                                                                             values=creature_list)
        creature_select.set('크증18%')
        creature_select.place(x=390 - 17, y=280 + 52)

        req_cool = self.req_cool = tkinter.ttk.Combobox(self.window, width=13,
                                                                               values=['X(지속딜만)', 'O(그로기포함)'])
        req_cool.set('X(지속딜만)')
        req_cool.place(x=390 - 17, y=310 + 52)

        calc_img = self.get_photo_image("ext_img/calc.png")
        tkinter.Button(self.window, image=calc_img, borderwidth=0, activebackground=self.dark_main,
                       command=self.calc_thread, bg=self.dark_main).place(x=390 - 35 + 150, y=7)

        stop_img = self.get_photo_image("ext_img/stop.png")
        tkinter.Button(self.window, image=stop_img, borderwidth=0, activebackground=self.dark_main,
                       command=self.stop_calc, bg=self.dark_main).place(x=390 - 35 + 150, y=62)

        timeline_img = self.get_photo_image("ext_img/timeline.png")
        tkinter.Button(self.window, image=timeline_img, borderwidth=0, activebackground=self.dark_main,
                       command=self.timeline_select, bg=self.dark_sub).place(x=345 + 165, y=340 - 100)

        custom_img = self.get_photo_image("ext_img/custom.png")
        tkinter.Button(self.window, image=custom_img, borderwidth=0, activebackground=self.dark_main,
                       command=lambda: self.create_custom_window(0), bg=self.dark_sub).place(x=435 + 165, y=340 - 100)

        save_img = self.get_photo_image("ext_img/SAVE.png")
        tkinter.Button(self.window, image=save_img, borderwidth=0, activebackground=self.dark_main,
                       command=self.save_checklist, bg=self.dark_sub).place(x=345 + 165, y=440 - 100)

        load_img = self.get_photo_image("ext_img/LOAD.png")
        tkinter.Button(self.window, image=load_img, borderwidth=0, activebackground=self.dark_main,
                       command=self.load_checklist, bg=self.dark_sub).place(x=435 + 165, y=440 - 100)

        change_name_img = self.get_photo_image("ext_img/name_change.png")
        tkinter.Button(self.window, image=change_name_img, borderwidth=0, activebackground=self.dark_main,
                       command=self.change_list_name, bg=self.dark_sub).place(x=435 + 165, y=405 - 100)

        inv_mod_list = ["미부여", "선택부여", "최적부여(버퍼X)"]
        inv_mod = self.inv_mod = tkinter.ttk.Combobox(self.window, width=10, values=inv_mod_list)
        inv_mod.place(x=785, y=285)
        inv_mod.set("미부여")
        inv_mod.bind("<<ComboboxSelected>>", self.update_inv)

        inv_type_list = self.inv_type_list
        inv_value_list1 = [6, 8, 10]
        inv_value_list2 = [3, 4, 5]

        inv_select1_1 = self.inv_select1_1 = tkinter.ttk.Combobox(self.window, width=4,
                                                                        values=inv_type_list)
        inv_select1_1.place(x=785, y=315)
        inv_select1_1.set("증뎀")
        inv_select1_2 = self.inv_select1_2 = tkinter.ttk.Combobox(self.window, width=3,
                                                                        values=inv_value_list1)
        inv_select1_2.place(x=842, y=315)
        inv_select1_2.set(10)
        inv_select2_1 = self.inv_select2_1 = tkinter.ttk.Combobox(self.window, width=4,
                                                                        values=inv_type_list)
        inv_select2_1.place(x=785, y=345)
        inv_select2_1.set("증뎀")
        inv_select2_2 = self.inv_select2_2 = tkinter.ttk.Combobox(self.window, width=3,
                                                                        values=inv_value_list2)
        inv_select2_2.place(x=842, y=345)
        inv_select2_2.set(5)

        inv_type_list2 = ["축스탯%/1각", "축스탯%/1각%", "축앞뎀%/1각", "축앞뎀%/1각%", "전직패", "축스탯%/1각+1"]
        inv_type_list2_1 = ["축스탯%/1각", "축스탯%/1각%", "축앞뎀%/1각", "축앞뎀%/1각%", "전직패", "축+1/1각"]
        inv_value_list3 = ['3%/60(상)', '3%/40(중)', '3%/20(하)']
        inv_value_list3_1 = ['3%/40(상)', '3%/30(중)', '3%/20(하)']
        inv_select3_1 = self.inv_select3_1 = tkinter.ttk.Combobox(self.window, width=12,
                                                                        values=inv_type_list2)
        inv_select3_1.place(x=785, y=385)
        inv_select3_1.set("축스탯%/1각")
        inv_select3_2 = self.inv_select3_2 = tkinter.ttk.Combobox(self.window, width=12,
                                                                        values=inv_value_list3)
        inv_select3_2.place(x=785, y=412)
        inv_select3_2.set('3%/60(상)')
        inv_select4_1 = self.inv_select4_1 = tkinter.ttk.Combobox(self.window, width=12,
                                                                        values=inv_type_list2_1)
        inv_select4_1.place(x=785, y=440)
        inv_select4_1.set("축스탯%/1각")
        inv_select4_2 = self.inv_select4_2 = tkinter.ttk.Combobox(self.window, width=12,
                                                                        values=inv_value_list3_1)
        inv_select4_2.place(x=785, y=467)
        inv_select4_2.set('3%/40(상)')
        inv_select3_1.bind("<<ComboboxSelected>>", self.update_inv_buf)
        inv_select4_1.bind("<<ComboboxSelected>>", self.update_inv_buf2)

        self.update_inv(0)

        know_image = self.get_photo_image("set_name/know_name.png")
        tkinter.Button(self.window, bg=self.dark_main, image=know_image,
                       command=self.create_knowledge_window).place(x=302, y=520)

        self.legend_on_tg = IntVar()

        select_default_lengend = self.select_default_lengend = tkinter.Button(self.window, relief='flat', borderwidth=0,
                                                activebackground=self.dark_main, bg=self.dark_main,
                                                image=self.default_img1n, command=lambda: self.change_default(0))
        select_default_lengend.place(x=492 + 15, y=516)

        select_default_chawon = self.select_default_chawon = tkinter.Button(self.window, relief='flat', borderwidth=0,
                                               activebackground=self.dark_main, bg=self.dark_main,
                                               image=self.default_img2f, command=lambda: self.change_default(1))
        select_default_chawon.place(x=522 + 15, y=516)

        select_default_old = self.select_default_old = tkinter.Button(self.window, relief='flat', borderwidth=0,
                                            activebackground=self.dark_main, bg=self.dark_main,
                                            image=self.default_img3f, command=lambda: self.change_default(2))
        select_default_old.place(x=552 + 15, y=516)

        tkinter.Checkbutton(self.window, variable=self.legend_on_tg, bg=self.dark_main,
                            activebackground=self.dark_main, bd=0).place(x=432, y=547)

        tkinter.Label(self.window, text="레전 적극 반영 여부(느려짐)", font=self.small_font, fg='white',
                      bg=self.dark_main, bd=0).place(x=450, y=551)

        default_tag_img = self.get_photo_image("ext_img/default_tag.png")
        tkinter.Label(self.window, bg=self.dark_main, image=default_tag_img).place(x=431, y=515)

        def donate():
            webbrowser.open('https://twip.kr/dawnclass16')

        donate_image = self.get_photo_image('ext_img/donate.png')
        donate_bt = tkinter.Button(self.window, image=donate_image, command=donate, borderwidth=0,
                                   bg=self.dark_main, activebackground=self.dark_main)
        donate_bt.place(x=622, y=520)

        def dunfaoff():
            webbrowser.open('https://dunfaoff.com/')

        dunfaoff_image = self.get_photo_image('ext_img/dunfaoff.png')
        dunfaoff_url = tkinter.Button(self.window, image=dunfaoff_image, command=dunfaoff, borderwidth=0,
                                      bg=self.dark_main, activebackground=self.dark_main)
        dunfaoff_url.place(x=535 + 219, y=410 - 402 + 32)

        def blog():
            webbrowser.open('https://blog.naver.com/dawnclass16/221837654941')

        blog_image = self.get_photo_image('ext_img/blog.png')
        blog_url = tkinter.Button(self.window, image=blog_image, command=blog, borderwidth=0,
                                  bg=self.dark_main, activebackground=self.dark_main)
        blog_url.place(x=615 + 219, y=410 - 402 + 32)

        def hamjung():
            tkinter.messagebox.showinfo("제작자 크레딧",
                                        "총제작자=Dawnclass(새벽반)\n이미지/그래픽=경철부동산\n직업/버퍼DB=대략볼록할철\n기타조언=히든 도비 4,5,6,7호\n\n오류 제보는 블로그 덧글이나 던조 쪽지로")

        maker_image = self.get_photo_image('ext_img/maker.png')
        maker = tkinter.Button(self.window, image=maker_image, command=hamjung, borderwidth=0,
                               bg=self.dark_main, activebackground=self.dark_main)

        def check_update(event):
            try:
                now_version_num = int(self.now_version[0] + self.now_version[2] + self.now_version[4])
                if event == 1:
                    html = urllib.request.urlopen("https://drive.google.com/open?id=1p8ZdzW_NzGKHHOtfPTuZSr1YgSEVtYCj")
                    bsObject = BeautifulSoup(html, "html.parser")
                    for meta in bsObject.head.find_all('meta'):
                        if meta.get('content').count('zip') == 1:
                            net_version = str(meta.get('content'))[-9:-4]
                            print(f"최신 업데이트 버전 = {net_version}")

                    return net_version

                if event == 0:
                    net_version = net_latest_version
                    net_version_num = int(net_version[0] + net_version[2] + net_version[4])
                    if now_version_num < net_version_num:
                        webbrowser.open('https://drive.google.com/open?id=1p8ZdzW_NzGKHHOtfPTuZSr1YgSEVtYCj')
                    else:
                        tkinter.messagebox.showinfo('버전확인', "최신버전입니다.")
            except:
                if event == 0:
                    tkinter.messagebox.showerror('에러', "업데이트 체크 실패(네트워크 오류)")
                if event == 1:
                    return "-"

        net_latest_version = check_update(1)
        version = tkinter.Button(self.window,
                                 text='현재 ' + self.now_version + '\n최신 ' + str(net_latest_version) + '\n업데이트',
                                 font=self.small_font, command=lambda: check_update(0))
        maker.place(x=622, y=585)
        version.place(x=630 - 3, y=645 + 3)

    def init_equipments(self):
        # 일반 에픽
        normal_equip_combinations = [
            (ArmorEquipPartCode, [SetCode.ARMOR_BEGIN, SetCode.ARMOR_END + 1]),
            (AccessoryEquipPartCode, [SetCode.ACCESSORY_BEGIN, SetCode.ACCESSORY_END + 1]),
            (SpecialEquipPartCode, [SetCode.SPECIAL_BEGIN, SetCode.SPECIAL_END + 1]),
            (TopNeckSubEquipPartCode, [SetCode.TOP_NECK_SUB_BEGIN, SetCode.TOP_NECK_SUB_END + 1]),
            (BottomBraceStoneEquipPartCode, [SetCode.BOTTOM_BRACE_STONE_BEGIN, SetCode.BOTTOM_BRACE_STONE_END + 1]),
            (ShoesRingEarringEquipPartCode, [SetCode.SHOES_RING_EARRING_BEGIN, SetCode.SHOES_RING_EARRING_END + 1])
        ]

        for comb in normal_equip_combinations:
            for equip_part_code in comb[0]:
                for set_code in range(comb[1][0], comb[1][1]):
                    base_code = f"{equip_part_code}{set_code:02}"

                    normal_equip_code = f"{base_code}{GodCheckCode.NON_GOD}"

                    self.normal_epic_list.append(normal_equip_code)

                    if equip_part_code in list(GodEquipPartCode.__members__.values()):
                        god_equip_code = f"{base_code}{GodCheckCode.GOD}"
                        self.normal_epic_list.append(god_equip_code)

        # 지혜의 산물
        self.know_list = ["13390150", "22390240", "23390450", "33390750", "21390340", "31390540", "32390650",
                          "11390850", "12390950", "13391050", "14391150", "15391250"]
        self.know_set_list = ["22400150", "22400250", "22400350", "22400450", "22400550", "21400640", "31400750",
                              "31400850", "31400950", "31401050", "31401150", "32401240", "32401340", "32401440"]
        self.know_jin_list = ["11410100", "11410110", "11410120", "11410130", "11410140", "11410150", "21420100",
                              "21420110", "21420120", "21420130", "21420140", "21420150", "33430100", "33430110",
                              "33430120", "33430130", "33430140", "33430150"]

        # 시로코 에픽
        self.siroco_equip_list = ["41510", "41520", "41530", "41540", "41550", "42510", "42520", "42530", "42540",
                                  "42550", "43510", "43520", "43530", "43540", "43550"]

        all_equipments = self.normal_epic_list + self.get_all_knowledge_equipment_list() + self.siroco_equip_list

        for equip_code in all_equipments:
            self.owned_equipments[f"tg{equip_code}"] = 0

    def refresh_db(self, load_excel_ext: Workbook = None, load_preset_ext: Workbook = None):
        load_excel = load_excel_ext
        load_preset = load_preset_ext

        if load_excel is None:
            load_excel = load_workbook("DATA.xlsx", data_only=True)
        if load_preset is None:
            load_preset = load_workbook("preset.xlsx", data_only=True)

        db_custom = load_preset["custom"]

        ## 초기 구동 엑셀
        db_one = load_excel["one"]

        opt_one = self.opt_one

        opt_one.clear()
        name_one = {}
        a = 1

        for row in db_one.rows:
            row_value = []
            row_value_cut = []

            for cell in row:
                row_value.append(cell.value)
                row_value_cut = row_value[2:]

            opt_one[db_one.cell(a, 1).value] = row_value_cut
            name_one[db_one.cell(a, 1).value] = row_value
            a = a + 1

        db_job = load_excel["lvl"]

        opt_job = self.opt_job
        opt_job_ele = self.opt_job_ele

        opt_job.clear()
        opt_job_ele.clear()
        u = 1

        for row in db_job.rows:
            row_value = []

            for cell in row:
                row_value.append(cell.value)

            opt_job[db_job.cell(u, 1).value] = row_value[3:]
            opt_job_ele[db_job.cell(u, 1).value] = row_value[:3]
            u = u + 1

        del opt_job["empty"]
        del opt_job["직업명"]

        c = 1
        db_buf = load_excel["buf"]

        opt_buf = self.opt_buf
        opt_buf.clear()

        for row in db_buf.rows:
            row_value = []
            row_value_cut = []

            for cell in row:
                row_value.append(cell.value)
                row_value_cut = row_value[2:]

            opt_buf[db_buf.cell(c, 1).value] = row_value_cut
            c = c + 1

        d = 1
        db_buflvl = load_excel["buflvl"]

        opt_buflvl = self.opt_buflvl
        opt_buflvl.clear()

        for row in db_buflvl.rows:
            row_value = []
            row_value_cut = []

            for cell in row:
                row_value.append(cell.value)
                row_value_cut = [0] + row_value[1:]

            opt_buflvl[db_buflvl.cell(d, 1).value] = row_value_cut
            d = d + 1

        level_db = load_excel["leveling"]
        jk = 1
        opt_leveling = self.opt_leveling
        opt_leveling.clear()

        for row in level_db.rows:
            row_value = []

            for cell in row:
                row_value.append(cell.value)

            row_value_cut = row_value[2:]
            opt_leveling[level_db.cell(jk, 1).value] = row_value_cut  ## DB 불러오기 ##
            jk = jk + 1

        if load_excel_ext is None:
            load_preset.close()
        if load_excel_ext is None:
            load_excel.close()

        # Info loaded from db
        self.wep_list.clear()
        wep_list = self.wep_list

        for i in range(0, 75):
            wep_list.append(name_one[str(i + 111001)][1])

        wep_list.append(name_one["111076"][1])

    def load_preset_name(self, load_preset_ext: Workbook = None):
        load_preset = load_preset_ext

        if load_preset is None:
            load_preset = load_workbook("preset.xlsx", data_only=True)

        db_custom = load_preset["custom"]

        for i in range(0, 20):
            self.save_name_list.append(db_custom.cell(i + 1, 5).value)

        if load_preset_ext is None:
            load_preset.close()

    def update_preset_version(self, load_preset_ext: Workbook = None):
        load_preset = load_preset_ext

        if load_preset is None:
            load_preset = load_workbook("preset.xlsx", data_only=True)

        db_custom = load_preset["custom"]

        try:
            if str(db_custom['K2'].value) != '1':
                self.create_update_window()

            print(f"Preset 엑셀 버전 = {db_custom['K1'].value}")
            print(f"클라이언트 버전 = {self.now_version}")

            if str(db_custom['K1'].value) != self.now_version:
                print("DB 업데이트")
                db_custom['K1'] = self.now_version
                self.auto_custom = 1
                load_preset.save("preset.xlsx")
                load_preset.close()
                calc_update.update_preset()  ## 업데이트: 외부모듈
        except PermissionError as error:
            tkinter.messagebox.showerror("에러", "업데이트 실패. 엑셀을 닫고 다시 실행해주세요.")
            self.window.destroy()

        if load_preset_ext is None:
            load_preset.close()

    def change_state_text(self, text: str):
        self.calc_state_label.configure(text=text)

    def change_case_count_text(self, text: str):
        self.case_count_label.configure(text=text)

    def change_total_count_text(self, text: str, color: str):
        self.total_count_label.configure(text=text)
        self.total_count_label["fg"] = color

    def change_weapon_list_num_text(self, weapon_num: int):
        text = f"무기 수={weapon_num} / 10"
        self.weapon_list_num_label.configure(text=text)

    def change_selected_weapon_img(self, index: int, img: PhotoImage):
        self.selected_weapon_label_list[index].configure(image=img)

    def get_all_knowledge_equipment_list(self):
        return self.know_list + self.know_set_list + self.know_jin_list

    def click_equipment(self, code: str):
        select_item = self.owned_equipments
        equip_buttons = self.equip_buttons

        if select_item[f"tg{code}"] == 0:
            equip_buttons[code]['image'] = self.image_list[code]
            select_item[f"tg{code}"] = 1
        elif select_item[f"tg{code}"] == 1:
            equip_buttons[code]['image'] = self.image_list2[code]
            select_item[f"tg{code}"] = 0

        if len(code) == 5:
            self.check_set(int('1' + code[2:4]))

    ## 선택한 모든 장비 체크 초기화
    def reset_equipments(self):
        select_item = self.owned_equipments

        for j in [1000, 2000, 3000, 4000]:
            if j == 1000:
                end_range = 536
            else:
                end_range = 356
            for i in range(j + 101, j + end_range):
                try:
                    select_item['tg{}0'.format(i)] = 0
                except KeyError as error:
                    passss = 1
                try:
                    select_item['tg{}1'.format(i)] = 0
                except KeyError as error:
                    passss = 1

        for i in self.get_all_knowledge_equipment_list():
            select_item[f"tg{i}"] = 0

        self.check_equipment()
        self.wep_list_reset()

        for i in range(101, 156):
            try:
                self.check_set(i)
            except:
                pass

    ## 세트태그 선택시 풀셋 전부 온오프
    def click_set(self, code: int):
        select_item = self.owned_equipments
        set_buttons = self.set_buttons
        equip_buttons = self.equip_buttons

        code_add = code - 100
        code_str = str(code)[1:3]
        set_checked = 0
        if code >= 116:  ##악세/특장/스까면
            if 116 <= code <= 119:
                for i in range(21, 24):  ## 악세부위에서
                    try:
                        if select_item['tg' + str(i) + code_str + '0'] == 1:  ##채택된 숫자를 찾는다
                            set_checked = set_checked + 1  ##그럼 변수에 +1을 더함
                    except KeyError as error:
                        c = 1
            elif 123 >= code >= 120:
                for i in range(31, 34):  ## 특장부위에서
                    try:
                        if select_item['tg' + str(i) + code_str + '0'] == 1:  ##채택된 숫자를 찾는다
                            set_checked = set_checked + 1  ##그럼 변수에 +1을 더함
                    except KeyError as error:
                        c = 1
            elif 131 >= code >= 128:
                for i in [11, 22, 31]:  ## 상목보부위에서
                    try:
                        if select_item['tg' + str(i) + code_str + '0'] == 1:  ##채택된 숫자를 찾는다
                            set_checked = set_checked + 1  ##그럼 변수에 +1을 더함
                    except KeyError as error:
                        c = 1
            elif 127 >= code >= 124:
                for i in [12, 21, 32]:  ## 하팔법부위에서
                    try:
                        if select_item['tg' + str(i) + code_str + '0'] == 1:  ##채택된 숫자를 찾는다
                            set_checked = set_checked + 1  ##그럼 변수에 +1을 더함
                    except KeyError as error:
                        c = 1
            elif 135 >= code >= 132:
                for i in [15, 23, 33]:  ## 신반귀부위에서
                    try:
                        if select_item['tg' + str(i) + code_str + '0'] == 1:  ##채택된 숫자를 찾는다
                            set_checked = set_checked + 1  ##그럼 변수에 +1을 더함
                    except KeyError as error:
                        c = 1
            elif 155 >= code >= 151:
                for i in [41, 42, 43]:  ##융합부위에서
                    try:
                        if select_item['tg' + str(i) + code_str + '0'] == 1:  ##채택된 숫자를 찾는다
                            set_checked = set_checked + 1  ##그럼 변수에 +1을 더함
                    except KeyError as error:
                        c = 1

            if set_checked == 3:  ## 채택 숫자가 3이면
                for i in range(11, 44):  ##모든 부위에서
                    try:
                        str_i = str(i)
                        equip_buttons[str_i + code_str + '0']['image'] = self.image_list2[
                            str_i + code_str + '0']  ##이미지도 오프로 바꿈
                        select_item['tg' + str_i + code_str + '0'] = 0  ##모든 체크를 0으로 만듬
                    except KeyError as error:
                        c = 1
                set_buttons[str(code)]['image'] = self.image_list_set2[str(code)]  ##세트이미지도 오프로 바꿈
            else:  ## 채택 숫자가 3미만이면
                for i in range(11, 44):  ##모든 부위에서
                    try:
                        str_i = str(i)
                        equip_buttons[str_i + code_str + '0']['image'] = self.image_list[
                            str_i + code_str + '0']  ##이미지도 온으로 바꿈
                        select_item['tg' + str_i + code_str + '0'] = 1  ##모든 체크를 1으로 만듬
                    except KeyError as error:
                        c = 1
                set_buttons[str(code)]['image'] = self.image_list_set[str(code)]  ##세트이미지도 온으로 바꿈


        else:
            for i in range(11, 16):  ## 방어구 부위에서
                try:
                    if select_item['tg' + str(i) + code_str + '0'] == 1:  ##채택된 숫자를 찾는다
                        set_checked = set_checked + 1  ##그럼 변수에 +1을 더함
                except KeyError as error:
                    c = 1

            if set_checked == 5:  ## 채택 숫자가 5이면
                for i in range(11, 16):  ## 방어구 부위에서
                    try:
                        equip_buttons[str(i) + code_str + '0']['image'] = self.image_list2[
                            str(i) + code_str + '0']  ##이미지도 오프로 바꿈
                        select_item['tg' + str(i) + code_str + '0'] = 0  ##모든 체크를 0으로 만듬
                    except KeyError as error:
                        c = 1
                set_buttons[str(code)]['image'] = self.image_list_set2[str(code)]  ##세트이미지도 오프로 바꿈

            else:  ## 채택 숫자가 5미만이면
                for i in range(11, 16):  ## 방어구 부위에서
                    try:
                        str_i = str(i)
                        equip_buttons[str_i + code_str + '0']['image'] = self.image_list[
                            str_i + code_str + '0']  ##이미지도 온으로 바꿈
                        select_item['tg' + str_i + code_str + '0'] = 1  ##모든 체크를 1으로 만듬
                    except KeyError as error:
                        c = 1
                set_buttons[str(code)]['image'] = self.image_list_set[str(code)]  ##세트이미지도 온으로 바꿈

    def wep_list_select(self):
        wep_name_list = self.wep_name_list
        wep_select = self.wep_select

        if wep_name_list.count(wep_select.get()) != 0:
            tkinter.messagebox.showerror('에러', "중복된 무기 선택")
            return

        if len(wep_name_list) != 10:
            wep_name_list.append(wep_select.get())
            self.selected_weapon_img_list.append(self.image_list_wep[wep_select.get()])
            self.change_weapon_list_num_text(len(wep_name_list))
            self.wep_img_list_refresh(self.selected_weapon_img_list)
        else:
            tkinter.messagebox.showerror('에러', "무기는 최대 10가지만 선택 가능합니다")
            return

    def wep_list_reset(self):
        wep_name_list = self.wep_name_list

        wep_name_list.clear()
        self.change_weapon_list_num_text(len(wep_name_list))

        self.selected_weapon_img_list.clear()
        self.wep_img_list_refresh(self.selected_weapon_img_list)

    def set_combobox_value(self, name: str, value: str):
        self.preset_values[name]().set(value)

    def get_combobox_value(self, name: str):
        return self.preset_values[name]().get()

    def calc_thread(self):
        threading.Thread(target=self.calc, args=(0,), daemon=True).start()

    # 정지
    def stop_calc(self):
        # TODO: thread fix
        self.exit_calc.value = 1
        time.sleep(1)
        self.exit_calc.value = 0

    ## 저장된 preset 불러오기
    def load_checklist(self):
        ask_msg1 = tkinter.messagebox.askquestion('확인', "저장된 내역을 불러오겠습니까?")
        for snum in range(0, 20):
            if self.save_select.get() == self.save_name_list[snum]:
                ssnum1 = snum
        if ask_msg1 == 'yes':
            load_preset3 = load_workbook("preset.xlsx")
            db_load_check = load_preset3["one"]
            db_load_cus = load_preset3["custom"]
            load_cell = db_load_check.cell
            load_cus = db_load_cus.cell
            k = 1

            select_item = self.owned_equipments

            for i in range(1, 316):
                if load_cell(i, 2 + ssnum1).value == 1:
                    try:
                        select_item['tg{}'.format(load_cell(i, 1).value)] = 1
                    except KeyError as error:
                        passss = 1
                elif load_cell(i, 2 + ssnum1).value == 0:
                    try:
                        select_item['tg{}'.format(load_cell(i, 1).value)] = 0
                    except KeyError as error:
                        passss = 1
            for i in range(52, 70):
                temp_opt = str(load_cus(i, 1).value)
                temp_val = str(load_cus(i, 2 + ssnum1).value)
                self.set_combobox_value(temp_opt, temp_val)
            for i in range(1, 20):
                load_cus(i, 2).value = str(load_cus(i + 25, 2 + ssnum1).value)
            for i in range(1, 8):
                load_cus(i, 8).value = str(load_cus(i + 44, 2 + ssnum1).value)
            load_cus(20, 2).value = str(load_cus(70, 2 + ssnum1).value)
            saved_wep_str = load_cus(71, 2 + ssnum1).value

            self.wep_name_list = eval(saved_wep_str)
            if not isinstance(self.wep_name_list, list):
                self.wep_name_list = []

            self.sync_wep_list()
            load_preset3.save("preset.xlsx")
            load_preset3.close()
            self.check_equipment()
            for i in range(101, 136):
                self.check_set(i)
            for i in range(151, 156):
                self.check_set(i)

            def load_inv():
                inv_select3_1 = self.inv_select3_1
                inv_select3_2 = self.inv_select3_2

                if inv_select3_1.get() == "축스탯%/1각":
                    inv_select3_2['values'] = ['3%/60(상)', '3%/40(중)', '3%/20(하)']
                elif inv_select3_1.get() == "축스탯%/1각%":
                    inv_select3_2['values'] = ['4%/3%(상)', '3%/3%(중)', '2%/3%(하)']
                elif inv_select3_1.get() == "축앞뎀%/1각":
                    inv_select3_2['values'] = ['4%/25(상)', '3%/25(중)', '2%/25(하)']
                elif inv_select3_1.get() == "축앞뎀%/1각%":
                    inv_select3_2['values'] = ['3%/3%(상)', '3%/2%(중)', '3%/1%(하)']
                elif inv_select3_1.get() == "전직패":
                    inv_select3_2['values'] = ['+185(상)', '+155(중)', '+125(하)']
                elif inv_select3_1.get() == "축스탯%/1각+1":
                    inv_select3_2['values'] = ['3%/+1(상)', '2%/+1(중)', '1%/+1(하)']

            def load_inv2():
                inv_select4_1 = self.inv_select4_1
                inv_select4_2 = self.inv_select4_2

                if inv_select4_1.get() == "축스탯%/1각":
                    inv_select4_2['values'] = ['3%/40(상)', '3%/30(중)', '3%/20(하)']
                elif inv_select4_1.get() == "축스탯%/1각%":
                    inv_select4_2['values'] = ['4%/2%(상)', '3%/2%(중)', '2%/2%(하)']
                elif inv_select4_1.get() == "축앞뎀%/1각":
                    inv_select4_2['values'] = ['3%/25(상)', '2%/25(중)', '1%/25(하)']
                elif inv_select4_1.get() == "축앞뎀%/1각%":
                    inv_select4_2['values'] = ['2%/3%(상)', '2%/2%(중)', '2%/1%(하)']
                elif inv_select4_1.get() == "전직패":
                    inv_select4_2['values'] = ['+145(상)', '+115(중)', '+85(하)']
                elif inv_select4_1.get() == "축+1/1각":
                    inv_select4_2['values'] = ['+1/30(상)', '+1/20(중)', '+1/10(하)']

            self.update_inv(0)

            def load_wep():
                wep_job_select = self.wep_job_select
                wep_type_select = self.wep_type_select
                wep_select = self.wep_select

                wep_type_select["values"] = list(calc_list_wep.DNF_wep_list[str(wep_job_select.get())].keys())
                wep_select["values"] = list(
                    calc_list_wep.DNF_wep_list[str(wep_job_select.get())][str(wep_type_select.get())])

            try:
                load_inv()
                load_inv2()
                load_wep()
            except:
                pass

            jobup_select = self.jobup_select
            jobtype_select = self.jobtype_select

            jobup_select["values"] = list(calc_list_job.DNF_job_list[jobtype_select.get()])
            tkinter.messagebox.showinfo("알림", "불러오기 완료")

    ## 현재값 preset에 저장하기
    def save_checklist(self):
        ask_msg2 = tkinter.messagebox.askquestion('확인', "저장하시겠습니까?")
        for snum in range(0, 20):
            if self.save_select.get() == self.save_name_list[snum]:
                ssnum2 = snum
        try:
            if ask_msg2 == 'yes':
                load_preset4 = load_workbook("preset.xlsx")
                db_save_check = load_preset4["one"]
                db_save_cus = load_preset4["custom"]
                save_cell = db_save_check.cell
                save_cus = db_save_cus.cell
                opt_save = {}
                for i in range(1, 316):
                    opt_save[save_cell(i, 1).value] = i

                select_item = self.owned_equipments

                for code in opt_save.keys():
                    try:
                        if select_item[f"tg{code}"] == 1:
                            save_cell(opt_save[code], 2 + ssnum2).value = 1
                    except KeyError as error:
                        passss1 = 1

                    try:
                        if select_item[f"tg{code}"] == 0:
                            save_cell(opt_save[code], 2 + ssnum2).value = 0
                    except KeyError as error:
                        passss1 = 1

                    passss = 1
                for i in range(52, 70):
                    temp_opt = str(save_cus(i, 1).value)
                    temp_val = self.get_combobox_value(temp_opt)
                    save_cus(i, 2 + ssnum2).value = temp_val
                for i in range(1, 20):
                    save_cus(i + 25, 2 + ssnum2).value = str(save_cus(i, 2).value)
                save_cus(70, 2 + ssnum2).value = str(save_cus(20, 2).value)
                save_cus(71, 2 + ssnum2).value = str(self.wep_name_list)
                for i in range(1, 8):
                    save_cus(i + 44, 2 + ssnum2).value = str(save_cus(i, 8).value)

                load_preset4.save("preset.xlsx")
                load_preset4.close()
                tkinter.messagebox.showinfo("알림", "저장 완료")

        except PermissionError as error:
            tkinter.messagebox.showerror("에러", "엑셀을 닫고 다시 시도해주세요.")

    ##잔향부여
    def update_inv(self, event):
        inv_mod = self.inv_mod

        inv_select1_1 = self.inv_select1_1
        inv_select1_2 = self.inv_select1_2
        inv_select2_1 = self.inv_select2_1
        inv_select2_2 = self.inv_select2_2
        inv_select3_1 = self.inv_select3_1
        inv_select3_2 = self.inv_select3_2
        inv_select4_1 = self.inv_select4_1
        inv_select4_2 = self.inv_select4_2

        if inv_mod.get() == "미부여" or inv_mod.get() == "최적부여(버퍼X)":
            if inv_mod.get() == "미부여":
                self.inv_tg = 0
            elif inv_mod.get() == "최적부여(버퍼X)":
                self.inv_tg = 2

            inv_select1_1['state'] = 'disabled'
            inv_select1_2['state'] = 'disabled'
            inv_select2_1['state'] = 'disabled'
            inv_select2_2['state'] = 'disabled'
            inv_select3_1['state'] = 'disabled'
            inv_select3_2['state'] = 'disabled'
            inv_select4_1['state'] = 'disabled'
            inv_select4_2['state'] = 'disabled'
        elif inv_mod.get() == "선택부여":
            self.inv_tg = 1
            inv_select1_1['state'] = 'normal'
            inv_select1_2['state'] = 'normal'
            inv_select2_1['state'] = 'normal'
            inv_select2_2['state'] = 'normal'
            inv_select3_1['state'] = 'normal'
            inv_select3_2['state'] = 'normal'
            inv_select4_1['state'] = 'normal'
            inv_select4_2['state'] = 'normal'

    def update_inv_buf(self, event):
        inv_select3_1 = self.inv_select3_1
        inv_select3_2 = self.inv_select3_2

        if inv_select3_1.get() == "축스탯%/1각":
            inv_select3_2['values'] = ['3%/60(상)', '3%/40(중)', '3%/20(하)']
            inv_select3_2.set('3%/60(상)')
        elif inv_select3_1.get() == "축스탯%/1각%":
            inv_select3_2['values'] = ['4%/3%(상)', '3%/3%(중)', '2%/3%(하)']
            inv_select3_2.set('4%/3%(상)')
        elif inv_select3_1.get() == "축앞뎀%/1각":
            inv_select3_2['values'] = ['4%/25(상)', '3%/25(중)', '2%/25(하)']
            inv_select3_2.set('4%/25(상)')
        elif inv_select3_1.get() == "축앞뎀%/1각%":
            inv_select3_2['values'] = ['3%/3%(상)', '3%/2%(중)', '3%/1%(하)']
            inv_select3_2.set('3%/3%(상)')
        elif inv_select3_1.get() == "전직패":
            inv_select3_2['values'] = ['+185(상)', '+155(중)', '+125(하)']
            inv_select3_2.set('+185(상)')
        elif inv_select3_1.get() == "축스탯%/1각+1":
            inv_select3_2['values'] = ['3%/+1(상)', '2%/+1(중)', '1%/+1(하)']
            inv_select3_2.set('3%/+1(상)')

    def update_inv_buf2(self, event):
        inv_select4_1 = self.inv_select4_1
        inv_select4_2 = self.inv_select4_2

        if inv_select4_1.get() == "축스탯%/1각":
            inv_select4_2['values'] = ['3%/40(상)', '3%/30(중)', '3%/20(하)']
            inv_select4_2.set('3%/40(상)')
        elif inv_select4_1.get() == "축스탯%/1각%":
            inv_select4_2['values'] = ['4%/2%(상)', '3%/2%(중)', '2%/2%(하)']
            inv_select4_2.set('4%/2%(상)')
        elif inv_select4_1.get() == "축앞뎀%/1각":
            inv_select4_2['values'] = ['3%/25(상)', '2%/25(중)', '1%/25(하)']
            inv_select4_2.set('3%/25(상)')
        elif inv_select4_1.get() == "축앞뎀%/1각%":
            inv_select4_2['values'] = ['2%/3%(상)', '2%/2%(중)', '2%/1%(하)']
            inv_select4_2.set('2%/3%(상)')
        elif inv_select4_1.get() == "전직패":
            inv_select4_2['values'] = ['+145(상)', '+115(중)', '+85(하)']
            inv_select4_2.set('+145(상)')
        elif inv_select4_1.get() == "축+1/1각":
            inv_select4_2['values'] = ['+1/30(상)', '+1/20(중)', '+1/10(하)']
            inv_select4_2.set('+1/30(상)')

    def change_default(self, value):
        if value == 0:
            self.default_base_equip = 0
            self.select_default_lengend['image'] = self.default_img1n
            self.select_default_chawon['image'] = self.default_img2f
            self.select_default_old['image'] = self.default_img3f
        elif value == 1:
            self.default_base_equip = 1
            self.select_default_lengend['image'] = self.default_img1f
            self.select_default_chawon['image'] = self.default_img2n
            self.select_default_old['image'] = self.default_img3f
        elif value == 2:
            self.default_base_equip = 2
            self.select_default_lengend['image'] = self.default_img1f
            self.select_default_chawon['image'] = self.default_img2f
            self.select_default_old['image'] = self.default_img3n

    def capture_screen(self, toplevel):
        nowx=toplevel.winfo_x()+8
        nowy=toplevel.winfo_y()
        xsize=int(toplevel.geometry().split('+')[0][:3])
        ysize=int(toplevel.geometry().split('+')[0][4:])+32
        #im=grab(bbox=(nowx, nowy, nowx+xsize, nowy+ysize))
        im=ImageGrab.grab((nowx, nowy, nowx+xsize, nowy+ysize))
        im.save('Screenshots/'+str(time.strftime('%y%m%d%H%M%S', time.localtime(time.time())))+'.png')


    ## 계산 함수 ##
    def calc(self, mode):
        select_item = self.owned_equipments
        jobup_select = self.jobup_select
        req_cool = self.req_cool

        try:
            result_window = self.result_window
            result_window.after(0,result_window.destroy) #기존 GIF 재생 정지
            result_window.after(50,self.time_delay2)
            result_window.after(50,self.time_delay4)
            result_window.after(500,self.time_delay1)
            result_window.after(500,self.time_delay3)
        except AttributeError as error:
            pass

        select_perfect: tkinter.ttk.Combobox = self.select_perfect

        if select_perfect.get()[0:5] == '세트필터↓' or select_perfect.get()[0:4] == '풀셋모드' or select_perfect.get() == '메타몽풀셋모드':
            set_perfect = 1 #세트 필터 하락
        else:
            set_perfect = 0
        if self.a_num_all > 20000000:
            if select_perfect.get()[0:4] != "풀셋모드" and select_perfect.get() != "메타몽풀셋모드":
                ask_really=tkinter.messagebox.askquestion('확인',"2천만 가지가 넘는 경우의 수는 풀셋/메타몽풀셋 모드를 권장합니다.\n그냥 진행하시겠습니까?")
                if ask_really == 'yes':
                    pass
                elif ask_really == 'no':
                    self.change_state_text(text='중지됨')
                    return
        self.change_state_text(text="조합 알고리즘 구동 준비중...")
        start_time = time.time()

        ## 갱신된 DB 불러오기 ##
        load_excel = load_workbook("DATA.xlsx", data_only=True)
        load_presetc = load_workbook("preset.xlsx", data_only=True)

        self.refresh_db(load_excel_ext=load_excel, load_preset_ext=load_presetc)

        db_job = load_excel["lvl"]
        db_one = load_excel["one"]

        betterang = int(db_one["J86"].value)

        #속강 커스텀 계산
        db_preset=load_presetc["custom"]
        ele_skill=int(self.opt_job_ele[jobup_select.get()][1])
        ele_in=(int(db_preset["B14"].value)+int(db_preset["B15"].value)+int(db_preset["B16"].value)+
                int(ele_skill)-int(db_preset["B18"].value)+int(db_preset["B19"].value)+13)
        cool_eff=float(db_preset["B2"].value)/100
        cool_eff2=float(db_preset["B20"].value)/100
        if req_cool.get()=='X(지속딜만)':
            cool_on=0
        else:
            cool_on=1

        #아리아 증폭 판정
        if db_preset["H7"].value == "항상증폭":
            aria_fix=0.3
            aria_dif=0
        elif db_preset["H7"].value == "템에따라":
            aria_fix=0.25
            aria_dif=1
        elif db_preset["H7"].value == "항상미증폭":
            aria_fix=0.25
            aria_dif=0

        self.show_number = 0

        #진각/2각에 따른 실마리, 쿨감 효율 차이
        if jobup_select.get()[-4:] == "(진각)":
            silmari=0
            active_eff_one=15
            active_eff_set=18-3
            cool_eff_dictnum=0
        else:
            silmari=1
            active_eff_one=21
            active_eff_set=27-3
            cool_eff_dictnum=27

        # 직업별 레벨링 효율차
        opt_job = self.opt_job

        job_lv1=opt_job[jobup_select.get()][11]
        job_lv2=opt_job[jobup_select.get()][12]
        job_lv3=opt_job[jobup_select.get()][13]
        job_lv4=opt_job[jobup_select.get()][14]
        job_lv5=opt_job[jobup_select.get()][15]
        job_lv6=opt_job[jobup_select.get()][16]
        job_pas0=opt_job[jobup_select.get()][0]
        job_pas1=opt_job[jobup_select.get()][1]
        job_pas2=opt_job[jobup_select.get()][2]
        job_pas3=opt_job[jobup_select.get()][3]
        #직업별 각성기 스증 효율차이
        job_ult1=opt_job[jobup_select.get()][17]
        job_ult2=opt_job[jobup_select.get()][18]
        job_ult3=opt_job[jobup_select.get()][19]

        ##무기 다중화
        check_wep_tg=0
        wep_pre_calced=[];cool_pre_calced=[];cool_pre_calced2=[];wep_num=[]

        if len(self.wep_name_list) == 0:
            wep_select = self.wep_select
            wep_name_list_temp = [wep_select.get()]
        else:
            wep_name_list_temp = self.wep_name_list
        for now_wep in wep_name_list_temp:
            for i, weapon_name in enumerate(self.wep_list):
                if now_wep == weapon_name:
                    wep_num.append((str(i+111001),))
            if now_wep.count("흑천의 주인")==1: wep_num.append(("111001",))
            elif now_wep.count("원초의 꿈")==1: wep_num.append(("111076",))
            wep_type=now_wep[1:now_wep.find(")")]
            for i in range(1,63):
                if jobup_select.get()==db_job['A'+str(i)].value:
                    for j in range(28,66):
                        if wep_type==db_job.cell(63,j).value:
                            wep_pre_calced.append(float(db_job.cell(i,j).value))
                            cool_pre_calced.append((1/float(db_job.cell(i,j+38).value)-1)*cool_eff*cool_on+1)
                            cool_pre_calced2.append((1/float(db_job.cell(i,j+38).value)-1)*cool_eff2+1)
        for i in range(len(wep_name_list_temp)):
            if wep_pre_calced[i]==0 and jobup_select.get()[0:4]!="(버프)" and check_wep_tg==0:
                tkinter.messagebox.showinfo('확인',"착용할 수 없는 무기를 선택했습니다. 이 경우 모든 무기 보정이 비활성화되어 마스터리/앞뎀 반영이 되지 않습니다.")
                check_wep_tg=1
                wep_pre_calced[i]=1
                cool_pre_calced[i]=1
                cool_pre_calced2[i]=1
            elif wep_pre_calced[i]==0 and jobup_select.get()[0:4]=="(버프)":
                tkinter.messagebox.showerror('에러',"버퍼가 사용할 수 없는 무기입니다.")
                self.change_state_text(text='중지됨')
                return

        ## 외부모듈: 경우의 수 만들기
        result_all_equ_list=make_all_equ_list(select_item,select_perfect.get())
        equip_list = self.equip_list

        equip_list["11"] = result_all_equ_list[0][0]
        equip_list["11_0"] = result_all_equ_list[0][1]
        equip_list["11_1"] = result_all_equ_list[0][2]
        equip_list["12"] = result_all_equ_list[0][3]
        equip_list["13"] = result_all_equ_list[0][4]
        equip_list["14"] = result_all_equ_list[0][5]
        equip_list["15"] = result_all_equ_list[0][6]
        equip_list["21"] = result_all_equ_list[0][7]
        equip_list["21_0"] = result_all_equ_list[0][8]
        equip_list["21_1"] = result_all_equ_list[0][9]
        equip_list["22"] = result_all_equ_list[0][10]
        equip_list["23"] = result_all_equ_list[0][11]
        equip_list["31"] = result_all_equ_list[0][12]
        equip_list["32"] = result_all_equ_list[0][13]
        equip_list["33"] = result_all_equ_list[0][14]
        equip_list["33_0"] = result_all_equ_list[0][15]
        equip_list["33_1"] = result_all_equ_list[0][16]

        set_num_dict=result_all_equ_list[1][0] ##전부
        set_num_dict1=result_all_equ_list[1][1] ##에픽만
        set_num_dict2=result_all_equ_list[1][2] ##신화만

        set_max_list1=[] ##세트 갯수 리스트
        set_max_list2=[]
        set_max_list3=[]
        set_max_list4=[]
        set_max_list5=[]
        set_max_list6=[]
        for i in range(1,36):
            if set_num_dict.get(str(i+100)) != None:
                if i < 16:
                    set_max_list1.append(set_num_dict.get(str(i+100))) #방
                elif i < 20:
                    set_max_list2.append(set_num_dict.get(str(i+100))) #악
                elif i < 24:
                    set_max_list3.append(set_num_dict.get(str(i+100))) #특
                elif i < 28:
                    set_max_list4.append(set_num_dict.get(str(i+100))) #상
                elif i < 32:
                    set_max_list5.append(set_num_dict.get(str(i+100))) #하
                elif i < 36:
                    set_max_list6.append(set_num_dict.get(str(i+100))) #신
            else:
                if i < 16:
                    set_max_list1.append(0)
                elif i < 20:
                    set_max_list2.append(0)
                elif i < 24:
                    set_max_list3.append(0)
                elif i < 28:
                    set_max_list4.append(0)
                elif i < 32:
                    set_max_list5.append(0)
                elif i < 36:
                    set_max_list6.append(0)

        know_dict={}
        ##단품산물
        for know_one in self.know_list:
            equip_part = know_one[0:2]

            if select_item[f"tg{know_one}"] == 1:
                equip_list[equip_part].append(know_one)

                if equip_part in ["11", "21", "33"]:
                    equip_list[f"{equip_part}_0"].append(equip_part)

        list41=[];list42=[];list43=[]
        ##시로코: 융합
        for i in ['41','42','43']:
            for j in ['51','52','53','54','55']:
                if select_item['tg{}{}0'.format(i,j)] == 1:
                    if i=='41':
                        list41.append(j[1])
                    elif i=='42':
                        list42.append(j[1])
                    elif i=='43':
                        list43.append(j[1])
        inv2_on_tg=0 #잔향 2부여 세트옵션 발동 여부
        if len(list41)!=0 and len(list42)!=0 and len(list43)!=0:
            inv2_on_tg=1
        if len(list41)==0:
            list41=['0']
        if len(list42)==0:
            list42=['0']
        if len(list43)==0:
            list43=['0']
        div_list=[list41,list42,list43]
        div_list_all=list(itertools.product(*div_list))
        list40=[]
        for i in div_list_all:
            list40.append('4'+i[0]+i[1]+i[2])
        max_div_set=0
        for i in list40:
            now_div=Counter(i[1]+i[2]+i[3])
            del now_div['0']
            try:
                now_div_set=max(now_div.values())
            except:
                now_div_set=0
            if now_div_set >= max_div_set:
                max_div_set=now_div_set
        list40_0=[]
        for i in list40:
            now_div=Counter(i[1]+i[2]+i[3])
            del now_div['0']
            try:
                now_div_set=max(now_div.values())
            except:
                now_div_set=0
            if now_div_set >= max_div_set:
                list40_0.append(i)

        ##딜러 선입력 데미지 옵션
        fixed_dam=0;fixed_cri=0;extra_dam=0;extra_cri=0;extra_bon=0
        extra_all=0;extra_att=0;extra_sta=0;extra_pas2=0
        ##버퍼 선입력 버프 옵션
        extra_cper=0;extra_bstat=0;extra_clvl=0
        extra_blvl=0;extra_batt=0;extra_cstat=0
        extra_stat=0

        #칭호
        style_select = self.style_select
        creature_select = self.creature_select

        style_calced = self.style_calced = style_select.get()
        creature_calced = self.creature_calced = creature_select.get()

        if style_calced == '증뎀10%':
            fixed_dam=10
        elif style_calced == '증뎀15%':
            fixed_dam=15
        elif style_calced == '추뎀10%':
            extra_bon=10
        elif style_calced == '속강32':
            ele_in=ele_in+32
        elif style_calced == '크증10%':
            fixed_cri=10

        if creature_calced == '증뎀10%':
            fixed_dam = max(fixed_dam, 10)
        elif creature_calced == '크증10%':
            fixed_cri = 10
        elif creature_calced == '모공15%':
            extra_all=15
        elif creature_calced == '크증18%':
            fixed_cri=18
            extra_pas2=1
        elif creature_calced == '물마독공18%':
            extra_att=18
            extra_pas2=1

        #잔향 부여 선기입 (직접 선택)
        inv1_opt = ""
        inv1_val = 0
        inv2_opt = ""
        inv2_val = 0
        inv3_opt = ""
        inv3_val = 0
        inv4_opt = ""
        inv4_val = 0

        if self.inv_tg ==1:
            inv_select1_1 = self.inv_select1_1
            inv_select1_2 = self.inv_select1_2
            inv_select2_1 = self.inv_select2_1
            inv_select2_2 = self.inv_select2_2

            inv1_opt=inv_select1_1.get()
            inv1_val=int(inv_select1_2.get())
            inv2_opt=inv_select2_1.get()
            inv2_val=int(inv_select2_2.get())
            if inv_select1_1.get()=="증뎀":
                extra_dam=extra_dam+int(inv_select1_2.get())
            elif inv_select1_1.get()=="크증":
                extra_cri=extra_cri+int(inv_select1_2.get())
            elif inv_select1_1.get()=="추뎀":
                extra_bon=extra_bon+int(inv_select1_2.get())
            elif inv_select1_1.get()=="모공":
                extra_all=extra_all+int(inv_select1_2.get())
            elif inv_select1_1.get()=="공%":
                extra_att=extra_att+int(inv_select1_2.get())
            elif inv_select1_1.get()=="스탯":
                extra_sta=extra_sta+int(inv_select1_2.get())

            if inv_select2_1.get()=="증뎀" and inv2_on_tg==1:
                extra_dam=extra_dam+int(inv_select2_2.get())
            elif inv_select2_1.get()=="크증" and inv2_on_tg==1:
                extra_cri=extra_cri+int(inv_select2_2.get())
            elif inv_select2_1.get()=="추뎀" and inv2_on_tg==1:
                extra_bon=extra_bon+int(inv_select2_2.get())
            elif inv_select2_1.get()=="모공" and inv2_on_tg==1:
                extra_all=extra_all+int(inv_select2_2.get())
            elif inv_select2_1.get()=="공%" and inv2_on_tg==1:
                extra_att=extra_att+int(inv_select2_2.get())
            elif inv_select2_1.get()=="스탯" and inv2_on_tg==1:
                extra_sta=extra_sta+int(inv_select2_2.get())

            inv_select3_1 = self.inv_select3_1
            inv_select3_2 = self.inv_select3_2
            inv_select4_1 = self.inv_select4_1
            inv_select4_2 = self.inv_select4_2

            inv3_opt=inv_select3_1.get()
            inv3_val=inv_select3_2.get()[:-3]
            inv4_opt=inv_select4_1.get()
            inv4_val=inv_select4_2.get()[:-3]
            if inv_select3_1.get()=="축스탯%/1각":
                if inv_select3_2.get()[-2:-1]=="상":
                    extra_bstat=(extra_bstat/100+1)*1.03*100-100;extra_cstat=extra_cstat+60
                elif inv_select3_2.get()[-2:-1]=="중":
                    extra_bstat=(extra_bstat/100+1)*1.03*100-100;extra_cstat=extra_cstat+40
                elif inv_select3_2.get()[-2:-1]=="하":
                    extra_bstat=(extra_bstat/100+1)*1.03*100-100;extra_cstat=extra_cstat+20
            elif inv_select3_1.get()=="축스탯%/1각%":
                if inv_select3_2.get()[-2:-1]=="상":
                    extra_bstat=(extra_bstat/100+1)*1.04*100-100;extra_cper=(extra_cper/100+1)*1.03*100-100
                elif inv_select3_2.get()[-2:-1]=="중":
                    extra_bstat=(extra_bstat/100+1)*1.03*100-100;extra_cper=(extra_cper/100+1)*1.03*100-100
                elif inv_select3_2.get()[-2:-1]=="하":
                    extra_bstat=(extra_bstat/100+1)*1.02*100-100;extra_cper=(extra_cper/100+1)*1.03*100-100
            elif inv_select3_1.get()=="축앞뎀%/1각":
                if inv_select3_2.get()[-2:-1]=="상":
                    extra_batt=(extra_batt/100+1)*1.04*100-100;extra_cstat=extra_cstat+25
                elif inv_select3_2.get()[-2:-1]=="중":
                    extra_batt=(extra_batt/100+1)*1.03*100-100;extra_cstat=extra_cstat+25
                elif inv_select3_2.get()[-2:-1]=="하":
                    extra_batt=(extra_batt/100+1)*1.02*100-100;extra_cstat=extra_cstat+25
            elif inv_select3_1.get()=="축앞뎀%/1각%":
                if inv_select3_2.get()[-2:-1]=="상":
                    extra_batt=(extra_batt/100+1)*1.03*100-100;extra_cper=(extra_cper/100+1)*1.03*100-100
                elif inv_select3_2.get()[-2:-1]=="중":
                    extra_batt=(extra_batt/100+1)*1.03*100-100;extra_cper=(extra_cper/100+1)*1.02*100-100
                elif inv_select3_2.get()[-2:-1]=="하":
                    extra_batt=(extra_batt/100+1)*1.03*100-100;extra_cper=(extra_cper/100+1)*1.01*100-100
            elif inv_select3_1.get()=="전직패":
                if inv_select3_2.get()[-2:-1]=="상":
                    extra_stat=extra_stat+185
                elif inv_select3_2.get()[-2:-1]=="중":
                    extra_stat=extra_stat+155
                elif inv_select3_2.get()[-2:-1]=="하":
                    extra_stat=extra_stat+125
            elif inv_select3_1.get()=="축스탯%/1각+1":
                if inv_select3_2.get()[-2:-1]=="상":
                    extra_bstat=(extra_bstat/100+1)*1.03*100-100;extra_clvl=extra_clvl+1
                elif inv_select3_2.get()[-2:-1]=="중":
                    extra_bstat=(extra_bstat/100+1)*1.02*100-100;extra_clvl=extra_clvl+1
                elif inv_select3_2.get()[-2:-1]=="하":
                    extra_bstat=(extra_bstat/100+1)*1.01*100-100;extra_clvl=extra_clvl+1

            if inv_select4_1.get()=="축스탯%/1각" and inv2_on_tg==1:
                if inv_select4_2.get()[-2:-1]=="상":
                    extra_bstat=(extra_bstat/100+1)*1.03*100-100;extra_cstat=extra_cstat+40
                elif inv_select4_2.get()[-2:-1]=="중":
                    extra_bstat=(extra_bstat/100+1)*1.03*100-100;extra_cstat=extra_cstat+30
                elif inv_select4_2.get()[-2:-1]=="하":
                    extra_bstat=(extra_bstat/100+1)*1.03*100-100;extra_cstat=extra_cstat+20
            elif inv_select4_1.get()=="축스탯%/1각%" and inv2_on_tg==1:
                if inv_select4_2.get()[-2:-1]=="상":
                    extra_bstat=(extra_bstat/100+1)*1.04*100-100;extra_cper=(extra_cper/100+1)*1.02*100-100
                elif inv_select4_2.get()[-2:-1]=="중":
                    extra_bstat=(extra_bstat/100+1)*1.03*100-100;extra_cper=(extra_cper/100+1)*1.02*100-100
                elif inv_select4_2.get()[-2:-1]=="하":
                    extra_bstat=(extra_bstat/100+1)*1.02*100-100;extra_cper=(extra_cper/100+1)*1.02*100-100
            elif inv_select4_1.get()=="축앞뎀%/1각" and inv2_on_tg==1:
                if inv_select4_2.get()[-2:-1]=="상":
                    extra_batt=(extra_batt/100+1)*1.03*100-100;extra_cstat=extra_cstat+25
                elif inv_select4_2.get()[-2:-1]=="중":
                    extra_batt=(extra_batt/100+1)*1.02*100-100;extra_cstat=extra_cstat+25
                elif inv_select4_2.get()[-2:-1]=="하":
                    extra_batt=(extra_batt/100+1)*1.01*100-100;extra_cstat=extra_cstat+25
            elif inv_select4_1.get()=="축앞뎀%/1각%" and inv2_on_tg==1:
                if inv_select4_2.get()[-2:-1]=="상":
                    extra_batt=(extra_batt/100+1)*1.02*100-100;extra_cper=(extra_cper/100+1)*1.03*100-100
                elif inv_select4_2.get()[-2:-1]=="중":
                    extra_batt=(extra_batt/100+1)*1.02*100-100;extra_cper=(extra_cper/100+1)*1.02*100-100
                elif inv_select4_2.get()[-2:-1]=="하":
                    extra_batt=(extra_batt/100+1)*1.02*100-100;extra_cper=(extra_cper/100+1)*1.01*100-100
            elif inv_select4_1.get()=="전직패" and inv2_on_tg==1:
                if inv_select4_2.get()[-2:-1]=="상":
                    extra_stat=extra_stat+145
                elif inv_select4_2.get()[-2:-1]=="중":
                    extra_stat=extra_stat+115
                elif inv_select4_2.get()[-2:-1]=="하":
                    extra_stat=extra_stat+85
            elif inv_select4_1.get()=="축+1/1각" and inv2_on_tg==1:
                if inv_select4_2.get()[-2:-1]=="상":
                    extra_cstat=extra_cstat+30;extra_blvl=extra_blvl+1
                elif inv_select4_2.get()[-2:-1]=="중":
                    extra_cstat=extra_cstat+20;extra_blvl=extra_blvl+1
                elif inv_select4_2.get()[-2:-1]=="하":
                    extra_cstat=extra_cstat+10;extra_blvl=extra_blvl+1
        if inv2_on_tg==0:
            inv2_opt="미충족"
            inv2_val=" X "
            inv4_opt="미충족"
            inv4_val=""
        if self.inv_tg == 0:
            inv1_opt="미부여"
            inv1_val=""
            inv2_opt=""
            inv2_val=""
            inv3_opt="미부여"
            inv3_val=""
            inv4_opt="미부여"
            inv4_val=""
        if self.inv_tg == 2:
            inv3_opt="미부여"
            inv3_val=""
            inv4_opt="미부여"
            inv4_val=""


        self.all_list_list_num = 0
        all_list_list = []

        ##풀셋모드##
        ##########################################################################################################################
        if select_perfect.get()[0:4] == '풀셋모드' or select_perfect.get() == '메타몽풀셋모드':
            active_bang5_0=[];active_bang5_1=[]
            active_bang2_0=[];active_bang3_1_0=[];active_bang3_1_1=[];active_bang3_2=[];active_bang3_3=[] #1:상의(1_1:신화), 2:하의, 3:신발 / 포함 어벨
            active_acc3_0=[];active_acc3_1=[]
            active_spe3_0=[];active_spe3_1=[]
            active_sang3_0=[];active_sang3_1=[];active_sang2=[]
            active_ha3_0=[];active_ha3_1=[];active_ha2_0=[];active_ha2_1=[]
            active_sin3_0=[];active_sin3_1=[];active_sin2_0=[];active_sin2_1=[]

            for i in range(1,36): ##경우의 수 가르기
                if i < 16:
                    if set_num_dict1.get(str(i+100))==5:
                        active_bang5_0.append(str(i+100))
                    if set_num_dict2.get(str(i+100))==5:
                        active_bang5_1.append(str(i+100))
                    if equip_list["13"].count('1'+str(i+300)+'0')==1:
                        if equip_list["14"].count('1'+str(i+400)+'0')==1:
                            active_bang2_0.append(str(i+100))
                            if equip_list["11_0"].count('1'+str(i+100)+'0')==1:
                                active_bang3_1_0.append(str(i+100))
                            if equip_list["11_1"].count('1'+str(i+100)+'1')==1:
                                active_bang3_1_1.append(str(i+100))
                            if equip_list["12"].count('1'+str(i+200)+'0')==1:
                                active_bang3_2.append(str(i+100))
                            if equip_list["15"].count('1'+str(i+500)+'0')==1:
                                active_bang3_3.append(str(i+100))
                else:
                    if set_num_dict1.get(str(i+100))==3:
                        if i < 20:
                           active_acc3_0.append(str(i+100))
                        elif i < 24:
                            active_spe3_0.append(str(i+100))
                        elif i < 28:
                            active_ha3_0.append(str(i+100))
                        elif i < 32:
                            active_sang3_0.append(str(i+100))
                        elif i < 36:
                            active_sin3_0.append(str(i+100))
                    if set_num_dict2.get(str(i+100))==3:
                        if i < 20:
                            active_acc3_1.append(str(i+100))
                        elif i < 24:
                            active_spe3_1.append(str(i+100))
                        elif i < 28:
                            active_ha3_1.append(str(i+100))
                        elif i < 32:
                            active_sang3_1.append(str(i+100))
                        elif i < 36:
                            active_sin3_1.append(str(i+100))
            for i in range(24,36):
                if equip_list["32"].count('3'+str(i+200)+'0')==1:
                    if equip_list["21_0"].count('2'+str(i+100)+'0')==1:
                        active_ha2_0.append(str(i+100))
                    if equip_list["21_1"].count('2'+str(i+100)+'1')==1:
                        active_ha2_1.append(str(i+100))
                elif equip_list["23"].count('2'+str(i+300)+'0')==1:
                    if equip_list["33_0"].count('3'+str(i+300)+'0')==1:
                        active_sin2_0.append(str(i+100))
                    if equip_list["33_1"].count('3'+str(i+300)+'1')==1:
                        active_sin2_1.append(str(i+100))
                elif equip_list["22"].count('2'+str(i+200)+'0')==1:
                    if equip_list["31"].count('3'+str(i+100)+'0')==1:
                        active_sang2.append(str(i+100))


            all_list_before_inv=[]
            all_list_god_before_inv=[]
            ##1. 533 풀셋
            items533=[active_bang5_0,active_acc3_0,active_spe3_0]
            items533_1=[active_bang5_1,active_acc3_0,active_spe3_0]
            items533_2=[active_bang5_0,active_acc3_1,active_spe3_0]
            items533_3=[active_bang5_0,active_acc3_0,active_spe3_1]
            if len(active_bang5_0)!=0 and len(active_acc3_0)!=0 and len(active_spe3_0)!=0:
                case_list=list(itertools.product(*items533))
                all_list_before_inv=all_list_before_inv+calc_fullset.making_cases(case_list,0,1)
            if len(active_bang5_1)!=0 and len(active_acc3_0)!=0 and len(active_spe3_0)!=0:
                case_list=list(itertools.product(*items533_1))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,1,1)
            if len(active_bang5_0)!=0 and len(active_acc3_1)!=0 and len(active_spe3_0)!=0:
                case_list=list(itertools.product(*items533_2))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,2,1)
            if len(active_bang5_0)!=0 and len(active_acc3_0)!=0 and len(active_spe3_1)!=0:
                case_list=list(itertools.product(*items533_3))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,3,1)

            bang_on_dict={} ##신화여부 상관없음
            bang_on_dict1={} ##신화
            bang_on_dict2={} ##노신화
            ##1_2. 32/33 스까셋
            for i in range(1,16):
                temp_bang_on=[0,0,0,0,0]
                temp_bang_on1=[0,0,0,0,0]
                temp_bang_on2=[0,0,0,0,0]
                if equip_list["11"].count('1'+str(100+i)+'0') != 0:
                    temp_bang_on[0]=1
                    temp_bang_on1[0]=1
                if equip_list["11"].count('1'+str(100+i)+'1')!=0:
                    temp_bang_on[0]=1
                    temp_bang_on2[0]=1
                if equip_list["12"].count('1'+str(200+i)+'0')!=0:
                    temp_bang_on[1]=1
                    temp_bang_on1[1]=1
                    temp_bang_on2[1]=1
                if equip_list["13"].count('1'+str(300+i)+'0')!=0:
                    temp_bang_on[2]=1
                    temp_bang_on1[2]=1
                    temp_bang_on2[2]=1
                if equip_list["14"].count('1'+str(400+i)+'0')!=0:
                    temp_bang_on[3]=1
                    temp_bang_on1[3]=1
                    temp_bang_on2[3]=1
                if equip_list["15"].count('1'+str(500+i)+'0')!=0:
                    temp_bang_on[4]=1
                    temp_bang_on1[4]=1
                    temp_bang_on2[4]=1
                bang_on_dict[str(100+i)]=temp_bang_on
                bang_on_dict1[str(100+i)]=temp_bang_on1
                bang_on_dict2[str(100+i)]=temp_bang_on2

            SG32_0=[]
            SG32_1=[]
            tem_tg=0
            for now_bang_dict in [bang_on_dict1,bang_on_dict2]:
                tem_tg=tem_tg+1
                for cases in ([0,1],[0,2],[0,3],[0,4],[1,2],[1,3],[1,4],[2,3],[2,4],[3,4]):
                    not_cases=[0,1,2,3,4]
                    not_cases.remove(cases[0]);not_cases.remove(cases[1])
                    for i in range(1,16):
                        stri=str(100+i)
                        if now_bang_dict.get(stri)[cases[0]]+now_bang_dict.get(stri)[cases[1]]==2:
                              for j in range(1,16):
                                  if i!=j:
                                      strj=str(100+j)
                                      if now_bang_dict.get(strj)[not_cases[0]]+now_bang_dict.get(strj)[not_cases[1]]+now_bang_dict.get(strj)[not_cases[2]]==3:
                                          temp_32=[0,0,0,0,0]
                                          temp_32[cases[0]]=stri;temp_32[cases[1]]=stri
                                          temp_32[not_cases[0]]=strj;temp_32[not_cases[1]]=strj;temp_32[not_cases[2]]=strj
                                          if tem_tg ==1:
                                              SG32_0.append(''.join(temp_32))
                                          elif tem_tg ==2:
                                              SG32_1.append(''.join(temp_32))
            items32_33=[SG32_0,active_acc3_0,active_spe3_0]
            items32_33_1=[SG32_1,active_acc3_0,active_spe3_0]
            items32_33_2=[SG32_0,active_acc3_1,active_spe3_0]
            items32_33_3=[SG32_0,active_acc3_0,active_spe3_1]
            if len(SG32_0)!=0 and len(active_acc3_0)!=0 and len(active_spe3_0)!=0:
                case_list=list(itertools.product(*items32_33))
                all_list_before_inv=all_list_before_inv+calc_fullset.making_cases(case_list,0,6)
            if len(SG32_1)!=0 and len(active_acc3_0)!=0 and len(active_spe3_0)!=0:
                case_list=list(itertools.product(*items32_33_1))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,1,6)
            if len(SG32_0)!=0 and len(active_acc3_1)!=0 and len(active_spe3_0)!=0:
                case_list=list(itertools.product(*items32_33_2))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,2,6)
            if len(SG32_0)!=0 and len(active_acc3_0)!=0 and len(active_spe3_1)!=0:
                case_list=list(itertools.product(*items32_33_3))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,3,6)



            ##2. 표준3332 풀셋
            temp_list2=[]
            temp_list_god2=[]
            items3332=[active_sang3_0,active_ha3_0,active_sin3_0,active_bang2_0]
            items3332_1=[active_sang3_1,active_ha3_0,active_sin3_0,active_bang2_0]
            items3332_2=[active_sang3_0,active_ha3_1,active_sin3_0,active_bang2_0]
            items3332_3=[active_sang3_0,active_ha3_0,active_sin3_1,active_bang2_0]
            if len(active_sang3_0)!=0 and len(active_ha3_0)!=0 and len(active_sin3_0)!=0 and len(active_bang2_0)!=0:
                case_list=list(itertools.product(*items3332))
                all_list_before_inv=all_list_before_inv+calc_fullset.making_cases(case_list,0,2)
            if len(active_sang3_1)!=0 and len(active_ha3_0)!=0 and len(active_sin3_0)!=0 and len(active_bang2_0)!=0:
                case_list=list(itertools.product(*items3332_1))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,1,2)
            if len(active_sang3_0)!=0 and len(active_ha3_1)!=0 and len(active_sin3_0)!=0 and len(active_bang2_0)!=0:
                case_list=list(itertools.product(*items3332_2))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,2,2)
            if len(active_sang3_0)!=0 and len(active_ha3_0)!=0 and len(active_sin3_1)!=0 and len(active_bang2_0)!=0:
                case_list=list(itertools.product(*items3332_3))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,3,2)
            ##3. 변형3332 풀셋
            #0번 상의변형
            temp_list3_0=[]
            temp_list_god3_0=[]
            items2333=[active_sang2,active_ha3_0,active_sin3_0,active_bang3_1_0]
            items2333_1=[active_sang2,active_ha3_0,active_sin3_0,active_bang3_1_1]
            items2333_2=[active_sang2,active_ha3_1,active_sin3_0,active_bang3_1_0]
            items2333_3=[active_sang2,active_ha3_0,active_sin3_1,active_bang3_1_0]
            if len(active_sang2)!=0 and len(active_ha3_0)!=0 and len(active_sin3_0)!=0 and len(active_bang3_1_0)!=0:
                case_list=list(itertools.product(*items2333))
                all_list_before_inv=all_list_before_inv+calc_fullset.making_cases(case_list,0,3)
            if len(active_sang2)!=0 and len(active_ha3_0)!=0 and len(active_sin3_0)!=0 and len(active_bang3_1_1)!=0:
                case_list=list(itertools.product(*items2333_1))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,1,3)
            if len(active_sang2)!=0 and len(active_ha3_1)!=0 and len(active_sin3_0)!=0 and len(active_bang3_1_0)!=0:
                case_list=list(itertools.product(*items2333_2))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,2,3)
            if len(active_sang2)!=0 and len(active_ha3_0)!=0 and len(active_sin3_1)!=0 and len(active_bang3_1_0)!=0:
                case_list=list(itertools.product(*items2333_3))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,3,3)
            #1번 하의변형
            temp_list3_1=[]
            temp_list_god3_1=[]
            items3233=[active_sang3_0,active_ha2_0,active_sin3_0,active_bang3_2]
            items3233_1=[active_sang3_1,active_ha2_0,active_sin3_0,active_bang3_2]
            items3233_2=[active_sang3_0,active_ha2_1,active_sin3_0,active_bang3_2]
            items3233_3=[active_sang3_0,active_ha2_0,active_sin3_1,active_bang3_2]
            if len(active_sang3_0)!=0 and len(active_ha2_0)!=0 and len(active_sin3_0)!=0 and len(active_bang3_2)!=0:
                case_list=list(itertools.product(*items3233))
                all_list_before_inv=all_list_before_inv+calc_fullset.making_cases(case_list,0,4)
            if len(active_sang3_1)!=0 and len(active_ha2_0)!=0 and len(active_sin3_0)!=0 and len(active_bang3_2)!=0:
                case_list=list(itertools.product(*items3233_1))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,1,4)
            if len(active_sang3_0)!=0 and len(active_ha2_1)!=0 and len(active_sin3_0)!=0 and len(active_bang3_2)!=0:
                case_list=list(itertools.product(*items3233_2))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,2,4)
            if len(active_sang3_0)!=0 and len(active_ha2_0)!=0 and len(active_sin3_1)!=0 and len(active_bang3_2)!=0:
                case_list=list(itertools.product(*items3233_3))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,3,4)
            #2번 신발변형
            temp_list3_2=[]
            temp_list_god3_2=[]
            items3323=[active_sang3_0,active_ha3_0,active_sin2_0,active_bang3_3]
            items3323_1=[active_sang3_1,active_ha3_0,active_sin2_0,active_bang3_3]
            items3323_2=[active_sang3_0,active_ha3_1,active_sin2_0,active_bang3_3]
            items3323_3=[active_sang3_0,active_ha3_0,active_sin2_1,active_bang3_3]
            if len(active_sang3_0)!=0 and len(active_ha3_0)!=0 and len(active_sin2_0)!=0 and len(active_bang3_3)!=0:
                case_list=list(itertools.product(*items3323))
                all_list_before_inv=all_list_before_inv+calc_fullset.making_cases(case_list,0,5)
            if len(active_sang3_1)!=0 and len(active_ha3_0)!=0 and len(active_sin2_0)!=0 and len(active_bang3_3)!=0:
                case_list=list(itertools.product(*items3323_1))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,1,5)
            if len(active_sang3_0)!=0 and len(active_ha3_1)!=0 and len(active_sin2_0)!=0 and len(active_bang3_3)!=0:
                case_list=list(itertools.product(*items3323_2))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,2,5)
            if len(active_sang3_0)!=0 and len(active_ha3_0)!=0 and len(active_sin2_1)!=0 and len(active_bang3_3)!=0:
                case_list=list(itertools.product(*items3323_3))
                all_list_god_before_inv=all_list_god_before_inv+calc_fullset.making_cases(case_list,3,5)

            all_list=[];all_list_god=[]
            for i in list40_0:
                for j in all_list_before_inv:
                    tempx=list(j)
                    tempx.append(i)
                    all_list.append(tuple(tempx))
                for j in all_list_god_before_inv:
                    tempx=list(j)
                    tempx.append(i)
                    all_list_god.append(tuple(tempx))

            fullseton=0
            all_list_num=len(all_list_god)+len(all_list)
            if all_list_num==0 and select_perfect.get() != '메타몽풀셋모드':
                tkinter.messagebox.showerror('에러',"풀셋이 없습니다.")
                self.change_state_text(text='중지됨')
                return
            if all_list_num==0 and select_perfect.get() == '메타몽풀셋모드':
                pass
            else:
                fullseton=1
                all_list_list.append([all_list,all_list_god,all_list_num])

            if select_perfect.get() == '메타몽풀셋모드':
                self.change_state_text(text='메타몽 계산중(오래걸릴수 있음)')
                evert_list=equip_list["11"]+equip_list["12"]+equip_list["13"]+equip_list["14"]+equip_list["15"]+equip_list["21"]+equip_list["22"]+equip_list["23"]+equip_list["31"]+equip_list["32"]+equip_list["33"]
                result_metamong=calc_fullset.meta_ful(set_num_dict,evert_list,bang_on_dict,list40_0)
                if result_metamong[2]!=0:
                    all_list_list.append(result_metamong)
                elif fullseton==1:
                    pass
                else:
                    tkinter.messagebox.showerror('에러',"풀셋이 없습니다.")
                    self.change_state_text(text='중지됨')
                    return

        else:
        ##레전기본값##
        ##########################################################################################################################
            if self.default_base_equip == 0:
                df11='11360';df12='12360';df13='13360';df14='14360';df15='15360'
                df21='21370';df22='22370';df23='23370';df31='31380';df32='32380';df33='33380'
            elif self.default_base_equip == 1:
                df11='11440';df12='12440';df13='13440';df14='14440';df15='15440'
                df21='21450';df22='22450';df23='23450';df31='31460';df32='32460';df33='33460'
            elif self.default_base_equip == 2:
                df11='11470';df12='12470';df13='13470';df14='14470';df15='15470'
                df21='21480';df22='22480';df23='23480';df31='31490';df32='32490';df33='33490'
            if self.legend_on_tg.get() == 1:
                if len(equip_list["11_0"])==0 or len(equip_list["12"])==0 or len(equip_list["13"])==0 or len(equip_list["14"])==0 or len(equip_list["15"])==0 or max(set_max_list1) < 3:
                    equip_list["11"].append(df11);equip_list["12"].append(df12);equip_list["13"].append(df13);equip_list["14"].append(df14);equip_list["15"].append(df15);equip_list["11_0"].append(df11)
                if len(equip_list["21_0"])==0 or len(equip_list["22"])==0 or len(equip_list["23"])==0 or max(set_max_list2) < 3:
                    equip_list["21"].append(df21);equip_list["22"].append(df22);equip_list["23"].append(df23);equip_list["21_0"].append(df21)
                if len(equip_list["31"])==0 or len(equip_list["32"])==0 or len(equip_list["33_0"])==0 or max(set_max_list3) < 3:
                    equip_list["31"].append(df31);equip_list["32"].append(df32);equip_list["33"].append(df33);equip_list["33_0"].append(df33)

            if len(equip_list["11_0"])==0:
                equip_list["11"].append(df11);equip_list["11_0"].append(df11)
            if len(equip_list["12"])==0:
                equip_list["12"].append(df12)
            if len(equip_list["13"])==0:
                equip_list["13"].append(df13)
            if len(equip_list["14"])==0:
                equip_list["14"].append(df14)
            if len(equip_list["15"])==0:
                equip_list["15"].append(df15)
            if len(equip_list["21_0"])==0:
                equip_list["21"].append(df21);equip_list["21_0"].append(df21)
            if len(equip_list["22"])==0:
                equip_list["22"].append(df22)
            if len(equip_list["23"])==0:
                equip_list["23"].append(df23)
            if len(equip_list["31"])==0:
                equip_list["31"].append(df31)
            if len(equip_list["32"])==0:
                equip_list["32"].append(df32)
            if len(equip_list["33_0"])==0:
                equip_list["33"].append(df33);equip_list["33_0"].append(df33)





        ##세트산물 계산##
        #########################################################################################################################
            know_bang1_list=['22400150','22400250','22400350','22400450','22400550']
            know_bang2_list=['31400750','31400850','31400950','31401050','31401150']
            know_acc_list=['32401240','32401340','32401440']

            for i in self.know_set_list: ##경우1:산물 하나
                if select_item['tg'+i]==1:
                    if int(i[4:6]) <6:
                        items0=[['99990'],['99990'],['99990'],['99990'],['99990'],equip_list["21_0"],[i],equip_list["23"],equip_list["31"],equip_list["32"],equip_list["33_0"],list40_0]
                        items1=[]
                        items2=[['99990'],['99990'],['99990'],['99990'],['99990'],equip_list["21_1"],[i],equip_list["23"],equip_list["31"],equip_list["32"],equip_list["33_0"],list40_0]
                        items3=[['99990'],['99990'],['99990'],['99990'],['99990'],equip_list["21_0"],[i],equip_list["23"],equip_list["31"],equip_list["32"],equip_list["33_1"],list40_0]
                    elif int(i[4:6])==6:
                        items0=[equip_list["11_0"],equip_list["12"],equip_list["13"],equip_list["14"],equip_list["15"],[i],equip_list["22"],equip_list["23"],['99990'],['99990'],['99990'],list40_0]
                        items1=[equip_list["11_1"],equip_list["12"],equip_list["13"],equip_list["14"],equip_list["15"],[i],equip_list["22"],equip_list["23"],['99990'],['99990'],['99990'],list40_0]
                        items2=[]
                        items3=[]
                    elif int(i[4:6]) <12:
                        items0=[['99990'],['99990'],['99990'],['99990'],['99990'],equip_list["21_0"],equip_list["22"],equip_list["23"],[i],equip_list["32"],equip_list["33_0"],list40_0]
                        items1=[]
                        items2=[['99990'],['99990'],['99990'],['99990'],['99990'],equip_list["21_1"],equip_list["22"],equip_list["23"],[i],equip_list["32"],equip_list["33_0"],list40_0]
                        items3=[['99990'],['99990'],['99990'],['99990'],['99990'],equip_list["21_0"],equip_list["22"],equip_list["23"],[i],equip_list["32"],equip_list["33_1"],list40_0]
                    elif int(i[4:6]) <15:
                        items0=[equip_list["11_0"],equip_list["12"],equip_list["13"],equip_list["14"],equip_list["15"],equip_list["31"],[i],equip_list["33_0"],['99990'],['99990'],['99990'],list40_0]
                        items1=[equip_list["11_1"],equip_list["12"],equip_list["13"],equip_list["14"],equip_list["15"],equip_list["31"],[i],equip_list["33_0"],['99990'],['99990'],['99990'],list40_0]
                        items2=[]
                        items3=[equip_list["11_0"],equip_list["12"],equip_list["13"],equip_list["14"],equip_list["15"],equip_list["31"],[i],equip_list["33_1"],['99990'],['99990'],['99990'],list40_0]
                    all_list_list.append(calc_setlist.make_list(items0,items1,items2,items3))

            know_bang1_on=[]
            for i in know_bang1_list:
                if select_item['tg'+i]==1:
                    know_bang1_on.append(i)
            if select_item['tg21400640']==1:  ##경우2:만유(팔찌)+방어구(목걸이)
                items0=[['99990'],['99990'],['99990'],['99990'],['99990'],['21400640'],know_bang1_on,equip_list["23"],['99990'],['99990'],['99990'],list40_0]
                items1=[]
                items2=[]
                items3=[]
                all_list_list.append(calc_setlist.make_list(items0,items1,items2,items3))

            know_acc_on=[]
            know_bang2_on=[]
            for i in know_acc_list:
                if select_item['tg'+i]==1:
                    know_acc_on.append(i)
            for i in know_bang2_list:
                if select_item['tg'+i]==1:
                    know_bang2_on.append(i)
            if len(know_acc_list)!=0 and len(know_bang2_on)!=0:  ##경우3: 악세(법석)+방어구(보장)
                items0=[['99990'],['99990'],['99990'],['99990'],['99990'],['99990'],['99990'],['99990'],know_bang2_on,know_acc_on,equip_list["33_0"],list40_0]
                items1=[]
                items2=[]
                items3=[['99990'],['99990'],['99990'],['99990'],['99990'],['99990'],['99990'],['99990'],know_bang2_on,know_acc_on,equip_list["33_1"],list40_0]
                all_list_list.append(calc_setlist.make_list(items0,items1,items2,items3))

            jin_sang=[]
            jin_pal=[]
            jin_gui=[]
            for i in self.know_jin_list:
                if select_item[f"tg{i}"] == 1:
                    if i[0:2]=='11':
                        jin_sang.append(i)
                    elif i[0:2]=='21':
                        jin_pal.append(i)
                    elif i[0:2]=='33':
                        jin_gui.append(i)

            for i in self.know_jin_list:
                if select_item['tg'+i]==1: ##경우4: 진레전산물
                    if i[0:2]=='11': ##상의만
                        items0=[[i],['12410'],['13410'],['14410'],['15410'],equip_list["21_0"],equip_list["22"],equip_list["23"],equip_list["31"],equip_list["32"],equip_list["33_0"],list40_0]
                        items1=[]
                        items2=[[i],['12410'],['13410'],['14410'],['15410'],equip_list["21_1"],equip_list["22"],equip_list["23"],equip_list["31"],equip_list["32"],equip_list["33_0"],list40_0]
                        items3=[[i],['12410'],['13410'],['14410'],['15410'],equip_list["21_0"],equip_list["22"],equip_list["23"],equip_list["31"],equip_list["32"],equip_list["33_1"],list40_0]
                        all_list_list.append(calc_setlist.make_list(items0,items1,items2,items3))
                    if i[0:2]=='21': ##팔찌만
                        items0=[equip_list["11_0"],equip_list["12"],equip_list["13"],equip_list["14"],equip_list["15"],[i],['22420'],['23420'],equip_list["31"],equip_list["32"],equip_list["33_0"],list40_0]
                        items1=[equip_list["11_1"],equip_list["12"],equip_list["13"],equip_list["14"],equip_list["15"],[i],['22420'],['23420'],equip_list["31"],equip_list["32"],equip_list["33_0"],list40_0]
                        items2=[]
                        items3=[equip_list["11_0"],equip_list["12"],equip_list["13"],equip_list["14"],equip_list["15"],[i],['22420'],['23420'],equip_list["31"],equip_list["32"],equip_list["33_1"],list40_0]
                        all_list_list.append(calc_setlist.make_list(items0,items1,items2,items3))
                    if i[0:2]=='33': ##귀걸만
                        items0=[equip_list["11_0"],equip_list["12"],equip_list["13"],equip_list["14"],equip_list["15"],equip_list["21_0"],equip_list["22"],equip_list["23"],['31430'],['32430'],[i],list40_0]
                        items1=[equip_list["11_1"],equip_list["12"],equip_list["13"],equip_list["14"],equip_list["15"],equip_list["21_0"],equip_list["22"],equip_list["23"],['31430'],['32430'],[i],list40_0]
                        items2=[equip_list["11_0"],equip_list["12"],equip_list["13"],equip_list["14"],equip_list["15"],equip_list["21_1"],equip_list["22"],equip_list["23"],['31430'],['32430'],[i],list40_0]
                        items3=[]
                        all_list_list.append(calc_setlist.make_list(items0,items1,items2,items3))

            if len(jin_sang)!=0 and len(jin_pal)!=0: ##상의+팔찌
                items0=[jin_sang,['12410'],['13410'],['14410'],['15410'],jin_pal,['22420'],['23420'],equip_list["31"],equip_list["32"],equip_list["33_0"],list40_0]
                items1=[]
                items2=[]
                items3=[jin_sang,['12410'],['13410'],['14410'],['15410'],jin_pal,['22420'],['23420'],equip_list["31"],equip_list["32"],equip_list["33_1"],list40_0]
                all_list_list.append(calc_setlist.make_list(items0,items1,items2,items3))
            if len(jin_sang)!=0 and len(jin_gui)!=0: ##상의+귀걸
                items0=[jin_sang,['12410'],['13410'],['14410'],['15410'],equip_list["21_0"],equip_list["22"],equip_list["23"],['31430'],['32430'],jin_gui,list40_0]
                items1=[]
                items2=[jin_sang,['12410'],['13410'],['14410'],['15410'],equip_list["21_1"],equip_list["22"],equip_list["23"],['31430'],['32430'],jin_gui,list40_0]
                items3=[]
                all_list_list.append(calc_setlist.make_list(items0,items1,items2,items3))
            if len(jin_pal)!=0 and len(jin_gui)!=0: ##팔찌+귀걸
                items0=[equip_list["11_0"],equip_list["12"],equip_list["13"],equip_list["14"],equip_list["15"],jin_pal,['22420'],['23420'],['31430'],['32430'],jin_gui,list40_0]
                items1=[equip_list["11_1"],equip_list["12"],equip_list["13"],equip_list["14"],equip_list["15"],jin_pal,['22420'],['23420'],['31430'],['32430'],jin_gui,list40_0]
                items2=[]
                items3=[]
                all_list_list.append(calc_setlist.make_list(items0,items1,items2,items3))
            if len(jin_sang)!=0 and len(jin_pal)!=0 and len(jin_gui)!=0: ##3개 전부
                items0=[jin_sang,['12410'],['13410'],['14410'],['15410'],jin_pal,['22420'],['23420'],['31430'],['32430'],jin_gui,list40_0]
                items1=[]
                items2=[]
                items3=[]
                all_list_list.append(calc_setlist.make_list(items0,items1,items2,items3))



        ##일반 경우의 수##
        #########################################################################################################################
            timp_list_num=0
            for i in all_list_list:
                timp_list_num=timp_list_num+int(i[2])

            items0=[equip_list["11_0"],equip_list["12"],equip_list["13"],equip_list["14"],equip_list["15"],equip_list["21_0"],equip_list["22"],equip_list["23"],equip_list["31"],equip_list["32"],equip_list["33_0"],list40_0]
            items1=[equip_list["11_1"],equip_list["12"],equip_list["13"],equip_list["14"],equip_list["15"],equip_list["21_0"],equip_list["22"],equip_list["23"],equip_list["31"],equip_list["32"],equip_list["33_0"],list40_0]
            items2=[equip_list["11_0"],equip_list["12"],equip_list["13"],equip_list["14"],equip_list["15"],equip_list["21_1"],equip_list["22"],equip_list["23"],equip_list["31"],equip_list["32"],equip_list["33_0"],list40_0]
            items3=[equip_list["11_0"],equip_list["12"],equip_list["13"],equip_list["14"],equip_list["15"],equip_list["21_0"],equip_list["22"],equip_list["23"],equip_list["31"],equip_list["32"],equip_list["33_1"],list40_0]
            all_list_list.append(calc_setlist.make_list(items0,items1,items2,items3))

        #########################################################################################################################

        for i in all_list_list:
            self.all_list_list_num += int(i[2])

        self.all_list_list_num *= len(wep_num)

        if self.all_list_list_num > 500000000:
            tkinter.messagebox.showerror('에러',"경우의 수가 5억가지가 넘습니다.\n진행이 불가능합니다.\n안 쓸 에픽 체크를 풀어주세요")
            self.change_state_text(text='중지됨')
            return
        elif self.all_list_list_num > 100000000:
            ask_msg2=tkinter.messagebox.askquestion('확인',"경우의 수가 1억가지가 넘습니다.\n메모리 과부하가 날 수 있고 30분이상 걸릴 수 있습니다.\n강행으로 인한 PC파손은 책임지지 않습니다.\n진행하시겠습니까?")
            if ask_msg2 == 'no':
                self.change_state_text(text='중지됨')
                return
        elif self.all_list_list_num > 30000000:
            ask_msg2=tkinter.messagebox.askquestion('확인',"경우의 수가 3천만가지가 넘습니다.\n다소 오래 걸릴 수 있습니다.\n강행으로 인한 PC파손은 책임지지 않습니다.\n진행하시겠습니까?")
            if ask_msg2 == 'no':
                self.change_state_text(text='중지됨')
                return
        if set_perfect ==1 and self.all_list_list_num > 30000000:
            tkinter.messagebox.showerror('에러',"정확도 높음 기능은 많은 경우의 수를 지원하지 않습니다.")
            self.change_state_text(text='중지됨')
            return
        if set_perfect !=1 and self.all_list_list_num < 10000:
            set_perfect=1


        #########################################################################################################################계산 시작

        self.change_state_text(text='계산 시작')
        
        save_list = {}  #딜러1번
        save_list0 = {}  #딜러2번
        save_list1 = {}  #버퍼1번
        save_list2 = {}  #버퍼2번
        save_list3 = {}  #버퍼3번

        job_name: str = jobup_select.get()[4:7]
        
        all_list_list = reduce(lambda a, b: (a[0] + b[0], a[1] + b[1]), all_list_list)
        normal_equip_list = all_list_list[0]
        god_equip_list = all_list_list[1]
        all_equip_cases = [(val, True) for val in god_equip_list] + [(val, False) for val in normal_equip_list]

        if select_perfect.get()[0:4] == '풀셋모드' or select_perfect.get() == '메타몽풀셋모드':
            num_thread = 1
        else:
            num_thread = os.cpu_count()

        count_per_thread = len(all_equip_cases) // num_thread
        leftover_per_thread = len(all_equip_cases) % num_thread

        if count_per_thread == 0:
            num_thread = 1
            all_list_per_thread = [all_equip_cases]
        elif num_thread == 1:
            all_list_per_thread = [all_equip_cases]
        else:
            all_list_per_thread = [all_equip_cases[i * count_per_thread:(i + 1) * count_per_thread] for i in range(num_thread)]

            for i in range(leftover_per_thread):
                all_list_per_thread[i].append(all_equip_cases[num_thread * count_per_thread + i])

        common_args = {
            "job_name": job_name,
            "opt_one": self.opt_one,
            "opt_buf": self.opt_buf,
            "opt_buflvl": self.opt_buflvl,
            "inv_tg": self.inv_tg,
            "inv_type_list": self.inv_type_list,
            "set_perfect": set_perfect,
            "wep_num": wep_num,
            "extra_dam": extra_dam,
            "extra_cri": extra_cri,
            "extra_bon": extra_bon,
            "extra_all": extra_all,
            "extra_att": extra_att,
            "extra_sta": extra_sta,
            "ele_in": ele_in,
            "fixed_dam": fixed_dam,
            "fixed_cri": fixed_cri,
            "extra_pas2": extra_pas2,
            "cool_eff_dictnum": cool_eff_dictnum,
            "betterang": betterang,

            "active_eff_one": active_eff_one,
            "job_lv1": job_lv1,
            "job_lv2": job_lv2,
            "job_lv3": job_lv3,
            "job_lv4": job_lv4,
            "job_lv5": job_lv5,
            "job_lv6": job_lv6,
            "silmari": silmari,
            "job_pas0": job_pas0,
            "job_pas1": job_pas1,
            "job_pas2": job_pas2,
            "job_pas3": job_pas3,
            "inv2_on_tg": inv2_on_tg,
            "job_ult1": job_ult1,
            "job_ult2": job_ult2,
            "job_ult3": job_ult3,
            "ele_skill": ele_skill,
            "wep_pre_calced": wep_pre_calced,
            "cool_eff": cool_eff,
            "cool_eff2": cool_eff2,
            "cool_on": cool_on,
            "cool_pre_calced": cool_pre_calced,
            "cool_pre_calced2": cool_pre_calced2,
            "wep_name_list_temp": wep_name_list_temp,
            "db_preset": db_preset,
            "extra_clvl": extra_clvl,
            "extra_stat": extra_stat,
            "inv1_opt": inv1_opt,
            "inv1_val": inv1_val,
            "inv2_opt": inv2_opt,
            "inv2_val": inv2_val,
            "inv3_opt": inv3_opt,
            "inv3_val": inv3_val,
            "inv4_opt": inv4_opt,
            "inv4_val": inv4_val,
            "extra_blvl": extra_blvl,
            "extra_cstat": extra_cstat,
            "extra_bstat": extra_bstat,
            "extra_batt": extra_batt,
            "extra_cper": extra_cper
        }

        self.exit_calc.value = 0
        self.count_num.value = 0
        self.count_all.value = 0

        for calc_wep_num in range(len(wep_num)):
            if num_thread == 1:
                init_global(self.exit_calc, self.count_num, self.count_all)

                calc_result = calc_per_thread(all_list_per_thread[0], 1, calc_wep_num, **common_args)

                local_save_list, local_save_list0, local_save_list1, local_save_list2, local_save_list3 = calc_result

                save_list.update(local_save_list)
                save_list0.update(local_save_list0)
                save_list1.update(local_save_list1)
                save_list2.update(local_save_list2)
                save_list3.update(local_save_list3)
            else:
                with ProcessPoolExecutor(max_workers=num_thread, initializer=init_global, initargs=(self.exit_calc, self.count_num, self.count_all)) as executor:
                    futures: List[Future] = []

                    print(f"Calc start: {num_thread} threads")

                    for i, list_per_thread in enumerate(all_list_per_thread):
                        fut: Future = executor.submit(calc_per_thread, list_per_thread, i + 1, calc_wep_num, **common_args)
                        futures.append(fut)

                    print(f"Thread spawn complete, calculating...")

                    executor.shutdown()

                    print(f"All calculations done")

                for fut in as_completed(futures):
                    if self.exit_calc.value == 1:
                        self.change_state_text(text='중지됨')
                        return

                    thread_save_list, thread_save_list0, thread_save_list1, thread_save_list2, thread_save_list3 = fut.result()

                    save_list.update(thread_save_list)
                    save_list0.update(thread_save_list0)
                    save_list1.update(thread_save_list1)
                    save_list2.update(thread_save_list2)
                    save_list3.update(thread_save_list3)

        ###### 결과 순위 매기기 #######################################################################################
        self.show_number = 0

        if jobup_select.get()[4:7] != "세인트" and jobup_select.get()[4:7] != "세라핌" and jobup_select.get()[4:7] != "헤카테":
            self.change_state_text(text='결과 집계중')

            ranking=[]
            ranking0=[]
            for j in range(0,5):
                try:
                    now_max=max(list(save_list.keys()))
                    ranking.append((now_max,save_list.get(now_max)))
                    del save_list[now_max]
                except ValueError as error:
                    pass
            for j in range(0,5):
                try:
                    now_max=max(list(save_list0.keys()))
                    ranking0.append((now_max,save_list0.get(now_max)))
                    del save_list0[now_max]
                except ValueError as error:
                    pass
            ranking=[ranking,ranking0]
            self.show_result(ranking,'deal',ele_skill,cool_eff)
        else: ##버퍼
            self.change_state_text(text='결과 집계중')

            ranking1=[];ranking2=[];ranking3=[]
            for j in range(0,5):
                try:
                    now_max1=max(list(save_list1.keys()))
                    ranking1.append((now_max1,save_list1.get(now_max1)))
                    del save_list1[now_max1]
                except ValueError as error:
                    pass
            for j in range(0,5):
                try:
                    now_max2=max(list(save_list2.keys()))
                    ranking2.append((now_max2,save_list2.get(now_max2)))
                    del save_list2[now_max2]
                except ValueError as error:
                    pass

            for j in range(0,5):
                try:
                    now_max3=max(list(save_list3.keys()))
                    ranking3.append((now_max3,save_list3.get(now_max3)))
                    del save_list3[now_max3]
                except ValueError as error:
                    pass
            ranking=[ranking1,ranking2,ranking3]
            self.show_result(ranking,'buf',ele_skill,cool_eff)

        load_presetc.close()
        load_excel.close()

        self.change_state_text(text='출력 완료')
        print("걸린 시간 = "+str(time.time() - start_time)+"초")


    def show_result(self, rank_list,job_type,ele_skill,cool_eff):
        jobup_select = self.jobup_select

        result_window = self.result_window = tkinter.Toplevel(self.window)
        result_window.attributes("-topmost", True)
        result_window.geometry("585x402")
        result_window.title("결과값")
        result_window.resizable(False,False)
        result_window.configure(bg=self.dark_main)
        canvas_res = self.canvas_res = Canvas(result_window, width=587, height=804, bd=0, bg=self.dark_main)
        canvas_res.place(x=-2,y=-2)
        if job_type=='deal':
            result_bg=self.get_photo_image('ext_img/bg_result.png')
        else:
            result_bg=self.get_photo_image('ext_img/bg_result2.png')
        canvas_res.create_image(0,0,image=result_bg,anchor='nw')
        random_npc_img=self.get_photo_image('ext_img/bg_result_'+random.choice(['1','2'])+'.png')
        random_npc=canvas_res.create_image(313-210,370,image=random_npc_img,anchor='nw')

        image_list = self.image_list
        image_list_wep = self.image_list_wep

        self.now_rank_num = 0
        self.set_name_toggle=0
        self.pause_gif = 0
        self.stop_gif = 0
        self.stop_gif2 = 0
        job_name=jobup_select.get()[:-4]
        job_up_name=jobup_select.get()[-4:]
        canvas_res.create_text(122,50,text="<직업>",font=self.guide_font,fill='white')
        canvas_res.create_text(122,70,text=job_name,font=self.guide_font,fill='white')
        canvas_res.create_text(122,87,text=job_up_name,font=self.guide_font,fill='white')

        ele_change_toggle=0
        rank_setting = self.rank_setting = [0, 0, 0, 0, 0]
        rank0_setting = self.rank0_setting = [0, 0, 0, 0, 0]
        rank_wep_name = self.deal_rank_wep_name[0] = [0, 0, 0, 0, 0]
        rank0_wep_name = self.deal_rank_wep_name[1] = [0, 0, 0, 0, 0]

        rank_ult=[0,0,0,0,0];rank0_ult=[0,0,0,0,0]
        if job_type=='deal': ########################### 딜러 ###########################
            req_cool = self.req_cool

            result_image_on = self.deal_result_image_on[0]
            result0_image_on = self.deal_result_image_on[1]

            rank_wep_img = self.deal_rank_wep_img[0]
            rank0_wep_img = self.deal_rank_wep_img[1]

            result_image_tag = self.deal_result_image_tag[0]
            result0_image_tag = self.deal_result_image_tag[1]

            rank_stat = self.deal_rank_stat1[0]
            rank_stat2 = self.deal_rank_stat2[0]
            rank_stat3 = self.deal_rank_stat3[0]

            rank0_stat = self.deal_rank_stat1[1]
            rank0_stat2 = self.deal_rank_stat2[1]
            rank0_stat3 = self.deal_rank_stat3[1]

            cool_check=req_cool.get()[0]
            if cool_check=='O':
                self.cool_eff_text = '(쿨감O)'
                cool_eff_text_all='그로기(쿨감O)'
            else:
                self.cool_eff_text = '(쿨감X)'
                cool_eff_text_all='그로기(쿨감X)'

            self.res_cool_what = canvas_res.create_text(122,114,text=cool_eff_text_all,font=self.small_font,fill='white')

            rank0_list = self.rank_list[0] = rank_list[1]
            rank1_list = self.rank_list[1] = rank_list[0]

            rank_dam = self.rank_dam[0] = [0, 0, 0, 0, 0]
            rank0_dam = self.rank_dam[1] = [0, 0, 0, 0, 0]

            rank_dam_tagk = self.rank_dam_tagk[0] = [0, 0, 0, 0, 0]
            rank0_dam_tagk = self.rank_dam_tagk[1] = [0, 0, 0, 0, 0]

            rank_dam_nolv = self.rank_dam_nolv[0] = [0, 0, 0, 0, 0]
            rank0_dam_nolv = self.rank_dam_nolv[1] = [0, 0, 0, 0, 0]

            rank_dam_noele = self.rank_dam_noele[0] = [0, 0, 0, 0, 0]
            rank0_dam_noele = self.rank_dam_noele[1] = [0, 0, 0, 0, 0]

            rank_inv = self.deal_rank_inv[0] = [0, 0, 0, 0, 0]
            rank0_inv = self.deal_rank_inv[1] = [0, 0, 0, 0, 0]

            rank_dam_tagk_nolv = self.rank_dam_tagk_nolv[0] = [0, 0, 0, 0, 0]
            rank0_dam_tagk_nolv = self.rank_dam_tagk_nolv[1] = [0, 0, 0, 0, 0]
            rank_dam_tagk_noele = self.rank_dam_tagk_noele[0] = [0, 0, 0, 0, 0]
            rank0_dam_tagk_noele = self.rank_dam_tagk_noele[1] = [0, 0, 0, 0, 0]
            rank_dam_onlyequ = [0, 0, 0, 0, 0]
            rank0_dam_onlyequ = [0, 0, 0, 0, 0]

            rss=[0,0,0,0,0];rss0=[0,0,0,0,0]

            rank_wep_img.clear()
            rank0_wep_img.clear()
            rank_wep_img.extend([0, 0, 0, 0, 0])
            rank0_wep_img.extend([0, 0, 0, 0, 0])

            deal_result_image_gif = self.deal_result_image_gif[0] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            deal_result0_image_gif = self.deal_result_image_gif[1] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            deal_result_image_gif_tg = self.deal_result_image_gif_tg[0] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            deal_result0_image_gif_tg = self.deal_result_image_gif_tg[1] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]

            deal_result_siroco_gif = self.deal_result_siroco_gif[0] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            deal_result0_siroco_gif = self.deal_result_siroco_gif[1] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            deal_result_siroco_gif_tg = self.deal_result_siroco_gif_tg[0] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            deal_result0_siroco_gif_tg = self.deal_result_siroco_gif_tg[1] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]

            result_image_on.clear()
            result0_image_on.clear()
            result_image_on.extend([{}, {}, {}, {}, {}])
            result0_image_on.extend([{}, {}, {}, {}, {}])
            result_image_tag.clear()
            result0_image_tag.clear()
            result_image_tag.extend([{}, {}, {}, {}, {}])
            result0_image_tag.extend([{}, {}, {}, {}, {}])
            for temp_rank in range(0,5):
                try:
                    rank_dam[temp_rank]=str(int(100*rank1_list[temp_rank][0]))+"%";rank0_dam[temp_rank]=str(int(100*rank0_list[temp_rank][0]))+"%";
                    rank_dam_nolv[temp_rank]=int(100*rank1_list[temp_rank][1][2]);rank0_dam_nolv[temp_rank]=int(100*rank0_list[temp_rank][1][2]);
                    rank_dam_noele[temp_rank]=int(100*rank1_list[temp_rank][1][3]);rank0_dam_noele[temp_rank]=int(100*rank0_list[temp_rank][1][3]);
                    rank_inv[temp_rank]=rank1_list[temp_rank][1][4];rank0_inv[temp_rank]=rank0_list[temp_rank][1][4];
                    rank_ult[temp_rank]=rank1_list[temp_rank][1][5];rank0_ult[temp_rank]=rank0_list[temp_rank][1][5];
                    rank_setting[temp_rank]=list(rank1_list[temp_rank][1][0]);rank0_setting[temp_rank]=list(rank0_list[temp_rank][1][0]);
                    rank_dam_onlyequ[temp_rank]=rank1_list[temp_rank][1][6];rank0_dam_onlyequ[temp_rank]=rank0_list[temp_rank][1][6];
                    rank_wep_name[temp_rank]=rank1_list[temp_rank][1][8];rank0_wep_name[temp_rank]=rank0_list[temp_rank][1][8]
                    rank_wep_img[temp_rank]=image_list_wep[rank_wep_name[temp_rank]];rank0_wep_img[temp_rank]=image_list_wep[rank0_wep_name[temp_rank]]
                    for i in rank_setting[temp_rank]:
                        if len(i)==4 and i[0]=='4': ## 융합 장비 다시 풀기
                            rank_setting[temp_rank].append('415'+i[1]+'0');
                            rank_setting[temp_rank].append('425'+i[2]+'0');
                            rank_setting[temp_rank].append('435'+i[3]+'0');
                            rank_setting[temp_rank].remove(i);
                    for i in rank0_setting[temp_rank]:
                        if len(i)==4 and i[0]=='4': ## 융합 장비 다시 풀기
                            rank0_setting[temp_rank].append('415'+i[1]+'0');
                            rank0_setting[temp_rank].append('425'+i[2]+'0');
                            rank0_setting[temp_rank].append('435'+i[3]+'0');
                            rank0_setting[temp_rank].remove(i);
                    rss[temp_rank]=rank1_list[temp_rank][1][1];rss0[temp_rank]=rank0_list[temp_rank][1][1]
                    rank_dam_tagk[temp_rank]=str(int((100*rank1_list[temp_rank][0])*(100+rss[temp_rank][6])/(121+rss[temp_rank][6])*1.23))+"%";rank0_dam_tagk[temp_rank]=str(int((100*rank0_list[temp_rank][0])*(100+rss0[temp_rank][6])/(121+rss0[temp_rank][6])*1.23))+"%"
                    rank_dam_tagk_nolv[temp_rank]=int((100+rss[temp_rank][6])/(121+rss[temp_rank][6])*1.23*(rank_dam_nolv[temp_rank]));rank0_dam_tagk_nolv[temp_rank]=int((100+rss0[temp_rank][6])/(121+rss0[temp_rank][6])*1.23*(rank0_dam_nolv[temp_rank]))
                    rank_dam_tagk_noele[temp_rank]=int((100+rss[temp_rank][6])/(121+rss[temp_rank][6])*1.23*(rank_dam_noele[temp_rank]));rank0_dam_tagk_noele[temp_rank]=int((100+rss0[temp_rank][6])/(121+rss0[temp_rank][6])*1.23*(rank0_dam_noele[temp_rank]))
                    for i in [11,12,13,14,15,21,22,23,31,32,33,41,42,43]:
                        for j in rank_setting[temp_rank]:
                            if len(j) != 6:
                                if j[0:2] == str(i):
                                    result_image_on[temp_rank][str(i)]=image_list[j]
                                    result_image_tag[temp_rank][str(i)]=j
                                    if i ==11 and j[4:5]=='1' and len(j)==5:
                                        deal_result_image_gif[temp_rank][0]=calc_gif.img_gif(j,0)
                                        deal_result_image_gif_tg[temp_rank][0]=1
                                    if i ==21 and j[4:5]=='1' and len(j)==5:
                                        deal_result_image_gif[temp_rank][1]=calc_gif.img_gif(j,0)
                                        deal_result_image_gif_tg[temp_rank][1]=1
                                    if i ==33 and j[4:5]=='1' and len(j)==5:
                                        deal_result_image_gif[temp_rank][2]=calc_gif.img_gif(j,0)
                                        deal_result_image_gif_tg[temp_rank][2]=1
                                    if i ==41 and j[3:5]!='00' and len(j)==5:
                                        deal_result_siroco_gif[temp_rank][0]=calc_gif.img_gif(j,1)
                                        deal_result_siroco_gif_tg[temp_rank][0]=1
                                    if i ==42 and j[3:5]!='00' and len(j)==5:
                                        deal_result_siroco_gif[temp_rank][1]=calc_gif.img_gif(j,1)
                                        deal_result_siroco_gif_tg[temp_rank][1]=1
                                    if i ==43 and j[3:5]!='00' and len(j)==5:
                                        deal_result_siroco_gif[temp_rank][2]=calc_gif.img_gif(j,1)
                                        deal_result_siroco_gif_tg[temp_rank][2]=1
                    for i in [11,12,13,14,15,21,22,23,31,32,33,41,42,43]:
                        for j in rank0_setting[temp_rank]:
                            if len(j) != 6:
                                if j[0:2] == str(i):
                                    result0_image_on[temp_rank][str(i)]=image_list[j]
                                    result0_image_tag[temp_rank][str(i)]=j
                                    if i ==11 and j[4:5]=='1' and len(j)==5:
                                        deal_result0_image_gif[temp_rank][0]=calc_gif.img_gif(j,0)
                                        deal_result0_image_gif_tg[temp_rank][0]=1
                                    if i ==21 and j[4:5]=='1' and len(j)==5:
                                        deal_result0_image_gif[temp_rank][1]=calc_gif.img_gif(j,0)
                                        deal_result0_image_gif_tg[temp_rank][1]=1
                                    if i ==33 and j[4:5]=='1' and len(j)==5:
                                        deal_result0_image_gif[temp_rank][2]=calc_gif.img_gif(j,0)
                                        deal_result0_image_gif_tg[temp_rank][2]=1
                                    if i ==41 and j[3:5]!='00' and len(j)==5:
                                        deal_result0_siroco_gif[temp_rank][0]=calc_gif.img_gif(j,1)
                                        deal_result0_siroco_gif_tg[temp_rank][0]=1
                                    if i ==42 and j[3:5]!='00' and len(j)==5:
                                        deal_result0_siroco_gif[temp_rank][1]=calc_gif.img_gif(j,1)
                                        deal_result0_siroco_gif_tg[temp_rank][1]=1
                                    if i ==43 and j[3:5]!='00' and len(j)==5:
                                        deal_result0_siroco_gif[temp_rank][2]=calc_gif.img_gif(j,1)
                                        deal_result0_siroco_gif_tg[temp_rank][2]=1
                except IndexError as error:
                    pass

            for i in range(0,5):
                try:
                    for j in rank_setting[i]:
                        for k in self.know_set_list:
                            if j==k:
                                if int(j[4:6])<12 and int(j[4:6])!=6:
                                    result_image_on[i]['11']=image_list['n11'+j[4:6]];result_image_tag[i]['11']='n11'+j[4:6]
                                    result_image_on[i]['12']=image_list['n12'+j[4:6]];result_image_tag[i]['12']='n12'+j[4:6]
                                    result_image_on[i]['13']=image_list['n13'+j[4:6]];result_image_tag[i]['13']='n13'+j[4:6]
                                    result_image_on[i]['14']=image_list['n14'+j[4:6]];result_image_tag[i]['14']='n14'+j[4:6]
                                    result_image_on[i]['15']=image_list['n15'+j[4:6]];result_image_tag[i]['15']='n15'+j[4:6]
                                elif int(j[4:6])==6:
                                    result_image_on[i]['31']=image_list['n31'+j[4:6]];result_image_tag[i]['31']='n31'+j[4:6]
                                    result_image_on[i]['32']=image_list['n32'+j[4:6]];result_image_tag[i]['32']='n32'+j[4:6]
                                    result_image_on[i]['33']=image_list['n33'+j[4:6]];result_image_tag[i]['33']='n33'+j[4:6]
                                elif int(j[4:6])<15:
                                    result_image_on[i]['21']=image_list['n21'+j[4:6]];result_image_tag[i]['21']='n21'+j[4:6]
                                    result_image_on[i]['22']=image_list['n22'+j[4:6]];result_image_tag[i]['22']='n22'+j[4:6]
                                    result_image_on[i]['23']=image_list['n23'+j[4:6]];result_image_tag[i]['23']='n23'+j[4:6]
                    for j in rank0_setting[i]:
                        for k in self.know_set_list:
                            if j==k:
                                if int(j[4:6])<12 and int(j[4:6])!=6:
                                    result0_image_on[i]['11']=image_list['n11'+j[4:6]];result0_image_tag[i]['11']='n11'+j[4:6]
                                    result0_image_on[i]['12']=image_list['n12'+j[4:6]];result0_image_tag[i]['12']='n12'+j[4:6]
                                    result0_image_on[i]['13']=image_list['n13'+j[4:6]];result0_image_tag[i]['13']='n13'+j[4:6]
                                    result0_image_on[i]['14']=image_list['n14'+j[4:6]];result0_image_tag[i]['14']='n14'+j[4:6]
                                    result0_image_on[i]['15']=image_list['n15'+j[4:6]];result0_image_tag[i]['15']='n15'+j[4:6]
                                elif int(j[4:6])==6:
                                    result0_image_on[i]['31']=image_list['n31'+j[4:6]];result0_image_tag[i]['31']='n31'+j[4:6]
                                    result0_image_on[i]['32']=image_list['n32'+j[4:6]];result0_image_tag[i]['32']='n32'+j[4:6]
                                    result0_image_on[i]['33']=image_list['n33'+j[4:6]];result0_image_tag[i]['33']='n33'+j[4:6]
                                elif int(j[4:6])<15:
                                    result0_image_on[i]['21']=image_list['n21'+j[4:6]];result0_image_tag[i]['21']='n21'+j[4:6]
                                    result0_image_on[i]['22']=image_list['n22'+j[4:6]];result0_image_tag[i]['22']='n22'+j[4:6]
                                    result0_image_on[i]['23']=image_list['n23'+j[4:6]];result0_image_tag[i]['23']='n23'+j[4:6]
                except:
                    pass
            # 0추스탯 1추공 2증 3크 4추
            # 6모 7공 8스탯 9속강 10지속 11스증 12특수
            # 13공속 14크확 / 15 특수액티브 / 16~19 패시브 /20 그로기포함/21 2각캐특수액티브 /22~27 액티브레벨링
            self.tagkgum_exist = 0
            self.tagk_tg = 0

            if rank_wep_name.count("(도)태극천제검")!=0: #태극검 하드 코딩
                self.tagkgum_exist = 1
                self.tagk_tg = 0

                self.tagkgum_img = self.get_photo_image('ext_img/tagk_um.png')

                tagkgum = self.tagkgum = tkinter.Button(result_window, command = lambda: self.change_tagk(ele_skill,),
                                                              image=self.tagkgum_img, bg=self.dark_main, borderwidth=0,
                                                              activebackground=self.dark_main)
                tagkgum.place(x=182,y=7)
                tagkgum.command = self.change_tagk
            if job_up_name =='(진각)':
                simari=0
                jingakgi=1
            else:
                simari=1
                jingakgi=0

            rank_stat.clear()
            rank_stat2.clear()
            rank_stat3.clear()

            rank0_stat.clear()
            rank0_stat2.clear()
            rank0_stat3.clear()

            rank_stat.extend([0,0,0,0,0])
            rank_stat2.extend([0,0,0,0,0])
            rank_stat3.extend([0,0,0,0,0])

            rank0_stat.extend([0,0,0,0,0])
            rank0_stat2.extend([0,0,0,0,0])
            rank0_stat3.extend([0,0,0,0,0])

            rank_stat_tagk = self.rank_stat_tagk_um[0] = [0, 0, 0, 0, 0]
            rank0_stat_tagk = self.rank_stat_tagk_um[1] = [0, 0, 0, 0, 0]
            rank_stat_tagk2 = self.rank_stat_tagk_yang[0] = [0, 0, 0, 0, 0]
            rank0_stat_tagk2 = self.rank_stat_tagk_yang[1] = [0, 0, 0, 0, 0]

            for i in range(0,5):
                try:
                    rank_stat[i]=("증뎀= "+str(int(round(rss[i][2],0)))+
                                  "%\n크증= "+str(int(round(rss[i][3],0)))+
                                  "%\n추뎀= "+str(int(round(rss[i][4],0)))+
                                  "%\n모공= "+str(int(round(rss[i][6],0)))+
                                  "%\n공%= "+str(int(round(rss[i][7],0)))+
                                  "%\n스탯= "+str(int(round(rss[i][8],0)))+
                                  "%\n속강= "+str(int(round(rss[i][9],0)))+
                                  "\n지속= "+str(int(round(rss[i][10],0)))+
                                  "%\n스증= "+str(int(round(rss[i][11],0)))+
                                  "%\n특수= "+str(int(round(rss[i][12],0)))+
                                  "%\n각성= "+str(int(round(rank_ult[i][3],0)))+
                                  "%\n공속= "+str(int(round(rss[i][13],0)))+
                                  "%\n크확= "+str(int(round(rss[i][14],0)))+"%")
                    if rank_wep_name[i]=="(도)태극천제검":
                        rank_stat_tagk[i]=("증뎀= "+str(int(round(rss[i][2],0)))+
                                          "%\n크증= "+str(int(round(rss[i][3],0)))+
                                          "%\n추뎀= "+str(int(round(rss[i][4],0)))+
                                          "%\n모공= "+str(-21+int(round(rss[i][6],0)))+
                                          "%\n공%= "+str(int(round(rss[i][7],0)))+
                                          "%\n스탯= "+str(int(round(rss[i][8],0)))+
                                          "%\n속강= "+str(int(round(rss[i][9],0)))+
                                          "\n지속= "+str(int(round(rss[i][10],0)))+
                                          "%\n스증= "+str(int(1.23*(100+round(rss[i][11],0))-100))+
                                          "%\n특수= "+str(int(round(rss[i][12],0)))+
                                          "%\n각성= "+str(int(round(rank_ult[i][3],0)))+
                                          "%\n공속= "+str(-70+int(round(rss[i][13],0)))+
                                          "%\n크확= "+str(int(round(rss[i][14],0)))+"%")
                    rank_stat2[i]=(str(int(round(rss[i][22],0)))+' : '+str(int(rank_dam_nolv[i]*(1+0.1014953549605047*(rss[i][22]+30))/(1+0.1014953549605047*30)))+'%'+
                                   "\n"+str(int(round(rss[i][23],0)))+' : '+str(int((1+rank_ult[i][0]/100)*rank_dam_nolv[i]*(1+0.2318898690189789*(rss[i][23]+11))/(1+0.2318898690189789*11)))+'%'+
                                   "\n"+str(int(round(rss[i][24],0)))+' : '+str(int(rank_dam_nolv[i]*(1+0.1014953549605047*(rss[i][24]+17))/(1+0.1014953549605047*17)))+'%'+
                                   "\n"+str(int(round(rss[i][25],0)))+' : '+str(int((1+rank_ult[i][1]/100)*rank_dam_nolv[i]*(1+simari*(0.2+0.05*rss[i][27]))/(1+simari*0.2)*(1+0.2318898690189789*(rss[i][25]+4))/(1+0.2318898690189789*4)))+'%'+
                                   "\n"+str(int(round(rss[i][26],0)))+' : '+str(int(rank_dam_nolv[i]*jingakgi*(1+0.1014953549605047*(rss[i][26]+5))/(1+0.1014953549605047*5)))+'%'+
                                   "\n"+str(int(round(rss[i][27],0)))+' : '+str(int((1+rank_ult[i][2]/100)*rank_dam_nolv[i]*jingakgi*(1+0.2318898690189789*(rss[i][27]+1))/(1+0.2318898690189789*1)))+'%')
                    if rank_wep_name[i]=="(도)태극천제검":
                        rank_stat_tagk2[i]=(str(int(round(rss[i][22],0)))+' : '+str(int(rank_dam_tagk_nolv[i]*(1+0.1014953549605047*(rss[i][22]+30))/(1+0.1014953549605047*30)))+'%'+
                                           "\n"+str(int(round(rss[i][23],0)))+' : '+str(int((1+rank_ult[i][0]/100)*rank_dam_tagk_nolv[i]*(1+0.2318898690189789*(rss[i][23]+11))/(1+0.2318898690189789*11)))+'%'+
                                           "\n"+str(int(round(rss[i][24],0)))+' : '+str(int(rank_dam_tagk_nolv[i]*(1+0.1014953549605047*(rss[i][24]+17))/(1+0.1014953549605047*17)))+'%'+
                                           "\n"+str(int(round(rss[i][25],0)))+' : '+str(int((1+rank_ult[i][1]/100)*rank_dam_tagk_nolv[i]*(1+simari*(0.2+0.05*rss[i][27]))/(1+simari*0.2)*(1+0.2318898690189789*(rss[i][25]+4))/(1+0.2318898690189789*4)))+'%'+
                                           "\n"+str(int(round(rss[i][26],0)))+' : '+str(int(rank_dam_tagk_nolv[i]*jingakgi*(1+0.1014953549605047*(rss[i][26]+5))/(1+0.1014953549605047*5)))+'%'+
                                           "\n"+str(int(round(rss[i][27],0)))+' : '+str(int((1+rank_ult[i][2]/100)*rank_dam_tagk_nolv[i]*jingakgi*(1+0.2318898690189789*(rss[i][27]+1))/(1+0.2318898690189789*1)))+'%')

                    rank_stat3[i]=(str(round(rss[i][16],1))+
                                   "\n"+str(int(round(rss[i][17],0)))+
                                   "\n"+str(int(round(rss[i][18],0)))+
                                   "\n"+str(int(round(rss[i][19],0))))

                    print(str(i+1)+'등')
                    print('순장비계수')
                    print(str(int(rank_dam_onlyequ[i]*100))+'%')
                    print()

                    rank0_stat[i]=("증뎀= "+str(int(round(rss0[i][2],0)))+
                                  "%\n크증= "+str(int(round(rss0[i][3],0)))+
                                  "%\n추뎀= "+str(int(round(rss0[i][4],0)))+
                                  "%\n모공= "+str(int(round(rss0[i][6],0)))+
                                  "%\n공%= "+str(int(round(rss0[i][7],0)))+
                                  "%\n스탯= "+str(int(round(rss0[i][8],0)))+
                                  "%\n속강= "+str(int(round(rss0[i][9],0)))+
                                  "\n지속= "+str(int(round(rss0[i][10],0)))+
                                  "%\n스증= "+str(int(round(rss0[i][11],0)))+
                                  "%\n특수= "+str(int(round(rss0[i][12],0)))+
                                  "%\n각성= "+str(int(round(rank0_ult[i][3],0)))+
                                  "%\n공속= "+str(int(round(rss0[i][13],0)))+
                                  "%\n크확= "+str(int(round(rss0[i][14],0)))+"%")
                    if rank_wep_name[i]=="(도)태극천제검":
                        rank0_stat_tagk[i]=("증뎀= "+str(int(round(rss0[i][2],0)))+
                                          "%\n크증= "+str(int(round(rss0[i][3],0)))+
                                          "%\n추뎀= "+str(int(round(rss0[i][4],0)))+
                                          "%\n모공= "+str(-21+int(round(rss0[i][6],0)))+
                                          "%\n공%= "+str(int(round(rss0[i][7],0)))+
                                          "%\n스탯= "+str(int(round(rss0[i][8],0)))+
                                          "%\n속강= "+str(int(round(rss0[i][9],0)))+
                                          "\n지속= "+str(int(round(rss0[i][10],0)))+
                                          "%\n스증= "+str(int(1.23*(100+round(rss0[i][11],0))-100))+
                                          "%\n특수= "+str(int(round(rss0[i][12],0)))+
                                          "%\n각성= "+str(int(round(rank0_ult[i][3],0)))+
                                          "%\n공속= "+str(-70+int(round(rss0[i][13],0)))+
                                          "%\n크확= "+str(int(round(rss0[i][14],0)))+"%")
                    rank0_stat2[i]=(str(int(round(rss0[i][22],0)))+' : '+str(int(rank0_dam_nolv[i]*(1+0.1014953549605047*(rss0[i][22]+30))/(1+0.1014953549605047*30)))+'%'+
                                   "\n"+str(int(round(rss0[i][23],0)))+' : '+str(int((1+rank0_ult[i][0]/100)*rank0_dam_nolv[i]*(1+0.2318898690189789*(rss0[i][23]+11))/(1+0.2318898690189789*11)))+'%'+
                                   "\n"+str(int(round(rss0[i][24],0)))+' : '+str(int(rank0_dam_nolv[i]*(1+0.1014953549605047*(rss0[i][24]+17))/(1+0.1014953549605047*17)))+'%'+
                                   "\n"+str(int(round(rss0[i][25],0)))+' : '+str(int((1+rank0_ult[i][1]/100)*rank0_dam_nolv[i]*(1+simari*(0.2+0.05*rss0[i][27]))/(1+simari*0.2)*(1+0.2318898690189789*(rss0[i][25]+4))/(1+0.2318898690189789*4)))+'%'+
                                   "\n"+str(int(round(rss0[i][26],0)))+' : '+str(int(rank0_dam_nolv[i]*jingakgi*(1+0.1014953549605047*(rss0[i][26]+5))/(1+0.1014953549605047*5)))+'%'+
                                   "\n"+str(int(round(rss0[i][27],0)))+' : '+str(int((1+rank0_ult[i][2]/100)*rank0_dam_nolv[i]*jingakgi*(1+0.2318898690189789*(rss0[i][27]+1))/(1+0.2318898690189789*1)))+'%')
                    if rank_wep_name[i]=="(도)태극천제검":
                        rank0_stat_tagk2[i]=(str(int(round(rss0[i][22],0)))+' : '+str(int(rank0_dam_tagk_nolv[i]*(1+0.1014953549605047*(rss0[i][22]+30))/(1+0.1014953549605047*30)))+'%'+
                                           "\n"+str(int(round(rss0[i][23],0)))+' : '+str(int((1+rank0_ult[i][0]/100)*rank0_dam_tagk_nolv[i]*(1+0.2318898690189789*(rss0[i][23]+11))/(1+0.2318898690189789*11)))+'%'+
                                           "\n"+str(int(round(rss0[i][24],0)))+' : '+str(int(rank0_dam_tagk_nolv[i]*(1+0.1014953549605047*(rss0[i][24]+17))/(1+0.1014953549605047*17)))+'%'+
                                           "\n"+str(int(round(rss0[i][25],0)))+' : '+str(int((1+rank0_ult[i][1]/100)*rank0_dam_tagk_nolv[i]*(1+simari*(0.2+0.05*rss0[i][27]))/(1+simari*0.2)*(1+0.2318898690189789*(rss0[i][25]+4))/(1+0.2318898690189789*4)))+'%'+
                                           "\n"+str(int(round(rss0[i][26],0)))+' : '+str(int(rank0_dam_tagk_nolv[i]*jingakgi*(1+0.1014953549605047*(rss0[i][26]+5))/(1+0.1014953549605047*5)))+'%'+
                                           "\n"+str(int(round(rss0[i][27],0)))+' : '+str(int((1+rank0_ult[i][2]/100)*rank0_dam_tagk_nolv[i]*jingakgi*(1+0.2318898690189789*(rss0[i][27]+1))/(1+0.2318898690189789*1)))+'%')

                    rank0_stat3[i]=(str(round(rss0[i][16],1))+
                                   "\n"+str(int(round(rss0[i][17],0)))+
                                   "\n"+str(int(round(rss0[i][18],0)))+
                                   "\n"+str(int(round(rss0[i][19],0))))

                except:
                    pass

            self.res_wep = canvas_res.create_text(12,22,text=rank_wep_name[0],font=self.guide_font,fill='white',anchor='w')
            if int(ele_skill) != 0:
                ele_change_toggle=1
                self.res_ele = canvas_res.create_text(122,149,text="자속강X="+str(rank_dam_noele[0])+"%",fill='white',font=self.small_font)

            self.res_dam = canvas_res.create_text(122,130,text=rank_dam[0],font=self.mid_font,fill='white')
            self.res_stat = canvas_res.create_text(50,293,text=rank_stat[0],fill='white')
            self.res_stat2 = canvas_res.create_text(163,263,text=rank_stat2[0],fill='white',anchor='w')
            self.res_stat3 = canvas_res.create_text(145+24,361,text=rank_stat3[0],fill='white')
            self.res_inv = canvas_res.create_text(122,174,text=rank_inv[0],font=self.guide_font,fill='white')

            gif_images = self.gif_images

            gif_images["11"] = canvas_res.create_image(57,57,image=result_image_on[0]['11'])
            gif_images["12"] = canvas_res.create_image(27,87,image=result_image_on[0]['12'])
            gif_images["13"] = canvas_res.create_image(27,57,image=result_image_on[0]['13'])
            gif_images["14"] = canvas_res.create_image(57,87,image=result_image_on[0]['14'])
            gif_images["15"] = canvas_res.create_image(27,117,image=result_image_on[0]['15'])
            gif_images["21"] = canvas_res.create_image(189,57,image=result_image_on[0]['21'])
            gif_images["22"] = canvas_res.create_image(219,57,image=result_image_on[0]['22'])
            gif_images["23"] = canvas_res.create_image(219,87,image=result_image_on[0]['23'])
            gif_images["31"] = canvas_res.create_image(189,87,image=result_image_on[0]['31'])
            gif_images["32"] = canvas_res.create_image(219,117,image=result_image_on[0]['32'])
            gif_images["33"] = canvas_res.create_image(189,117,image=result_image_on[0]['33'])
            gif_images["41"] = canvas_res.create_image(27,87,image=result_image_on[0]['41'])
            gif_images["42"] = canvas_res.create_image(219,87,image=result_image_on[0]['42'])
            gif_images["43"] = canvas_res.create_image(189,87,image=result_image_on[0]['43'])

            cn1=0
            cn4=5

            res_item_list = self.res_item_list = [{}, {}, {}, {}, {}]
            res_dam_list = self.res_dam_list = [0, 0, 0, 0, 0]
            self.res_wep_img.clear()

            for j in range(0,5):
                try:
                    for i in [11,12,13,14,15,21,22,23,31,32,33]:
                        res_item_list[j][str(i)]=canvas_res.create_image(268+cn1*29,67+78*j,image=result_image_on[j][str(i)])
                        cn1=cn1+1
                    for i in [41,42,43]:
                        res_item_list[j][str(i)]=canvas_res.create_image(268+cn4*29,67-30+78*j,image=result_image_on[j][str(i)])
                        cn4=cn4+1
                    cn1=0
                    cn4=5
                    res_dam_list[j]=canvas_res.create_text(358,34+78*j,text=rank_dam[j],font=self.mid_font,fill='white')
                    self.res_wep_img.append(canvas_res.create_image(304, 36+78*j, image=rank_wep_img[j]))
                except KeyError as error:
                    cn1=0
                    cn4=5
            length=len(rank1_list)
            if deal_result_image_gif_tg[0][0]==1:
                self.play_gif( 0,0,0,gif_images["11"],deal_result_image_gif,0,1,1)
            if deal_result_image_gif_tg[0][1]==1:
                self.play_gif( 0,0,1,gif_images["21"],deal_result_image_gif,0,1,1)
            if deal_result_image_gif_tg[0][2]==1:
                self.play_gif( 0,0,2,gif_images["33"],deal_result_image_gif,0,1,1)
            if deal_result_siroco_gif_tg[0][0]==1:
                self.play_gif( 0,0,0,gif_images["41"],deal_result_siroco_gif,0,1,1)
            if deal_result_siroco_gif_tg[0][1]==1:
                self.play_gif( 0,0,1,gif_images["42"],deal_result_siroco_gif,0,1,1)
            if deal_result_siroco_gif_tg[0][2]==1:
                self.play_gif( 0,0,2,gif_images["43"],deal_result_siroco_gif,0,1,1)

            for i in range(0,5):
                for j in [11,21,33]:
                    temp=int(j/10)-1
                    if deal_result_image_gif_tg[i][temp]==1:
                        self.play_gif(0,i,temp,res_item_list[i][str(j)],deal_result_image_gif,1,0,1)
                for j in [41,42,43]:
                    temp=j-41
                    if deal_result_siroco_gif_tg[i][temp]==1:
                        self.play_gif(0,i,temp,res_item_list[i][str(j)],deal_result_siroco_gif,1,0,1)

            self.tg_groggy = 0

            groggy_bt = self.groggy_bt = tkinter.Button(result_window,command=lambda: self.change_groggy(ele_skill),
                                                              image=self.tg_groggy_img1,
                                                              fg='white',bg=self.dark_main,borderwidth=0,activebackground=self.dark_main)
            groggy_bt.place(x=190,y=325)
            canvas_res.create_text(217, 382, text="버전=\n" + self.now_version, fill='white', anchor='c')

            self.result_cool_canvas_list = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
            canvas_res.create_image(210,455,image=self.result_upbox_img)
            canvas_res.create_image(210,606,image=self.result_downbox_img)
            canvas_res.create_image(59,558,image=self.result_sidebox_img)
            self.show_result_dealer()

        elif job_type=='buf': ########################### 버퍼 ###########################
            load_presetr=load_workbook("preset.xlsx", data_only=True)
            r_preset=load_presetr["custom"]
            res_buf = self.res_buf

            result_image_on1 = self.buff_result_image_on[0]
            result_image_on2 = self.buff_result_image_on[1]
            result_image_on3 = self.buff_result_image_on[2]

            rank_wep_img1 = self.buff_rank_wep_img[0]
            rank_wep_img2 = self.buff_rank_wep_img[1]
            rank_wep_img3 = self.buff_rank_wep_img[2]

            self.rank_type_buf = 3
            rank_setting1=[0,0,0,0,0]
            rank_setting2=[0,0,0,0,0]
            rank_setting3=[0,0,0,0,0]
            rank_inv1 = self.buff_rank_inv[0] = [0, 0, 0, 0, 0]
            rank_inv2 = self.buff_rank_inv[1] = [0, 0, 0, 0, 0]
            rank_inv3 = self.buff_rank_inv[2] = [0, 0, 0, 0, 0]

            result_image_on1.clear()
            result_image_on2.clear()
            result_image_on3.clear()

            result_image_on1.extend([{}, {}, {}, {}, {}])
            result_image_on2.extend([{}, {}, {}, {}, {}])
            result_image_on3.extend([{}, {}, {}, {}, {}])

            result_image_on1_tag = self.result_image_on_tag[0] = [{}, {}, {}, {}, {}]
            result_image_on2_tag = self.result_image_on_tag[1] = [{}, {}, {}, {}, {}]
            result_image_on3_tag = self.result_image_on_tag[2] = [{}, {}, {}, {}, {}]
            buff_result_image_gif1 = self.buff_result_image_gif[0] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            buff_result_image_gif2 = self.buff_result_image_gif[1] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            buff_result_image_gif3 = self.buff_result_image_gif[2] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            buff_result_image_gif1_tg = self.buff_result_image_gif_tg[0] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            buff_result_image_gif2_tg = self.buff_result_image_gif_tg[1] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            buff_result_image_gif3_tg = self.buff_result_image_gif_tg[2] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            buff_result_siroco_gif1 = self.buff_result_siroco_gif[0] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            buff_result_siroco_gif2 = self.buff_result_siroco_gif[1] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            buff_result_siroco_gif3 = self.buff_result_siroco_gif[2] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            buff_result_siroco_gif1_tg = self.buff_result_siroco_gif_tg[0] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            buff_result_siroco_gif2_tg = self.buff_result_siroco_gif_tg[1] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            buff_result_siroco_gif3_tg = self.buff_result_siroco_gif_tg[2] = [[0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0], [0, 0, 0]]
            rank_buf1 = self.rank_buf[0] = [0, 0, 0, 0, 0]
            rank_buf2 = self.rank_buf[1] = [0, 0, 0, 0, 0]
            rank_buf3 = self.rank_buf[2] = [0, 0, 0, 0, 0]
            rank_buf_ex1 = self.rank_buf_ex[0] = [0, 0, 0, 0, 0]
            rank_buf_ex2 = self.rank_buf_ex[1] = [0, 0, 0, 0, 0]
            rank_buf_ex3 = self.rank_buf_ex[2] = [0, 0, 0, 0, 0]
            rank_neo_buf1 = self.rank_neo_buf[0] = [0, 0, 0, 0, 0]
            rank_neo_buf2 = self.rank_neo_buf[1] = [0, 0, 0, 0, 0]
            rank_neo_buf3 = self.rank_neo_buf[2] = [0, 0, 0, 0, 0]

            rank_wep_name1 = self.buff_rank_wep_name[0] = [0, 0, 0, 0, 0]
            rank_wep_name2 = self.buff_rank_wep_name[1] = [0, 0, 0, 0, 0]
            rank_wep_name3 = self.buff_rank_wep_name[2] = [0, 0, 0, 0, 0]

            rank_wep_img1.clear()
            rank_wep_img2.clear()
            rank_wep_img3.clear()

            rank_wep_img1.extend([0, 0, 0, 0, 0])
            rank_wep_img2.extend([0, 0, 0, 0, 0])
            rank_wep_img3.extend([0, 0, 0, 0, 0])

            ## rank_setting[rank]=rank_list[a][rank][b][c]
            ## a: 0=축복,1=크오,2=합계
            ## b: 0=계수,1=스펙or증가량
            ## c: b에서 1 선택시, 0=스펙, 1=증가량
            for temp_rank in range(0,5):
                try:
                    rank_setting3[temp_rank]=rank_list[2][temp_rank][1][0]
                    rank_setting2[temp_rank]=rank_list[1][temp_rank][1][0]
                    rank_setting1[temp_rank]=rank_list[0][temp_rank][1][0]
                    rank_inv3[temp_rank]=rank_list[2][temp_rank][1][2]
                    rank_inv2[temp_rank]=rank_list[1][temp_rank][1][2]
                    rank_inv1[temp_rank]=rank_list[0][temp_rank][1][2]
                    rank_neo_buf3[temp_rank]=int(rank_list[2][temp_rank][1][4]/10)
                    rank_neo_buf2[temp_rank]=int(rank_list[1][temp_rank][1][4]/10)
                    rank_neo_buf1[temp_rank]=int(rank_list[0][temp_rank][1][4]/10)
                    rank_buf3[temp_rank]=int(rank_list[2][temp_rank][0]/10)
                    rank_buf2[temp_rank]=int(rank_list[1][temp_rank][0]/10)
                    rank_buf1[temp_rank]=int(rank_list[0][temp_rank][0]/10)
                    rank_buf_ex3[temp_rank]=rank_list[2][temp_rank][1][1]
                    rank_buf_ex2[temp_rank]=rank_list[1][temp_rank][1][1]
                    rank_buf_ex1[temp_rank]=rank_list[0][temp_rank][1][1]
                    rank_wep_name3[temp_rank]=rank_list[2][temp_rank][1][5]
                    rank_wep_name2[temp_rank]=rank_list[1][temp_rank][1][5]
                    rank_wep_name1[temp_rank]=rank_list[0][temp_rank][1][5]
                    rank_wep_img3[temp_rank]=image_list_wep[rank_wep_name3[temp_rank]]
                    rank_wep_img2[temp_rank]=image_list_wep[rank_wep_name2[temp_rank]]
                    rank_wep_img1[temp_rank]=image_list_wep[rank_wep_name1[temp_rank]]
                    for i in rank_setting3[temp_rank]:
                        if len(i)==4 and i[0]=='4': ## 융합 장비 다시 풀기
                            rank_setting3[temp_rank].append('415'+i[1]+'0')
                            rank_setting3[temp_rank].append('425'+i[2]+'0')
                            rank_setting3[temp_rank].append('435'+i[3]+'0')
                            rank_setting3[temp_rank].remove(i)
                    for i in rank_setting2[temp_rank]:
                        if len(i)==4 and i[0]=='4': ## 융합 장비 다시 풀기
                            rank_setting2[temp_rank].append('415'+i[1]+'0')
                            rank_setting2[temp_rank].append('425'+i[2]+'0')
                            rank_setting2[temp_rank].append('435'+i[3]+'0')
                            rank_setting2[temp_rank].remove(i)
                    for i in rank_setting1[temp_rank]:
                        if len(i)==4 and i[0]=='4': ## 융합 장비 다시 풀기
                            rank_setting1[temp_rank].append('415'+i[1]+'0')
                            rank_setting1[temp_rank].append('425'+i[2]+'0')
                            rank_setting1[temp_rank].append('435'+i[3]+'0')
                            rank_setting1[temp_rank].remove(i)
                    for i in [11,12,13,14,15,21,22,23,31,32,33,41,42,43]:
                        for j in rank_setting3[temp_rank]:
                            if len(j) != 6:
                                if j[0:2] == str(i):
                                    result_image_on3[temp_rank][str(i)]=image_list[j]
                                    result_image_on3_tag[temp_rank][str(i)]=j
                                    for k in [11,21,33]:
                                        if i ==k and j[4:5]=='1' and len(j)==5:
                                            buff_result_image_gif3[temp_rank][int(str(k+90)[1:2])]=calc_gif.img_gif(j,0)
                                            buff_result_image_gif3_tg[temp_rank][int(str(k+90)[1:2])]=1
                                    if i ==41 and j[3:5]!='00' and len(j)==5:
                                        buff_result_siroco_gif3[temp_rank][0]=calc_gif.img_gif(j,1)
                                        buff_result_siroco_gif3_tg[temp_rank][0]=1
                                    if i ==42 and j[3:5]!='00' and len(j)==5:
                                        buff_result_siroco_gif3[temp_rank][1]=calc_gif.img_gif(j,1)
                                        buff_result_siroco_gif3_tg[temp_rank][1]=1
                                    if i ==43 and j[3:5]!='00' and len(j)==5:
                                        buff_result_siroco_gif3[temp_rank][2]=calc_gif.img_gif(j,1)
                                        buff_result_siroco_gif3_tg[temp_rank][2]=1
                        for j in rank_setting2[temp_rank]:
                            if len(j) != 6:
                                if j[0:2] == str(i):
                                    result_image_on2[temp_rank][str(i)]=image_list[j]
                                    result_image_on2_tag[temp_rank][str(i)]=j
                                    for k in [11,21,33]:
                                        if i ==k and j[4:5]=='1' and len(j)==5:
                                            buff_result_image_gif2[temp_rank][int(str(k+90)[1:2])]=calc_gif.img_gif(j,0)
                                            buff_result_image_gif2_tg[temp_rank][int(str(k+90)[1:2])]=1
                                    if i ==41 and j[3:5]!='00' and len(j)==5:
                                        buff_result_siroco_gif2[temp_rank][0]=calc_gif.img_gif(j,1)
                                        buff_result_siroco_gif2_tg[temp_rank][0]=1
                                    if i ==42 and j[3:5]!='00' and len(j)==5:
                                        buff_result_siroco_gif2[temp_rank][1]=calc_gif.img_gif(j,1)
                                        buff_result_siroco_gif2_tg[temp_rank][1]=1
                                    if i ==43 and j[3:5]!='00' and len(j)==5:
                                        buff_result_siroco_gif2[temp_rank][2]=calc_gif.img_gif(j,1)
                                        buff_result_siroco_gif2_tg[temp_rank][2]=1
                        for j in rank_setting1[temp_rank]:
                            if len(j) != 6:
                                if j[0:2] == str(i):
                                    result_image_on1[temp_rank][str(i)]=image_list[j] ##
                                    result_image_on1_tag[temp_rank][str(i)]=j
                                    for k in [11,21,33]:
                                        if i ==k and j[4:5]=='1' and len(j)==5:
                                            buff_result_image_gif1[temp_rank][int(str(k+90)[1:2])]=calc_gif.img_gif(j,0)
                                            buff_result_image_gif1_tg[temp_rank][int(str(k+90)[1:2])]=1
                                    if i ==41 and j[3:5]!='00' and len(j)==5:
                                        buff_result_siroco_gif1[temp_rank][0]=calc_gif.img_gif(j,1)
                                        buff_result_siroco_gif1_tg[temp_rank][0]=1
                                    if i ==42 and j[3:5]!='00' and len(j)==5:
                                        buff_result_siroco_gif1[temp_rank][1]=calc_gif.img_gif(j,1)
                                        buff_result_siroco_gif1_tg[temp_rank][1]=1
                                    if i ==43 and j[3:5]!='00' and len(j)==5:
                                        buff_result_siroco_gif1[temp_rank][2]=calc_gif.img_gif(j,1)
                                        buff_result_siroco_gif1_tg[temp_rank][2]=1
                except IndexError as error:
                    pass

            """
            self.buf_jingak_exist = 0
            self.buf_jingak_tg = 0
            if job_name=='(버프)세인트': ## 진각 전환 하드 코딩
                def change_buf_jingak():
                    rank_type_buf = self.rank_type_buf
    
                    if rank_type_buf==1:
                        c_rank0_buf=rank_neo_buf1
                        c_rank_buf=rank_buf1
                    elif rank_type_buf==2:
                        c_rank0_buf=rank_neo_buf2
                        c_rank_buf=rank_buf2
                    elif rank_type_buf==3:
                        c_rank0_buf=rank_neo_buf3
                        c_rank_buf=rank_buf3
                    if self.buf_jingak_tg == 0:
                        self.buf_jingak["image"] = self.buf_jingak_img[0]
                        canvas_res.itemconfig(res_buf,text=c_rank0_buf[self.now_rank_num],fill='gray')
                        self.buf_jingak_tg = 1
                    elif self.buf_jingak_tg == 1:
                        self.buf_jingak["image"] = self.buf_jingak_img[1]
                        canvas_res.itemconfig(res_buf,text=c_rank_buf[self.now_rank_num],fill='white')
                        self.buf_jingak_tg = 0
    
                self.buf_jingak_exist = 1
                buf_jingak = self.buf_jingak = tkinter.Button(result_window,command=change_buf_jingak,image=self.buf_jingak_img[1],bg=self.dark_main,borderwidth=0,activebackground=self.dark_main)
                buf_jingak.place(x=172,y=7)
                buf_jingak.image=self.buf_jingak_img[1]
                buf_jingak.command=change_buf_jingak
            """

            self.res_wep = canvas_res.create_text(12,22,text=rank_wep_name3[0],font=self.guide_font,fill='white',anchor='w')
            canvas_res.create_text(122-55,193,text=rank_inv1[0],font=self.small_font,fill='white',anchor="w")
            res_buf = self.res_buf = canvas_res.create_text(122,125,text=rank_buf3[0],font=self.mid_font,fill='white')
            self.res_buf_type_what = canvas_res.create_text(122,145,text="총합 기준",font=self.guide_font,fill='white')
            self.res_buf_ex = [
                canvas_res.create_text(64, 283, text=rank_buf_ex3[0][0], font=self.small_font, fill='white'),
                canvas_res.create_text(183, 261, text=rank_buf_ex3[0][1], font=self.small_font, fill='white'),
                canvas_res.create_text(183, 318, text=rank_buf_ex3[0][2], font=self.small_font, fill='white')
            ]

            gif_images = self.gif_images

            gif_images["11"]=canvas_res.create_image(57,52,image=result_image_on3[0]['11'])
            gif_images["12"]=canvas_res.create_image(27,82,image=result_image_on3[0]['12'])
            gif_images["13"]=canvas_res.create_image(27,52,image=result_image_on3[0]['13'])
            gif_images["14"]=canvas_res.create_image(57,82,image=result_image_on3[0]['14'])
            gif_images["15"]=canvas_res.create_image(27,112,image=result_image_on3[0]['15'])
            gif_images["21"]=canvas_res.create_image(189,52,image=result_image_on3[0]['21'])
            gif_images["22"]=canvas_res.create_image(219,52,image=result_image_on3[0]['22'])
            gif_images["23"]=canvas_res.create_image(219,82,image=result_image_on3[0]['23'])
            gif_images["31"]=canvas_res.create_image(189,82,image=result_image_on3[0]['31'])
            gif_images["32"]=canvas_res.create_image(219,112,image=result_image_on3[0]['32'])
            gif_images["33"]=canvas_res.create_image(189,112,image=result_image_on3[0]['33'])
            gif_images["41"]=canvas_res.create_image(27,82,image=result_image_on3[0]['41'])
            gif_images["42"]=canvas_res.create_image(219,82,image=result_image_on3[0]['42'])
            gif_images["43"]=canvas_res.create_image(189,82,image=result_image_on3[0]['43'])

            cn1=0
            cn6=5
            res_img_list = self.res_img_list = {}
            res_buf_list = self.res_buf_list = {}
            self.res_wep_img.clear()  # TODO: memory leak?

            for j in range(0,5):
                try:
                    for i in [11,12,13,14,15,21,22,23,31,32,33]:
                        temp_res=canvas_res.create_image(268+cn1*29,67+78*j,image=result_image_on3[j][str(i)])
                        res_img_list[str(j)+str(i)]=temp_res
                        cn1=cn1+1
                    for i in [41,42,43]:
                        temp_res=canvas_res.create_image(268+cn6*29,67-30+78*j,image=result_image_on3[j][str(i)])
                        res_img_list[str(j)+str(i)]=temp_res
                        cn6=cn6+1
                    cn1=0
                    cn6=5
                    temp_buf=canvas_res.create_text(358,34+78*j,text=rank_buf3[j],font=self.mid_font,fill='white')
                    self.res_wep_img.append(canvas_res.create_image(304,36+78*j,image=rank_wep_img3[j]))
                    res_buf_list[j]=temp_buf
                except KeyError as error:
                    cn1=0
                    cn6=5
            length=len(rank_list[0])

            rank_type_but1=tkinter.Button(result_window,command=lambda: self.change_rank_type(1),image=self.type1_img,bg=self.dark_main,borderwidth=0,activebackground=self.dark_main)
            rank_type_but1.place(x=8,y=337)
            rank_type_but2=tkinter.Button(result_window,command=lambda: self.change_rank_type(2),image=self.type2_img,bg=self.dark_main,borderwidth=0,activebackground=self.dark_main)
            rank_type_but2.place(x=84,y=337)
            rank_type_but3=tkinter.Button(result_window,command=lambda: self.change_rank_type(3),image=self.type3_img,bg=self.dark_main,borderwidth=0,activebackground=self.dark_main)
            rank_type_but3.place(x=160,y=337)

            if buff_result_image_gif3_tg[0][0]==1:
                self.play_gif( 0,0,0,gif_images["11"],buff_result_image_gif3,0,1,1)
            if buff_result_image_gif3_tg[0][1]==1:
                self.play_gif( 0,0,1,gif_images["21"],buff_result_image_gif3,0,1,1)
            if buff_result_image_gif3_tg[0][2]==1:
                self.play_gif( 0,0,2,gif_images["33"],buff_result_image_gif3,0,1,1)
            if buff_result_siroco_gif3_tg[0][0]==1:
                self.play_gif( 0,0,0,gif_images["41"],buff_result_siroco_gif3,0,1,1)
            if buff_result_siroco_gif3_tg[0][1]==1:
                self.play_gif( 0,0,1,gif_images["42"],buff_result_siroco_gif3,0,1,1)
            if buff_result_siroco_gif3_tg[0][2]==1:
                self.play_gif( 0,0,2,gif_images["43"],buff_result_siroco_gif3,0,1,1)

            for i in range(0,5):
                for j in [11,21,33]:
                    temp=int(j/10)-1
                    if buff_result_image_gif3_tg[i][temp]==1:
                        self.play_gif(0,i,temp,res_img_list[str(i)+str(j)],buff_result_image_gif3,1,0,1)
                for j in [41,42,43]:
                    temp=j-41
                    if buff_result_siroco_gif3_tg[i][temp]==1:
                        self.play_gif(0,i,temp,res_img_list[str(i)+str(j)],buff_result_siroco_gif3,1,0,1)

            load_presetr.close()

        res_bt1=tkinter.Button(result_window,command=lambda: self.change_rank(0,job_type,ele_skill,rank_setting,rank_ult),image=self.show_detail_img,bg=self.dark_blue,borderwidth=0,activebackground=self.dark_blue);res_bt1.place(x=486,y=20+78*0)
        res_bt2=tkinter.Button(result_window,command=lambda: self.change_rank(1,job_type,ele_skill,rank_setting,rank_ult),image=self.show_detail_img,bg=self.dark_blue,borderwidth=0,activebackground=self.dark_blue)
        res_bt3=tkinter.Button(result_window,command=lambda: self.change_rank(2,job_type,ele_skill,rank_setting,rank_ult),image=self.show_detail_img,bg=self.dark_blue,borderwidth=0,activebackground=self.dark_blue)
        res_bt4=tkinter.Button(result_window,command=lambda: self.change_rank(3,job_type,ele_skill,rank_setting,rank_ult),image=self.show_detail_img,bg=self.dark_blue,borderwidth=0,activebackground=self.dark_blue)
        res_bt5=tkinter.Button(result_window,command=lambda: self.change_rank(4,job_type,ele_skill,rank_setting,rank_ult),image=self.show_detail_img,bg=self.dark_blue,borderwidth=0,activebackground=self.dark_blue)

        if length>1:
            res_bt2.place(x=486,y=20+78*1)
        if length>2:
            res_bt3.place(x=486,y=20+78*2)
        if length>3:
            res_bt4.place(x=486,y=20+78*3)
        if length>4:
            res_bt5.place(x=486,y=20+78*4)

        show_tag_but=tkinter.Button(result_window,command=lambda: self.show_set_name(job_type),image=self.show_tag_img,bg=self.dark_sub,borderwidth=0,activebackground=self.dark_sub)
        show_tag_but.place(x=173,y=158-26)
        show_tag_but.image=self.show_tag_img

        capture_but=tkinter.Button(result_window,command=lambda: self.capture_screen(result_window),image=self.capture_img,bg=self.dark_sub,borderwidth=0,activebackground=self.dark_sub)
        capture_but.place(x=173-164,y=158-26)
        capture_but.image=self.capture_img
        canvas_res.image=result_bg,random_npc_img
        res_bt1.image=self.show_detail_img
        self.place_center(result_window, 0)

    def show_result_dealer(self):
        result_window = self.result_window
        canvas_res = self.canvas_res
        now_rank_num = self.now_rank_num

        result_window.geometry("585x710")

        rank0_list = self.rank_list[0]
        rank1_list = self.rank_list[1]

        set_list=[]
        cool_list=[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0] ##최종값
        cool_list1=[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0] ##쿨감
        cool_list2=[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0] ##쿨회복
        font_overstrike=tkinter.font.Font(family="맑은 고딕", size=10,overstrike=1)
        font_normal=tkinter.font.Font(family="맑은 고딕", size=10)

        if self.tg_groggy == 0:
            now_list_list=list(rank1_list[now_rank_num][1][0])
            now_damage_groggy=rank1_list[now_rank_num][0]
            now_damage_sustain=rank1_list[now_rank_num][1][7]
            now_base_array=rank1_list[now_rank_num][1][1]
        elif self.tg_groggy == 1:
            now_list_list=list(rank0_list[now_rank_num][1][0])
            now_damage_groggy=rank0_list[now_rank_num][1][7]
            now_damage_sustain=rank0_list[now_rank_num][0]
            now_base_array=rank0_list[now_rank_num][1][1]
        if now_damage_groggy>=now_damage_sustain:
            ratio_tg='groggy'
            groggy_sustain_ratio=round(now_damage_groggy/now_damage_sustain*100,1)
            move_gauge=(groggy_sustain_ratio-100)
        elif now_damage_groggy<now_damage_sustain:
            ratio_tg='sustain'
            groggy_sustain_ratio=round(now_damage_sustain/now_damage_groggy*100,1)
            move_gauge=-(groggy_sustain_ratio-100)
        if move_gauge>=70: move_gauge=70
        elif move_gauge<-70: move_gauge=-70

        for now_equ in now_list_list:
            if len(now_equ)==5:
                set_list.append('1'+now_equ[2:4])
        for i in range(101,150):
            if set_list.count(str(i))==2: now_list_list.append(str(i)+'1')
            elif set_list.count(str(i))==3: now_list_list.append(str(i)+'2')
            elif set_list.count(str(i))==4: now_list_list.append(str(i)+'2')
            elif set_list.count(str(i))==5: now_list_list.append(str(i)+'3')
        penalty_score=0;bonus_score=0;utility_score=0
        for now_equ in now_list_list:
            now_equ_opt=self.opt_leveling.get(now_equ)
            for i in range(0,18):
                penalty_score+=now_equ_opt[54]
                bonus_score+=now_equ_opt[55]
                utility_score+=now_equ_opt[56]
                cool_list1[i]=100-(1-cool_list1[i]/100)*(1-now_equ_opt[18+i]/100)*100
                cool_list2[i]+=now_equ_opt[36+i]
        for i in range(0,18):
            cool_list[i]=-round(100-(1-cool_list1[i]/100)/(1+cool_list2[i]/100)*100,1)

        result_cool_canvas_list = self.result_cool_canvas_list
        if result_cool_canvas_list != [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]:
            for i in range(0,18):
                if cool_list[i]==0: fg='gray'
                else: fg='white'
                canvas_res.itemconfig(result_cool_canvas_list[i],text=str(cool_list[i])+"%",fill=fg)
                canvas_res.delete('not_overlap')
        else:
            canvas_res.create_text(58,425,text='<쿨감%>',font=self.guide_font,fill='white')
            for i in range(0,18):
                lvl_type_list=[1,5,10,15,20,25,30,35,40,45,50,60,70,75,80,85,95,100]
                canvas_res.create_text(20,450+i*14,text='Lv.',fill='white',anchor='w')
                canvas_res.create_text(43,450+i*14,text=str(lvl_type_list[i]),fill='white',anchor='c')
                if cool_list[i]==0: fg='gray'
                else: fg='white'
                result_cool_canvas_list[i]=canvas_res.create_text(98,450+i*14,text=str(cool_list[i])+"%",fill=fg,anchor='e')
            canvas_res.create_image(210,445,image=self.result_gauge_bar_img)

        canvas_res.create_image(210+int(move_gauge),445,image=self.result_gauge_img,tags=('not_overlap',))
        canvas_res.create_text(130,424,text="지속딜",fill='sky blue',anchor='w',tags=('not_overlap',))
        canvas_res.create_text(290,424,text="그로기",fill='pink1',anchor='e',tags=('not_overlap',))
        if 120>groggy_sustain_ratio>100 and ratio_tg=='groggy': ratio_color='pink1'
        elif groggy_sustain_ratio>=120 and ratio_tg=='groggy': ratio_color='OrangeRed2'
        elif 120>groggy_sustain_ratio>100 and ratio_tg=='sustain': ratio_color='sky blue'
        elif groggy_sustain_ratio>=120 and ratio_tg=='sustain': ratio_color='RoyalBlue1'
        else: ratio_color='ivory2'

        if groggy_sustain_ratio>=120:
            if ratio_tg=='groggy':
                ratio_value_text="<그로기 특화>";ratio_value_fg='OrangeRed2';check_list3_fg='white';check_list3_tg=font_normal
            if ratio_tg=='sustain':
                ratio_value_text="<지속딜 특화>";ratio_value_fg='RoyalBlue1';check_list3_fg='gray';check_list3_tg=font_overstrike
        elif groggy_sustain_ratio>=110:
            if ratio_tg=='groggy':
                ratio_value_text="<그로기 우세>";ratio_value_fg='pink1';check_list3_fg='white';check_list3_tg=font_normal
            if ratio_tg=='sustain':
                ratio_value_text="<지속딜 우세>";ratio_value_fg='sky blue';check_list3_fg='gray';check_list3_tg=font_overstrike
        else:
            ratio_value_text="<올라운더>";ratio_value_fg='ivory2';check_list3_fg='gray';check_list3_tg=font_overstrike
        canvas_res.create_text(213+int(move_gauge),465,text=str(round(groggy_sustain_ratio-100,1))+"%",font=self.guide_font,fill=ratio_color,anchor='c',tags=('not_overlap',))
        canvas_res.create_text(210,483,text=ratio_value_text,font=self.guide_font,fill=ratio_value_fg,anchor='c',tags=('not_overlap',))

        canvas_res.create_text(189,521,text="<세팅 체크리스트>",font=self.guide_font,fill='white',tags=('not_overlap',))
        if (-sum(cool_list)/18)>15:
            check_list4_fg="white";check_list4_tg=font_normal
        else:
            check_list4_fg="gray";check_list4_tg=font_overstrike
        if penalty_score>0:
            check_list2_fg="pink1";check_list2_tg=font_normal
        else:
            check_list2_fg="gray";check_list2_tg=font_overstrike
        if bonus_score>0:
            check_list6_fg="sky blue";check_list6_tg=font_normal
        else:
            check_list6_fg="gray";check_list6_tg=font_overstrike
        if utility_score>0:
            check_list7_fg="sky blue";check_list7_tg=font_normal
        else:
            check_list7_fg="gray";check_list7_tg=font_overstrike
        if now_base_array[13]<10 or now_base_array[14]<10:
            check_list1_fg="pink1";check_list1_tg=font_normal
        else:
            check_list1_fg="gray";check_list1_tg=font_overstrike
        check_list5_fg="gray";check_list5_tg=font_overstrike
        for special_equ in ['11130','11131','12130','13130','14130','15130','15140']:
            if now_list_list.count(special_equ)!=0:
                check_list5_fg="white";check_list5_tg=font_normal

        canvas_res.create_text(130,545,text="※ 크확/속도 낮음",fill=check_list1_fg,tags=('not_overlap',),anchor='w',font=check_list1_tg)
        canvas_res.create_text(130,565,text="※ 지연,랜덤딜/패널티",fill=check_list2_fg,tags=('not_overlap',),anchor='w',font=check_list2_tg)
        canvas_res.create_text(130,595,text="※ 그로기 딜비중 높음",fill=check_list3_fg,tags=('not_overlap',),anchor='w',font=check_list3_tg)
        canvas_res.create_text(130,615,text="※ 쿨타임 감소율 높음",fill=check_list4_fg,tags=('not_overlap',),anchor='w',font=check_list4_tg)
        canvas_res.create_text(130,635,text="※ 탈리신/흐름셋 채용",fill=check_list5_fg,tags=('not_overlap',),anchor='w',font=check_list5_tg)
        canvas_res.create_text(130,665,text="※ 특수 데미지 옵션",fill=check_list6_fg,tags=('not_overlap',),anchor='w',font=check_list6_tg)
        canvas_res.create_text(130,685,text="※ 높은 유틸성",fill=check_list7_fg,tags=('not_overlap',),anchor='w',font=check_list7_tg)

        def show_explain_ratio(event): show_check_list(8)
        def del_explain_ratio(event): del_check_list(8)
        def show_explain_damage(event): show_check_list(9)
        def del_explain_damage(event): del_check_list(9)
        def show_check_list1(event): show_check_list(1)
        def del_check_list1(event): del_check_list(1)
        def show_check_list2(event): show_check_list(2)
        def del_check_list2(event): del_check_list(2)
        def show_check_list3(event): show_check_list(3)
        def del_check_list3(event): del_check_list(3)
        def show_check_list4(event): show_check_list(4)
        def del_check_list4(event): del_check_list(4)
        def show_check_list5(event): show_check_list(5)
        def del_check_list5(event): del_check_list(5)
        def show_check_list6(event): show_check_list(6)
        def del_check_list6(event): del_check_list(6)
        def show_check_list7(event): show_check_list(7)
        def del_check_list7(event): del_check_list(7)
        def show_check_style(event): show_check_list('style')
        def del_check_style(event): del_check_list('style')

        def show_check_list(num):
            canvas_res.create_image(445,560,image=self.result_showbox_img,tags=('not_overlap','mouse_overlap'))
            try:
                canvas_res.create_text(340,560,text=calc_result.result_check_list_explain[num-1],tags=('not_overlap','mouse_overlap'),anchor='w',width=210)
            except:
                if num=='style':
                    style_check_list=style_compare()
                    canvas_res.create_text(340,560,text=style_check_list,tags=('not_overlap','mouse_overlap'),anchor='w',width=210)
        def del_check_list(num):
            canvas_res.delete("mouse_overlap")

        result_checklist_img = self.result_checklist_img

        explain_ratio_bt=tkinter.Label(result_window,image=result_checklist_img,font=self.small_font,bd=0,bg=self.result_sub);
        explain_ratio_bt.place(x=265,y=470)
        explain_ratio_bt.bind("<Enter>",show_explain_ratio);explain_ratio_bt.bind("<Leave>",del_explain_ratio)

        explain_damage_bt=tkinter.Label(result_window,image=result_checklist_img,font=self.small_font,bd=0,bg=self.result_sub);
        explain_damage_bt.place(x=146,y=76)
        explain_damage_bt.bind("<Enter>",show_explain_damage);explain_damage_bt.bind("<Leave>",del_explain_damage)

        check_list1_bt=tkinter.Label(result_window,image=result_checklist_img,font=self.small_font,bd=0,bg=self.result_sub);
        check_list1_bt.place(x=265,y=533)
        check_list1_bt.bind("<Enter>",show_check_list1);check_list1_bt.bind("<Leave>",del_check_list1)
        check_list2_bt=tkinter.Label(result_window,image=result_checklist_img,font=self.small_font,bd=0,bg=self.result_sub);
        check_list2_bt.place(x=265,y=553)
        check_list2_bt.bind("<Enter>",show_check_list2);check_list2_bt.bind("<Leave>",del_check_list2)
        check_list3_bt=tkinter.Label(result_window,image=result_checklist_img,font=self.small_font,bd=0,bg=self.result_sub);
        check_list3_bt.place(x=265,y=583)
        check_list3_bt.bind("<Enter>",show_check_list3);check_list3_bt.bind("<Leave>",del_check_list3)
        check_list4_bt=tkinter.Label(result_window,image=result_checklist_img,font=self.small_font,bd=0,bg=self.result_sub);
        check_list4_bt.place(x=265,y=603)
        check_list4_bt.bind("<Enter>",show_check_list4);check_list4_bt.bind("<Leave>",del_check_list4)
        check_list5_bt=tkinter.Label(result_window,image=result_checklist_img,font=self.small_font,bd=0,bg=self.result_sub);
        check_list5_bt.place(x=265,y=623)
        check_list5_bt.bind("<Enter>",show_check_list5);check_list5_bt.bind("<Leave>",del_check_list5)
        check_list6_bt=tkinter.Label(result_window,image=result_checklist_img,font=self.small_font,bd=0,bg=self.result_sub);
        check_list6_bt.place(x=265,y=653)
        check_list6_bt.bind("<Enter>",show_check_list6);check_list6_bt.bind("<Leave>",del_check_list6)
        check_list7_bt=tkinter.Label(result_window,image=result_checklist_img,font=self.small_font,bd=0,bg=self.result_sub);
        check_list7_bt.place(x=265,y=673)
        check_list7_bt.bind("<Enter>",show_check_list7);check_list7_bt.bind("<Leave>",del_check_list7)

        def style_compare():
            style_calced = self.style_calced
            creature_calced = self.creature_calced

            now_updam=now_base_array[2]
            now_cridam=now_base_array[3]
            now_bondam=now_base_array[4]
            now_ele=now_base_array[9]
            change_updam=15;change_cridam=10;change_bondam=10;change_ele=32
            no_style_default=1

            if style_calced == '증뎀10%' and creature_calced != '증뎀10%': no_style_default=(now_updam-10+100)/(now_updam+100);now_updam=now_updam-10
            elif style_calced == '증뎀15%':
                if creature_calced == '증뎀10%':
                    updam_diff = -5
                else:
                    updam_diff = -15

                no_style_default = (now_updam + updam_diff + 100) / (now_updam + 100)
                now_updam += updam_diff
            elif style_calced == '크증10%':
                if creature_calced == '크증10%' or creature_calced == '크증18%':
                    cridam_diff = 0
                else:
                    cridam_diff = -10

                no_style_default = (now_cridam + cridam_diff + 100) / (now_cridam + 100)
                now_cridam += cridam_diff
            elif style_calced=='추뎀10%': no_style_default=(now_bondam-10+100)/(now_bondam+100);now_bondam=now_bondam-10
            elif style_calced=='속강32': no_style_default=((now_ele-32)*0.0045+1.05)/((now_ele)*0.0045+1.05);now_ele=now_ele-32

            if creature_calced == '크증10%' or creature_calced == '크증18%':
                change_cridam = 0
            elif creature_calced == '증뎀10%':
                change_updam = 5

            style_updam=str(round((now_updam+change_updam+100)/(now_updam+100)*no_style_default*100-100,2))+"%"
            style_cridam=str(round((now_cridam+change_cridam+100)/(now_cridam+100)*no_style_default*100-100,2))+"%"
            style_bondam=str(round((now_bondam+change_bondam+100)/(now_bondam+100)*no_style_default*100-100,2))+"%"
            style_ele=str(round(((now_ele+change_ele)*0.0045+1.05)/((now_ele)*0.0045+1.05)*no_style_default*100-100,2))+"%"

            result_explain="현 계산값에서 데미지 칭호를 변경했을 때의 증감율을 나타냅니다. 특정 액티브 스킬공격력 증가 옵션은 계산되지 않습니다.\n\n\n"
            result_str='증뎀15% 칭호 : '+style_updam+'\n크증10% 칭호 : '+style_cridam+'\n추뎀10% 칭호 : '+style_bondam+'\n속강+32 칭호 : '+style_ele

            return result_explain+result_str

        style_compare_bt=tkinter.Label(result_window,image=self.style_compare_img,bd=0,bg=self.result_sub)
        style_compare_bt.place(x=255,y=508)
        style_compare_bt.bind("<Enter>",show_check_style);style_compare_bt.bind("<Leave>",del_check_style)

    def play_gif(self, count_frame, now_rank, now_pc, show_res, gif_list, mode, mode2, mode3):
        #now_rank:순위
        #now_pc:0(상의,하의),1(팔찌,반지),2(귀걸,보장)
        #show_res:이미지 재생될 canvas 객체
        #gif_list:딜벞 구분
        #mode:0(인포창),1(리스트)
        #mode2:0(정지불가),1(정지가능)  > 순위 바꾸기 정지
        #mode3:0(정지불가),1(정지가능)  > 버퍼 정렬변경 정지
        result_window = self.result_window
        canvas_res = self.canvas_res

        now_frame=gif_list[now_rank][now_pc][int(count_frame)]
        count_frame += 0.3
        if self.pause_gif == 0 or mode == 1:
            canvas_res.itemconfig(show_res,image=now_frame)
        while self.stop_gif == 1 and mode2 == 1:
            return
        while self.stop_gif2 == 1 and mode3 == 1:
            return
        else:
            if count_frame >=len(gif_list[now_rank][now_pc]):
                result_window.after(30, self.play_gif, 0,now_rank,now_pc,show_res,gif_list,mode,mode2,mode3)
            else:
                result_window.after(30, self.play_gif, count_frame,now_rank,now_pc,show_res,gif_list,mode,mode2,mode3)

    def time_delay1(self):
        self.stop_gif = 0
    def time_delay2(self):
        self.stop_gif = 1
        threading.Timer(0.035, self.time_delay1).start()
    def time_delay(self):
        threading.Timer(0, self.time_delay2).start()
    def time_delay3(self):
        self.stop_gif2 = 0
    def time_delay4(self):
        self.stop_gif2 = 1
        threading.Timer(0.035, self.time_delay3).start()
    def time_delayy(self):
        threading.Timer(0, self.time_delay4).start()

    ## 지속딜 <> 그로기 전환
    def change_groggy(self, ele_skill):
        threading.Timer(0.07, self.change_groggy2,args=(ele_skill,)).start()
        threading.Timer(0, self.time_delayy).start()

    def change_groggy2(self, ele_skill):
        result_window = self.result_window
        canvas_res = self.canvas_res
        gif_images = self.gif_images
        rank_wep_name = self.deal_rank_wep_name[0]
        rank0_wep_name = self.deal_rank_wep_name[1]
        rank_stat = self.deal_rank_stat1[0]
        rank_stat2 = self.deal_rank_stat2[0]
        rank_stat3 = self.deal_rank_stat3[0]
        rank0_stat = self.deal_rank_stat1[1]
        rank0_stat2 = self.deal_rank_stat2[1]
        rank0_stat3 = self.deal_rank_stat3[1]
        res_dam = self.res_dam
        res_stat = self.res_stat
        res_stat2 = self.res_stat2
        res_stat3 = self.res_stat3
        rank_dam_noele = self.rank_dam_noele[0]
        rank0_dam_noele = self.rank_dam_noele[1]
        rank_inv = self.deal_rank_inv[0]
        rank0_inv = self.deal_rank_inv[1]
        deal_result_image_gif = self.deal_result_image_gif[0]
        deal_result0_image_gif = self.deal_result_image_gif[1]
        deal_result_image_gif_tg = self.deal_result_image_gif_tg[0]
        deal_result0_image_gif_tg = self.deal_result_image_gif_tg[1]
        deal_result_siroco_gif = self.deal_result_siroco_gif[0]
        deal_result0_siroco_gif = self.deal_result_siroco_gif[1]
        deal_result_siroco_gif_tg = self.deal_result_siroco_gif_tg[0]
        deal_result0_siroco_gif_tg = self.deal_result_siroco_gif_tg[1]
        res_item_list = self.res_item_list

        self.now_rank_num = 0
        if self.tagkgum_exist == 1:
            self.tagkgum["image"] = self.tagkgum_img
            self.tagk_tg = 0

        if self.tg_groggy == 0:
            self.groggy_bt["image"] = self.tg_groggy_img2
            cool_what = f"지속딜{self.cool_eff_text}"
            image_changed = self.deal_result_image_on[1]
            change_dam_noele=rank0_dam_noele
            change_dam = self.rank_dam[1]
            change_stat=rank0_stat
            change_stat2=rank0_stat2
            rank_changed = self.rank_dam[1]
            change_stat3=rank0_stat3
            change_inv=rank0_inv
            image_gif_changed_tg=deal_result0_image_gif_tg
            image_gif_changed=deal_result0_image_gif
            siroco_gif_changed_tg=deal_result0_siroco_gif_tg
            siroco_gif_changed=deal_result0_siroco_gif
            wep_changed=rank0_wep_name
            wep_img_changed=self.deal_rank_wep_img[1]

            self.tg_groggy = 1
        elif self.tg_groggy == 1:
            self.groggy_bt["image"] = self.tg_groggy_img1
            cool_what = f"그로기{self.cool_eff_text}"
            image_changed = self.deal_result_image_on[0]
            change_dam_noele=rank_dam_noele
            change_dam = self.rank_dam[0]
            change_stat=rank_stat
            change_stat2=rank_stat2
            rank_changed = self.rank_dam[0]
            change_stat3=rank_stat3
            change_inv=rank_inv
            image_gif_changed_tg=deal_result_image_gif_tg
            image_gif_changed=deal_result_image_gif
            siroco_gif_changed_tg=deal_result_siroco_gif_tg
            siroco_gif_changed=deal_result_siroco_gif
            wep_changed=rank_wep_name
            wep_img_changed=self.deal_rank_wep_img[0]
            self.tg_groggy = 0

        ###############
        canvas_res.itemconfig(self.res_wep,text=wep_changed[0],fill="white")
        canvas_res.itemconfig(self.res_cool_what,text=cool_what)
        if int(ele_skill) != 0:
            canvas_res.itemconfig(self.res_ele,text="자속강X="+str(change_dam_noele[0])+"%")
        canvas_res.itemconfig(res_dam,text=change_dam[0])
        canvas_res.itemconfig(res_stat,text=change_stat[0])
        canvas_res.itemconfig(res_stat2,text=change_stat2[0])
        canvas_res.itemconfig(res_stat3,text=change_stat3[0])
        canvas_res.itemconfig(self.res_inv,text=change_inv[0])
        canvas_res.itemconfig(gif_images["11"],image=image_changed[0]['11'])
        canvas_res.itemconfig(gif_images["12"],image=image_changed[0]['12'])
        canvas_res.itemconfig(gif_images["13"],image=image_changed[0]['13'])
        canvas_res.itemconfig(gif_images["14"],image=image_changed[0]['14'])
        canvas_res.itemconfig(gif_images["15"],image=image_changed[0]['15'])
        canvas_res.itemconfig(gif_images["21"],image=image_changed[0]['21'])
        canvas_res.itemconfig(gif_images["22"],image=image_changed[0]['22'])
        canvas_res.itemconfig(gif_images["23"],image=image_changed[0]['23'])
        canvas_res.itemconfig(gif_images["31"],image=image_changed[0]['31'])
        canvas_res.itemconfig(gif_images["32"],image=image_changed[0]['32'])
        canvas_res.itemconfig(gif_images["33"],image=image_changed[0]['33'])
        canvas_res.itemconfig(gif_images["41"],image=image_changed[0]['41'])
        canvas_res.itemconfig(gif_images["42"],image=image_changed[0]['42'])
        canvas_res.itemconfig(gif_images["43"],image=image_changed[0]['43'])
        self.stop_gif = 1
        self.stop_gif2 = 1
        time.sleep(0.2)
        self.stop_gif = 0
        self.stop_gif2 = 0
        if image_gif_changed_tg[0][0]==1:
            result_window.after(0,self.play_gif,0,0,0,gif_images["11"],image_gif_changed,0,1,1)
        if image_gif_changed_tg[0][1]==1:
            result_window.after(0,self.play_gif,0,0,1,gif_images["21"],image_gif_changed,0,1,1)
        if image_gif_changed_tg[0][2]==1:
            result_window.after(0,self.play_gif,0,0,2,gif_images["33"],image_gif_changed,0,1,1)
        if siroco_gif_changed_tg[0][0]==1:
            result_window.after(0,self.play_gif,0,0,0,gif_images["41"],siroco_gif_changed,0,1,1)
        if siroco_gif_changed_tg[0][1]==1:
            result_window.after(0,self.play_gif,0,0,1,gif_images["42"],siroco_gif_changed,0,1,1)
        if siroco_gif_changed_tg[0][2]==1:
            result_window.after(0,self.play_gif,0,0,2,gif_images["43"],siroco_gif_changed,0,1,1)

        for j in range(0,5):
            try:
                for i in [11,12,13,14,15,21,22,23,31,32,33,41,42,43]:
                    canvas_res.itemconfig(res_item_list[j][str(i)],image=image_changed[j][str(i)])
                canvas_res.itemconfig(self.res_dam_list[j],text=rank_changed[j],fill='white')
                canvas_res.itemconfig(self.res_wep_img[j],image=wep_img_changed[j])
            except KeyError as error:
                pass
        for i in range(0,5):
            for j in [11,21,33]:
                temp=int(j/10)-1
                if image_gif_changed_tg[i][temp]==1:
                    #result_window.after(0,self.play_gif,0,i,temp,res_item_list[i][str(j)],image_gif_changed,1,0,1)
                    self.play_gif(0,i,temp,res_item_list[i][str(j)],image_gif_changed,1,0,1)
            for j in [41,42,43]:
                temp=j-41
                if siroco_gif_changed_tg[i][temp]==1:
                    #result_window.after(0,self.play_gif,0,i,temp,res_item_list[i][str(j)],siroco_gif_changed,1,0,1)
                    self.play_gif(0,i,temp,res_item_list[i][str(j)],siroco_gif_changed,1,0,1)
        self.show_result_dealer()

    ##태극검 음양 전환
    def change_tagk(self, ele_skill):
        canvas_res = self.canvas_res
        rank_wep_name = self.deal_rank_wep_name[0]
        rank0_wep_name = self.deal_rank_wep_name[1]
        rank_dam = self.rank_dam[0]
        rank0_dam = self.rank_dam[1]
        rank_stat = self.deal_rank_stat1[0]
        rank_stat2 = self.deal_rank_stat2[0]
        rank0_stat = self.deal_rank_stat1[1]
        rank0_stat2 = self.deal_rank_stat2[1]
        res_dam = self.res_dam
        res_stat = self.res_stat
        res_stat2 = self.res_stat2
        rank_dam_tagk = self.rank_dam_tagk[0]
        rank0_dam_tagk = self.rank_dam_tagk[1]
        rank_dam_noele = self.rank_dam_noele[0]
        rank0_dam_noele = self.rank_dam_noele[1]
        rank_dam_tagk_noele = self.rank_dam_tagk_noele[0]
        rank0_dam_tagk_noele = self.rank_dam_tagk_noele[1]
        rank_stat_tagk = self.rank_stat_tagk_um[0]
        rank0_stat_tagk = self.rank_stat_tagk_um[1]
        rank_stat_tagk2 = self.rank_stat_tagk_yang[0]
        rank0_stat_tagk2 = self.rank_stat_tagk_yang[1]
        res_dam_list = self.res_dam_list

        now = self.now_rank_num
        tagkgum_img=self.get_photo_image('ext_img/tagk_um.png')
        tagkgum_img2=self.get_photo_image('ext_img/tagk_ang.png')
        if self.tg_groggy == 0:
            c_rank_dam_tagk=rank_dam_tagk
            c_rank_stat_tagk=rank_stat_tagk
            c_rank_stat_tagk2=rank_stat_tagk2
            c_rank_dam_tagk_noele=rank_dam_tagk_noele
            c_rank_dam=rank_dam
            c_rank_stat=rank_stat
            c_rank_stat2=rank_stat2
            c_rank_dam_noele=rank_dam_noele
            c_rank_dam=rank_dam
            c_rank_wep_name=rank_wep_name
        elif self.tg_groggy == 1:
            c_rank_dam_tagk=rank0_dam_tagk
            c_rank_stat_tagk=rank0_stat_tagk
            c_rank_stat_tagk2=rank0_stat_tagk2
            c_rank_dam_tagk_noele=rank0_dam_tagk_noele
            c_rank_dam=rank0_dam
            c_rank_stat=rank0_stat
            c_rank_stat2=rank0_stat2
            c_rank_dam_noele=rank0_dam_noele
            c_rank_dam=rank0_dam
            c_rank_wep_name=rank0_wep_name
        if c_rank_wep_name[self.now_rank_num] == "(도)태극천제검":
            if self.tagk_tg == 0:
                canvas_res.itemconfig(res_dam,text=c_rank_dam_tagk[now])
                canvas_res.itemconfig(res_stat,text=c_rank_stat_tagk[now])
                canvas_res.itemconfig(res_stat2,text=c_rank_stat_tagk2[now])
                if ele_skill !=0:
                    canvas_res.itemconfig(self.res_ele,text="자속강X="+str(c_rank_dam_tagk_noele[now])+"%")
                canvas_res.itemconfig(self.res_wep,fill='red')

            elif self.tagk_tg == 1:
                canvas_res.itemconfig(res_dam,text=c_rank_dam[now])
                canvas_res.itemconfig(res_stat,text=c_rank_stat[now])
                canvas_res.itemconfig(res_stat2,text=c_rank_stat2[now])
                if ele_skill !=0:
                    canvas_res.itemconfig(self.res_ele,text="자속강X="+str(c_rank_dam_noele[now])+"%")
                canvas_res.itemconfig(self.res_wep,fill='white')
        if self.tagk_tg == 0:
            self.tagk_tg = 1
            self.tagkgum["image"] = tagkgum_img2
            for i in range(0,5):
                try:
                    if c_rank_wep_name[i]=="(도)태극천제검":
                        canvas_res.itemconfig(res_dam_list[i],text=c_rank_dam_tagk[i],fill='red')
                except:
                    pass
        elif self.tagk_tg == 1:
            self.tagk_tg = 0
            self.tagkgum["image"] = tagkgum_img
            for i in range(0,5):
                try:
                    if c_rank_wep_name[i]=="(도)태극천제검":
                        canvas_res.itemconfig(res_dam_list[i],text=c_rank_dam[i],fill='white')
                except:
                    pass

    def change_rank(self, rank_number, job_type, ele_skill, rank_setting, rank_ult):
        threading.Timer(0.05, self.change_rank2,args=(rank_number,job_type,ele_skill)).start()
        threading.Timer(0, self.time_delay).start()

    def change_rank_type(self, in_type):
        threading.Timer(0.05, self.change_rank_type2,args=(in_type,)).start()
        threading.Timer(0, self.time_delayy).start()

    ## 순위 선택 변경
    def change_rank2(self, now, job_type, ele_skill):
        result_window = self.result_window
        canvas_res = self.canvas_res
        gif_images = self.gif_images
        rank_dam = self.rank_dam[0]
        rank0_dam = self.rank_dam[1]

        self.now_rank_num = now
        if job_type =='deal':
            res_dam = self.res_dam
            res_stat = self.res_stat
            res_stat2 = self.res_stat2
            res_stat3 = self.res_stat3

            rank_stat = self.deal_rank_stat1[0]
            rank_stat2 = self.deal_rank_stat2[0]
            rank_stat3 = self.deal_rank_stat3[0]

            rank0_stat = self.deal_rank_stat1[1]
            rank0_stat2 = self.deal_rank_stat2[1]
            rank0_stat3 = self.deal_rank_stat3[1]

            rank_dam_tagk = self.rank_dam_tagk[0]
            rank0_dam_tagk = self.rank_dam_tagk[1]

            rank_dam_noele = self.rank_dam_noele[0]
            rank0_dam_noele = self.rank_dam_noele[1]

            rank_dam_tagk_noele = self.rank_dam_tagk_noele[0]
            rank0_dam_tagk_noele = self.rank_dam_tagk_noele[1]

            rank_inv = self.deal_rank_inv[0]
            rank0_inv = self.deal_rank_inv[1]

            deal_result_image_gif = self.deal_result_image_gif[0]
            deal_result0_image_gif = self.deal_result_image_gif[1]
            deal_result_image_gif_tg = self.deal_result_image_gif_tg[0]
            deal_result0_image_gif_tg = self.deal_result_image_gif_tg[1]

            deal_result_siroco_gif = self.deal_result_siroco_gif[0]
            deal_result0_siroco_gif = self.deal_result_siroco_gif[1]
            deal_result_siroco_gif_tg = self.deal_result_siroco_gif_tg[0]
            deal_result0_siroco_gif_tg = self.deal_result_siroco_gif_tg[1]

            tagkgum_exist = self.tagkgum_exist
            tagk_tg = self.tagk_tg

            try:
                if self.tg_groggy == 0:
                    image_changed = self.deal_result_image_on[0][now]
                    c_rank_wep = self.deal_rank_wep_name[0]
                    if c_rank_wep[now]=="(도)태극천제검" and tagkgum_exist==1 and tagk_tg==1:
                        c_rank_dam=rank_dam_tagk
                        c_rank_stat = self.rank_stat_tagk_um[0]
                        c_rank_stat2 = self.rank_stat_tagk_yang[0]
                        c_rank_dam_noele=rank_dam_tagk_noele
                        canvas_res.itemconfig(self.res_wep, fill="red")
                    else:
                        c_rank_dam=rank_dam
                        c_rank_stat=rank_stat
                        c_rank_stat2=rank_stat2
                        c_rank_dam_noele=rank_dam_noele
                        canvas_res.itemconfig(self.res_wep, fill="white")
                    c_rank_stat3=rank_stat3
                    c_rank_inv=rank_inv
                    image_gif_changed=deal_result_image_gif
                    siroco_gif_changed=deal_result_siroco_gif
                    image_gif_changed_tg=deal_result_image_gif_tg
                    siroco_gif_changed_tg=deal_result_siroco_gif_tg
                elif self.tg_groggy == 1:
                    image_changed = self.deal_result_image_on[1][now]
                    c_rank_wep = self.deal_rank_wep_name[1]
                    if c_rank_wep[now]=="(도)태극천제검" and tagkgum_exist==1 and tagk_tg==1:
                        c_rank_dam=rank0_dam_tagk
                        c_rank_stat = self.rank_stat_tagk_um[1]
                        c_rank_stat2 = self.rank_stat_tagk_yang[1]
                        c_rank_dam_noele=rank0_dam_tagk_noele
                        canvas_res.itemconfig(self.res_wep, fill="red")
                    else:
                        c_rank_dam=rank0_dam
                        c_rank_stat=rank0_stat
                        c_rank_stat2=rank0_stat2
                        c_rank_dam_noele=rank0_dam_noele
                        canvas_res.itemconfig(self.res_wep, fill="white")
                    c_rank_stat3=rank0_stat3
                    c_rank_inv=rank0_inv
                    image_gif_changed=deal_result0_image_gif
                    siroco_gif_changed=deal_result0_siroco_gif
                    image_gif_changed_tg=deal_result0_image_gif_tg
                    siroco_gif_changed_tg=deal_result0_siroco_gif_tg



                canvas_res.itemconfig(res_dam,text=c_rank_dam[now])
                canvas_res.itemconfig(res_stat,text=c_rank_stat[now])
                canvas_res.itemconfig(res_stat2,text=c_rank_stat2[now])
                canvas_res.itemconfig(res_stat3,text=c_rank_stat3[now])
                canvas_res.itemconfig(self.res_inv,text=c_rank_inv[now])
                if ele_skill !=0:
                    canvas_res.itemconfig(self.res_ele,text="자속강X="+str(c_rank_dam_noele[now])+"%")
                self.show_result_dealer()
            except KeyError as error:
                c=1


        elif job_type =='buf':
            rank_wep_name1 = self.buff_rank_wep_name[0]
            rank_wep_name2 = self.buff_rank_wep_name[1]
            rank_wep_name3 = self.buff_rank_wep_name[2]

            result_image_on1 = self.buff_result_image_on[0]
            result_image_on2 = self.buff_result_image_on[1]
            result_image_on3 = self.buff_result_image_on[2]

            res_buf_ex1 = self.res_buf_ex[0]
            res_buf_ex2 = self.res_buf_ex[1]
            res_buf_ex3 = self.res_buf_ex[2]

            buff_result_image_gif1 = self.buff_result_image_gif[0]
            buff_result_image_gif2 = self.buff_result_image_gif[1]
            buff_result_image_gif3 = self.buff_result_image_gif[2]
            buff_result_image_gif1_tg = self.buff_result_image_gif_tg[0]
            buff_result_image_gif2_tg = self.buff_result_image_gif_tg[1]
            buff_result_image_gif3_tg = self.buff_result_image_gif_tg[2]

            buff_result_siroco_gif1 = self.buff_result_siroco_gif[0]
            buff_result_siroco_gif2 = self.buff_result_siroco_gif[1]
            buff_result_siroco_gif3 = self.buff_result_siroco_gif[2]
            buff_result_siroco_gif1_tg = self.buff_result_siroco_gif_tg[0]
            buff_result_siroco_gif2_tg = self.buff_result_siroco_gif_tg[1]
            buff_result_siroco_gif3_tg = self.buff_result_siroco_gif_tg[2]

            try:
                rank_type_buf = self.rank_type_buf

                if rank_type_buf==1:
                    image_changed=result_image_on1[now]
                    rank_changed = self.rank_buf[0][now]
                    rank_buf_ex_changed = self.rank_buf_ex[0]
                    image_gif_changed=buff_result_image_gif1
                    image_gif_changed_tg=buff_result_image_gif1_tg
                    siroco_gif_changed=buff_result_siroco_gif1
                    siroco_gif_changed_tg=buff_result_siroco_gif1_tg
                    c_rank_wep=rank_wep_name1
                elif rank_type_buf==2:
                    image_changed=result_image_on2[now]
                    rank_changed = self.rank_buf[1][now]
                    rank_buf_ex_changed = self.rank_buf_ex[1]
                    image_gif_changed=buff_result_image_gif2
                    image_gif_changed_tg=buff_result_image_gif2_tg
                    siroco_gif_changed=buff_result_siroco_gif2
                    siroco_gif_changed_tg=buff_result_siroco_gif2_tg
                    c_rank_wep=rank_wep_name2
                elif rank_type_buf==3:
                    image_changed=result_image_on3[now]
                    rank_changed = self.rank_buf[2][now]
                    rank_buf_ex_changed = self.rank_buf_ex[2]
                    image_gif_changed=buff_result_image_gif3
                    image_gif_changed_tg=buff_result_image_gif3_tg
                    siroco_gif_changed=buff_result_siroco_gif3
                    siroco_gif_changed_tg=buff_result_siroco_gif3_tg
                    c_rank_wep=rank_wep_name3
                canvas_res.itemconfig(self.res_buf, text=rank_changed)
                canvas_res.itemconfig(res_buf_ex1,text=rank_buf_ex_changed[now][0]) # TODO: non-existing setting causes error
                canvas_res.itemconfig(res_buf_ex2,text=rank_buf_ex_changed[now][1])
                canvas_res.itemconfig(res_buf_ex3,text=rank_buf_ex_changed[now][2])

                if self.buf_jingak_exist == 1:
                    self.buf_jingak["image"] = self.buf_jingak_img[1]
                    canvas_res.itemconfig(self.res_buf, fill='white')
                    self.buf_jingak_tg = 0

            except KeyError as error:
                c=1
        canvas_res.itemconfig(self.res_wep, text=c_rank_wep[now])
        canvas_res.itemconfig(gif_images["11"],image=image_changed['11'])
        canvas_res.itemconfig(gif_images["12"],image=image_changed['12'])
        canvas_res.itemconfig(gif_images["13"],image=image_changed['13'])
        canvas_res.itemconfig(gif_images["14"],image=image_changed['14'])
        canvas_res.itemconfig(gif_images["15"],image=image_changed['15'])
        canvas_res.itemconfig(gif_images["21"],image=image_changed['21'])
        canvas_res.itemconfig(gif_images["22"],image=image_changed['22'])
        canvas_res.itemconfig(gif_images["23"],image=image_changed['23'])
        canvas_res.itemconfig(gif_images["31"],image=image_changed['31'])
        canvas_res.itemconfig(gif_images["32"],image=image_changed['32'])
        canvas_res.itemconfig(gif_images["33"],image=image_changed['33'])
        canvas_res.itemconfig(gif_images["41"],image=image_changed['41'])
        canvas_res.itemconfig(gif_images["42"],image=image_changed['42'])
        canvas_res.itemconfig(gif_images["43"],image=image_changed['43'])
        if image_gif_changed_tg[now][0]==1:
            result_window.after(0,self.play_gif,0,now,0,gif_images["11"],image_gif_changed,0,1,1)
        if image_gif_changed_tg[now][1]==1:
            result_window.after(0,self.play_gif,0,now,1,gif_images["21"],image_gif_changed,0,1,1)
        if image_gif_changed_tg[now][2]==1:
            result_window.after(0,self.play_gif,0,now,2,gif_images["33"],image_gif_changed,0,1,1)
        if siroco_gif_changed_tg[now][0]==1:
            result_window.after(0,self.play_gif,0,now,0,gif_images["41"],siroco_gif_changed,0,1,1)
        if siroco_gif_changed_tg[now][1]==1:
            result_window.after(0,self.play_gif,0,now,1,gif_images["42"],siroco_gif_changed,0,1,1)
        if siroco_gif_changed_tg[now][2]==1:
            result_window.after(0,self.play_gif,0,now,2,gif_images["43"],siroco_gif_changed,0,1,1)

    ## 에픽 이미지 세트옵션 보이기 전환
    def show_set_name(self, job_type):
        canvas_res = self.canvas_res
        image_list = self.image_list
        image_list_tag = self.image_list_tag
        gif_images = self.gif_images
        now_rank_num = self.now_rank_num
        result_image_tag = self.deal_result_image_tag[0]
        result0_image_tag = self.deal_result_image_tag[1]

        if job_type == "deal":
            if self.tg_groggy == 0:
                temp_image_tag=result_image_tag
            elif self.tg_groggy == 1:
                temp_image_tag=result0_image_tag

            if self.set_name_toggle == 0:
                self.set_name_toggle = 1
                self.pause_gif = 1
                canvas_res.itemconfig(gif_images["11"],image=image_list_tag[temp_image_tag[now_rank_num]['11']])
                canvas_res.itemconfig(gif_images["12"],image=image_list_tag[temp_image_tag[now_rank_num]['12']])
                canvas_res.itemconfig(gif_images["13"],image=image_list_tag[temp_image_tag[now_rank_num]['13']])
                canvas_res.itemconfig(gif_images["14"],image=image_list_tag[temp_image_tag[now_rank_num]['14']])
                canvas_res.itemconfig(gif_images["15"],image=image_list_tag[temp_image_tag[now_rank_num]['15']])
                canvas_res.itemconfig(gif_images["21"],image=image_list_tag[temp_image_tag[now_rank_num]['21']])
                canvas_res.itemconfig(gif_images["22"],image=image_list_tag[temp_image_tag[now_rank_num]['22']])
                canvas_res.itemconfig(gif_images["23"],image=image_list_tag[temp_image_tag[now_rank_num]['23']])
                canvas_res.itemconfig(gif_images["31"],image=image_list_tag[temp_image_tag[now_rank_num]['31']])
                canvas_res.itemconfig(gif_images["32"],image=image_list_tag[temp_image_tag[now_rank_num]['32']])
                canvas_res.itemconfig(gif_images["33"],image=image_list_tag[temp_image_tag[now_rank_num]['33']])
                canvas_res.itemconfig(gif_images["41"],image=image_list_tag[temp_image_tag[now_rank_num]['41']])
                canvas_res.itemconfig(gif_images["42"],image=image_list_tag[temp_image_tag[now_rank_num]['42']])
                canvas_res.itemconfig(gif_images["43"],image=image_list_tag[temp_image_tag[now_rank_num]['43']])
            elif self.set_name_toggle == 1:
                self.set_name_toggle = 0
                self.pause_gif = 0
                canvas_res.itemconfig(gif_images["11"],image=image_list[temp_image_tag[now_rank_num]['11']])
                canvas_res.itemconfig(gif_images["12"],image=image_list[temp_image_tag[now_rank_num]['12']])
                canvas_res.itemconfig(gif_images["13"],image=image_list[temp_image_tag[now_rank_num]['13']])
                canvas_res.itemconfig(gif_images["14"],image=image_list[temp_image_tag[now_rank_num]['14']])
                canvas_res.itemconfig(gif_images["15"],image=image_list[temp_image_tag[now_rank_num]['15']])
                canvas_res.itemconfig(gif_images["21"],image=image_list[temp_image_tag[now_rank_num]['21']])
                canvas_res.itemconfig(gif_images["22"],image=image_list[temp_image_tag[now_rank_num]['22']])
                canvas_res.itemconfig(gif_images["23"],image=image_list[temp_image_tag[now_rank_num]['23']])
                canvas_res.itemconfig(gif_images["31"],image=image_list[temp_image_tag[now_rank_num]['31']])
                canvas_res.itemconfig(gif_images["32"],image=image_list[temp_image_tag[now_rank_num]['32']])
                canvas_res.itemconfig(gif_images["33"],image=image_list[temp_image_tag[now_rank_num]['33']])
                canvas_res.itemconfig(gif_images["41"],image=image_list[temp_image_tag[now_rank_num]['41']])
                canvas_res.itemconfig(gif_images["42"],image=image_list[temp_image_tag[now_rank_num]['42']])
                canvas_res.itemconfig(gif_images["43"],image=image_list[temp_image_tag[now_rank_num]['43']])
        elif job_type == "buf":
            rank_type_buf = self.rank_type_buf

            if rank_type_buf in [1, 2, 3]:
                temp_image_tag = self.result_image_on_tag[rank_type_buf - 1]

            if self.set_name_toggle == 0:
                self.set_name_toggle = 1
                self.pause_gif = 1
                canvas_res.itemconfig(gif_images["11"],image=image_list_tag[temp_image_tag[now_rank_num]['11']])
                canvas_res.itemconfig(gif_images["12"],image=image_list_tag[temp_image_tag[now_rank_num]['12']])
                canvas_res.itemconfig(gif_images["13"],image=image_list_tag[temp_image_tag[now_rank_num]['13']])
                canvas_res.itemconfig(gif_images["14"],image=image_list_tag[temp_image_tag[now_rank_num]['14']])
                canvas_res.itemconfig(gif_images["15"],image=image_list_tag[temp_image_tag[now_rank_num]['15']])
                canvas_res.itemconfig(gif_images["21"],image=image_list_tag[temp_image_tag[now_rank_num]['21']])
                canvas_res.itemconfig(gif_images["22"],image=image_list_tag[temp_image_tag[now_rank_num]['22']])
                canvas_res.itemconfig(gif_images["23"],image=image_list_tag[temp_image_tag[now_rank_num]['23']])
                canvas_res.itemconfig(gif_images["31"],image=image_list_tag[temp_image_tag[now_rank_num]['31']])
                canvas_res.itemconfig(gif_images["32"],image=image_list_tag[temp_image_tag[now_rank_num]['32']])
                canvas_res.itemconfig(gif_images["33"],image=image_list_tag[temp_image_tag[now_rank_num]['33']])
                canvas_res.itemconfig(gif_images["41"],image=image_list_tag[temp_image_tag[now_rank_num]['41']])
                canvas_res.itemconfig(gif_images["42"],image=image_list_tag[temp_image_tag[now_rank_num]['42']])
                canvas_res.itemconfig(gif_images["43"],image=image_list_tag[temp_image_tag[now_rank_num]['43']])
            elif self.set_name_toggle == 1:
                self.set_name_toggle = 0
                self.pause_gif = 0
                canvas_res.itemconfig(gif_images["11"],image=image_list[temp_image_tag[now_rank_num]['11']])
                canvas_res.itemconfig(gif_images["12"],image=image_list[temp_image_tag[now_rank_num]['12']])
                canvas_res.itemconfig(gif_images["13"],image=image_list[temp_image_tag[now_rank_num]['13']])
                canvas_res.itemconfig(gif_images["14"],image=image_list[temp_image_tag[now_rank_num]['14']])
                canvas_res.itemconfig(gif_images["15"],image=image_list[temp_image_tag[now_rank_num]['15']])
                canvas_res.itemconfig(gif_images["21"],image=image_list[temp_image_tag[now_rank_num]['21']])
                canvas_res.itemconfig(gif_images["22"],image=image_list[temp_image_tag[now_rank_num]['22']])
                canvas_res.itemconfig(gif_images["23"],image=image_list[temp_image_tag[now_rank_num]['23']])
                canvas_res.itemconfig(gif_images["31"],image=image_list[temp_image_tag[now_rank_num]['31']])
                canvas_res.itemconfig(gif_images["32"],image=image_list[temp_image_tag[now_rank_num]['32']])
                canvas_res.itemconfig(gif_images["33"],image=image_list[temp_image_tag[now_rank_num]['33']])
                canvas_res.itemconfig(gif_images["41"],image=image_list[temp_image_tag[now_rank_num]['41']])
                canvas_res.itemconfig(gif_images["42"],image=image_list[temp_image_tag[now_rank_num]['42']])
                canvas_res.itemconfig(gif_images["43"],image=image_list[temp_image_tag[now_rank_num]['43']])


    ## 버퍼용 축복/1각/종합 버프력 전환
    def change_rank_type2(self, in_type):
        result_window = self.result_window
        canvas_res = self.canvas_res
        gif_images = self.gif_images
        rank_wep_name1 = self.buff_rank_wep_name[0]
        rank_wep_name2 = self.buff_rank_wep_name[1]
        rank_wep_name3 = self.buff_rank_wep_name[2]
        rank_wep_img1 = self.buff_rank_wep_img[0]
        rank_wep_img2 = self.buff_rank_wep_img[1]
        rank_wep_img3 = self.buff_rank_wep_img[2]
        result_image_on1 = self.buff_result_image_on[0]
        result_image_on2 = self.buff_result_image_on[1]
        result_image_on3 = self.buff_result_image_on[2]
        res_img_list = self.res_img_list
        res_buf_ex1 = self.res_buf_ex[0]
        res_buf_ex2 = self.res_buf_ex[1]
        res_buf_ex3 = self.res_buf_ex[2]

        buff_result_image_gif1 = self.buff_result_image_gif[0]
        buff_result_image_gif2 = self.buff_result_image_gif[1]
        buff_result_image_gif3 = self.buff_result_image_gif[2]
        buff_result_image_gif1_tg = self.buff_result_image_gif_tg[0]
        buff_result_image_gif2_tg = self.buff_result_image_gif_tg[1]
        buff_result_image_gif3_tg = self.buff_result_image_gif_tg[2]

        buff_result_siroco_gif1 = self.buff_result_siroco_gif[0]
        buff_result_siroco_gif2 = self.buff_result_siroco_gif[1]
        buff_result_siroco_gif3 = self.buff_result_siroco_gif[2]
        buff_result_siroco_gif1_tg = self.buff_result_siroco_gif_tg[0]
        buff_result_siroco_gif2_tg = self.buff_result_siroco_gif_tg[1]
        buff_result_siroco_gif3_tg = self.buff_result_siroco_gif_tg[2]

        self.now_rank_num = 0
        if in_type==1:
            self.rank_type_buf = 1
            image_changed=result_image_on1[0]
            image_changed_all=result_image_on1
            rank_changed = self.rank_buf[0]
            rank_buf_ex_changed = self.rank_buf_ex[0]
            type_changed="축복 기준"
            image_gif_changed=buff_result_image_gif1
            image_gif_changed_tg=buff_result_image_gif1_tg
            siroco_gif_changed=buff_result_siroco_gif1
            siroco_gif_changed_tg=buff_result_siroco_gif1_tg
            c_rank_wep=rank_wep_name1
            wep_img_changed=rank_wep_img1
        elif in_type==2:
            self.rank_type_buf = 2
            image_changed=result_image_on2[0]
            image_changed_all=result_image_on2
            rank_changed = self.rank_buf[1]
            rank_buf_ex_changed = self.rank_buf_ex[1]
            type_changed="1각 기준"
            image_gif_changed=buff_result_image_gif2
            image_gif_changed_tg=buff_result_image_gif2_tg
            siroco_gif_changed=buff_result_siroco_gif2
            siroco_gif_changed_tg=buff_result_siroco_gif2_tg
            c_rank_wep=rank_wep_name2
            wep_img_changed=rank_wep_img2
        elif in_type==3:
            self.rank_type_buf = 3
            image_changed=result_image_on3[0]
            image_changed_all=result_image_on3
            rank_changed = self.rank_buf[2]
            rank_buf_ex_changed = self.rank_buf_ex[2]
            type_changed="총합 기준"
            image_gif_changed=buff_result_image_gif3
            image_gif_changed_tg=buff_result_image_gif3_tg
            siroco_gif_changed=buff_result_siroco_gif3
            siroco_gif_changed_tg=buff_result_siroco_gif3_tg
            c_rank_wep=rank_wep_name3
            wep_img_changed=rank_wep_img3
        canvas_res.itemconfig(self.res_wep, text=c_rank_wep[0])
        canvas_res.itemconfig(self.res_buf_type_what, text=type_changed)
        canvas_res.itemconfig(res_buf_ex1,text=rank_buf_ex_changed[0][0])
        canvas_res.itemconfig(res_buf_ex2,text=rank_buf_ex_changed[0][1])
        canvas_res.itemconfig(res_buf_ex3,text=rank_buf_ex_changed[0][2])
        canvas_res.itemconfig(self.res_buf, text=rank_changed[0])
        canvas_res.itemconfig(gif_images["11"],image=image_changed['11'])
        canvas_res.itemconfig(gif_images["12"],image=image_changed['12'])
        canvas_res.itemconfig(gif_images["13"],image=image_changed['13'])
        canvas_res.itemconfig(gif_images["14"],image=image_changed['14'])
        canvas_res.itemconfig(gif_images["15"],image=image_changed['15'])
        canvas_res.itemconfig(gif_images["21"],image=image_changed['21'])
        canvas_res.itemconfig(gif_images["22"],image=image_changed['22'])
        canvas_res.itemconfig(gif_images["23"],image=image_changed['23'])
        canvas_res.itemconfig(gif_images["31"],image=image_changed['31'])
        canvas_res.itemconfig(gif_images["32"],image=image_changed['32'])
        canvas_res.itemconfig(gif_images["33"],image=image_changed['33'])
        canvas_res.itemconfig(gif_images["41"],image=image_changed['41'])
        canvas_res.itemconfig(gif_images["42"],image=image_changed['42'])
        canvas_res.itemconfig(gif_images["43"],image=image_changed['43'])
        self.stop_gif = 1
        self.stop_gif2 = 1
        time.sleep(0.2)
        self.stop_gif = 0
        self.stop_gif2 = 0

        if image_gif_changed_tg[0][0]==1:
            result_window.after(0,self.play_gif,0,0,0,gif_images["11"],image_gif_changed,0,1,1)
        if image_gif_changed_tg[0][1]==1:
            result_window.after(0,self.play_gif,0,0,1,gif_images["21"],image_gif_changed,0,1,1)
        if image_gif_changed_tg[0][2]==1:
            result_window.after(0,self.play_gif,0,0,2,gif_images["33"],image_gif_changed,0,1,1)
        if siroco_gif_changed_tg[0][0]==1:
            result_window.after(0,self.play_gif,0,0,0,gif_images["41"],siroco_gif_changed,0,1,1)
        if siroco_gif_changed_tg[0][1]==1:
            result_window.after(0,self.play_gif,0,0,1,gif_images["42"],siroco_gif_changed,0,1,1)
        if siroco_gif_changed_tg[0][2]==1:
            result_window.after(0,self.play_gif,0,0,2,gif_images["43"],siroco_gif_changed,0,1,1)
        cn2=0
        for j in range(0,5):
                try:
                    for i in [11,12,13,14,15,21,22,23,31,32,33]:
                        canvas_res.itemconfig(res_img_list[str(j)+str(i)],image=image_changed_all[j][str(i)])
                        cn2=cn2+2
                    cn2=0
                    canvas_res.itemconfig(self.res_buf_list[j],text=rank_changed[j],font=self.mid_font,fill='white')
                    canvas_res.itemconfig(self.res_wep_img[j], image=wep_img_changed[j])
                except KeyError as error:
                    c=1
        for i in range(0,5):
            for j in [11,21,33]:
                temp=int(j/10)-1
                if image_gif_changed_tg[i][temp]==1:
                    result_window.after(0,self.play_gif,0,i,temp,res_img_list[str(i)+str(j)],image_gif_changed,1,0,1)
            for j in [41,42,43]:
                temp=j-41
                if siroco_gif_changed_tg[i][temp]==1:
                    result_window.after(0,self.play_gif,0,i,temp,res_img_list[str(i)+str(j)],siroco_gif_changed,1,0,1)

        if self.buf_jingak_exist == 1:
            self.buf_jingak["image"] = self.buf_jingak_img[1]
            canvas_res.itemconfig(self.res_buf, fill='white')
            self.buf_jingak_tg = 0


    ## 실시간 갱신 카운터 (1: 계산 카운트 / 2: 경우의 수 카운트)
    def update_count(self):
        while True:
            text = f"{self.count_num.value}유효/{self.count_all.value}무효\n{self.all_list_list_num}전체"
            self.change_case_count_text(text)
            time.sleep(0.1)

    def update_count2(self):
        select_item = self.owned_equipments

        while True:
            self.a_num_all = 0
            a_num=[0,0,0,0,0,0,0,0,0,0,0]
            for i in range(101,136):
                try:
                    a_num[0]=a_num[0]+select_item['tg1{}0'.format(i)]+select_item['tg1{}1'.format(i)]
                except KeyError as error:
                    pass
                try:
                    a_num[1]=a_num[1]+select_item['tg1{}0'.format(i+100)]
                except KeyError as error:
                    pass
                try:
                    a_num[2]=a_num[2]+select_item['tg1{}0'.format(i+200)]
                except KeyError as error:
                    pass
                try:
                    a_num[3]=a_num[3]+select_item['tg1{}0'.format(i+300)]
                except KeyError as error:
                    pass
                try:
                    a_num[4]=a_num[4]+select_item['tg1{}0'.format(i+400)]
                except KeyError as error:
                    pass
                try:
                    a_num[5]=a_num[5]+select_item['tg2{}0'.format(i)]+select_item['tg2{}1'.format(i)]
                except KeyError as error:
                    pass
                try:
                    a_num[6]=a_num[6]+select_item['tg2{}0'.format(i+100)]
                except KeyError as error:
                    pass
                try:
                    a_num[7]=a_num[7]+select_item['tg2{}0'.format(i+200)]
                except KeyError as error:
                    pass
                try:
                    a_num[8]=a_num[8]+select_item['tg3{}0'.format(i)]
                except KeyError as error:
                    pass
                try:
                    a_num[9]=a_num[9]+select_item['tg3{}0'.format(i+100)]
                except KeyError as error:
                    pass
                try:
                    a_num[10]=a_num[10]+select_item['tg3{}0'.format(i+200)]+select_item['tg3{}1'.format(i+200)]
                except KeyError as error:
                    pass
            for i in range(0,5):
                if a_num[i]==0:
                    a_num[0]=a_num[0]+1;a_num[1]=a_num[1]+1;a_num[2]=a_num[2]+1;a_num[3]=a_num[3]+1;a_num[4]=a_num[4]+1;
                    break
            if a_num[5]+a_num[6]+a_num[7]<3:
                a_num[5]=a_num[5]+1;a_num[6]=a_num[6]+1;a_num[7]=a_num[7]+1
            if a_num[8]+a_num[9]+a_num[10]<3:
                a_num[8]=a_num[8]+1;a_num[9]=a_num[9]+1;a_num[10]=a_num[10]+1
            if a_num[5]==0:
                a_num[5]=a_num[5]+1
            if a_num[6]==0:
                a_num[6]=a_num[6]+1
            if a_num[7]==0:
                a_num[7]=a_num[7]+1
            if a_num[8]==0:
                a_num[8]=a_num[8]+1
            if a_num[9]==0:
                a_num[9]=a_num[9]+1
            if a_num[10]==0:
                a_num[10]=a_num[10]+1
            wep_num = len(self.wep_name_list)
            if wep_num==0: wep_num=1
            self.a_num_all = a_num[0]*a_num[1]*a_num[2]*a_num[3]*a_num[4]*a_num[5]*a_num[6]*a_num[7]*a_num[8]*a_num[9]*a_num[10]*wep_num

            if self.a_num_all > 10000000:
                total_case_text_color = "red"
            else:
                total_case_text_color = "white"

            self.change_total_count_text(f"경우의 수={self.a_num_all}", total_case_text_color)
            time.sleep(1)

    def update_thread(self):
        threading.Thread(target=self.update_count, daemon=True).start()
    def update_thread2(self):
        threading.Thread(target=self.update_count2, daemon=True).start()


    ## 실제 저장 토글값과 이미지 표시값 동기화
    def check_equipment(self):
        select_item = self.owned_equipments
        equip_buttons = self.equip_buttons

        for i in range(11010,43551):
            try:
                str_i = str(i)
                if select_item[f"tg{i}"]==0:
                    equip_buttons[str_i]['image'] = self.image_list2[str_i]
                elif select_item[f"tg{i}"]==1:
                    equip_buttons[str_i]['image'] = self.image_list[str_i]
            except:
                pass

        for i in self.get_all_knowledge_equipment_list():
            try:
                str_i = str(i)
                if select_item[f"tg{i}"]==0:
                    equip_buttons[str_i]['image'] = self.image_list2[str_i]
                elif select_item[f"tg{i}"]==1:
                    equip_buttons[str_i]['image'] = self.image_list[str_i]
            except:
                pass


    ## 세트명 태그 점등 여부와 실제 토글값 동기화
    def check_set(self, code):
        select_item = self.owned_equipments
        image_list_set = self.image_list_set
        image_list_set2 = self.image_list_set2
        set_buttons = self.set_buttons

        code_str=str(code)[1:3]
        set_checked=0
        if code < 116:
            for i in [11,12,13,14,15]:
                if select_item['tg'+str(i)+code_str+'0']==1:
                    set_checked=set_checked+1
        elif code < 120:
            for i in [21,22,23]:
                if select_item['tg'+str(i)+code_str+'0']==1:
                    set_checked=set_checked+1
        elif code < 124:
            for i in [31,32,33]:
                if select_item['tg'+str(i)+code_str+'0']==1:
                    set_checked=set_checked+1
        elif code < 128:
            for i in [12,21,32]:
                if select_item['tg'+str(i)+code_str+'0']==1:
                    set_checked=set_checked+1
        elif code < 132:
            for i in [11,22,31]:
                if select_item['tg'+str(i)+code_str+'0']==1:
                    set_checked=set_checked+1
        elif code < 136:
            for i in [15,23,33]:
                if select_item['tg'+str(i)+code_str+'0']==1:
                    set_checked=set_checked+1
        elif 150 < code < 156:
            for i in [41,42,43]:
                if select_item['tg'+str(i)+code_str+'0']==1:
                    set_checked=set_checked+1

        if code < 116:
            if set_checked==5:
                set_buttons[str(code)]['image']=image_list_set[str(code)]
            else:
                set_buttons[str(code)]['image']=image_list_set2[str(code)]
        else:
            if set_checked==3:
                set_buttons[str(code)]['image']=image_list_set[str(code)]
            else:
                set_buttons[str(code)]['image']=image_list_set2[str(code)]

    ## 딜러 프로필 보기 기능
    def show_profile2(self, name, server):
        try:
            def_result=calc_profile.make_profile(name,server)
        except:
            tkinter.messagebox.showerror('에러',"API조회 에러(네트워크 오류)")
            return
        if def_result=={'error':'Not found'}:
            tkinter.messagebox.showerror('에러',"서버/캐릭명을 확인하세요.")
            return
        if def_result=={'error':'buffer'}:
            tkinter.messagebox.showerror('에러',"버퍼는 지원하지 않습니다.")
            return
        setting_str=def_result[0]
        setting_dict=def_result[1]

        profile_window = tkinter.Toplevel(self.window)
        profile_window.geometry("415x277")
        profile_window.resizable(False,False)
        canvas = Canvas(profile_window, width=417, height=297, bd=0)
        canvas.place(x=-2,y=-2)
        cha_bg=self.get_photo_image('ext_img/bg_info.png')
        canvas.create_image(0,0,image=cha_bg,anchor='nw')
        cha_img=self.get_photo_image('my_cha.png')
        canvas.create_image(123,70,image=cha_img)
        image_on={}
        def play_gif_cha(count_frame,now_pc,show_res,gif_list):
            #now_pc:0(상의,하의),1(팔찌,반지),2(귀걸,보장)
            #show_res:이미지 재생될 canvas 객체
            now_frame=gif_list[now_pc][int(count_frame)]
            count_frame += 1
            canvas.itemconfig(show_res,image=now_frame)
            if count_frame >=len(gif_list[now_pc]):
                profile_window.after(100, play_gif_cha, 0,now_pc,show_res,gif_list)
            else:
                profile_window.after(100, play_gif_cha, count_frame,now_pc,show_res,gif_list)

        image_list = self.image_list
        cha_god_gif=[None,None,None]
        for i in [11,12,13,14,15,21,22,23,31,32,33]:
            for j in setting_dict['장비']:
                if len(j)!=6:
                    if j[0:2] == str(i):
                        image_on[str(i)]=image_list[j]
                    if len(j)==5 and j[4]=='1':
                        cha_god_gif[int(j[0])-1]=calc_gif.img_gif(j,0)
        cha_siroco_gif=[None,None,None]
        for i in setting_dict['장비']:
            if len(i)==4 and i[0]=='4':
                image_on['41']=image_list['415'+i[1]+'0']
                image_on['42']=image_list['425'+i[2]+'0']
                image_on['43']=image_list['435'+i[3]+'0']
                if i[1]!='0':
                    cha_siroco_gif[0]=calc_gif.img_gif('415'+i[1]+'0',1)
                if i[2]!='0':
                    cha_siroco_gif[1]=calc_gif.img_gif('425'+i[2]+'0',1)
                if i[3]!='0':
                    cha_siroco_gif[2]=calc_gif.img_gif('435'+i[3]+'0',1)
        img11=canvas.create_image(57,52,image=image_on['11'])
        img12=canvas.create_image(27,82,image=image_on['12'])
        img13=canvas.create_image(27,52,image=image_on['13'])
        img14=canvas.create_image(57,82,image=image_on['14'])
        img15=canvas.create_image(27,112,image=image_on['15'])
        img21=canvas.create_image(189,52,image=image_on['21'])
        img22=canvas.create_image(219,52,image=image_on['22'])
        img23=canvas.create_image(219,82,image=image_on['23'])
        img31=canvas.create_image(189,82,image=image_on['31'])
        img32=canvas.create_image(219,112,image=image_on['32'])
        img33=canvas.create_image(189,112,image=image_on['33'])
        img41=canvas.create_image(27,82,image=image_on['41'])
        img42=canvas.create_image(219,82,image=image_on['42'])
        img43=canvas.create_image(189,82,image=image_on['43'])
        for i in range(0,3):
            if i==0: show_res1=img11;show_res2=img41
            elif i==1: show_res1=img21;show_res2=img42
            elif i==2: show_res1=img33;show_res2=img43
            if cha_god_gif[i]!=None: play_gif_cha(0,i,show_res1,cha_god_gif)
            if cha_siroco_gif[i]!=None: play_gif_cha(0,i,show_res2,cha_siroco_gif)


        canvas.create_text(233,17,text=setting_dict['무기명'],font=self.guide_font,fill='white',anchor='e')
        canvas.create_text(10,170,text=setting_dict['캐릭명'],fill='white',anchor='w')
        canvas.create_text(10,185,text=setting_dict['직업명'],fill='white',anchor='w')
        canvas.create_text(233,170,text=setting_dict['모험단'],fill='white',anchor='e')
        canvas.create_text(233,185,text=server,fill='white',anchor='e')

        show_font=tkinter.font.Font(family="맑은 고딕", size=15, weight='bold')

        if float(setting_dict['종합점수'][:-1]) >=110: rank_color='gold2'
        elif float(setting_dict['종합점수'][:-1]) >=100: rank_color='deep sky blue'
        elif float(setting_dict['종합점수'][:-1]) >=80: rank_color='white'
        else: rank_color='silver'


        canvas.create_text(18,217,text='장비%=',font=show_font,fill='white',anchor='w')
        canvas.create_text(18+73,217,text=setting_dict['장비딜'],font=show_font,fill='white',anchor='w')
        canvas.create_text(20+189,220,text='쿨감기대값\n='+setting_dict['쿨감'],font=self.small_font,fill='white', anchor='c')
        canvas.create_text(18,252,text='세팅%=',font=show_font,fill='white',anchor='w')
        canvas.create_text(18+73,252,text=setting_dict['종합점수'],font=show_font,fill=rank_color,anchor='w')

        canvas.create_text(254,22,text='강화/스탯\n',font=self.guide_font,fill='white',anchor='w')
        canvas.create_text(254,6+25,text=setting_dict['스탯'],font=self.mid_font,fill='white',anchor='w')
        canvas.create_text(254+83,14,text='무기='+setting_dict['무기강화'],font=self.guide_font,fill='white',anchor='w')
        canvas.create_text(254+83,33,text='스탯='+setting_dict['스탯상세'],font=self.guide_font,fill='white',anchor='w')

        canvas.create_text(254,22+50,text='속강작\n',font=self.guide_font,fill='white',anchor='w')
        canvas.create_text(254,6+75,text=setting_dict['속강'],font=self.mid_font,fill='white',anchor='w')
        if setting_dict['속강종류']=='화': ele_type='火'; ele_color='red'
        elif setting_dict['속강종류']=='수': ele_type='水'; ele_color='blue'
        elif setting_dict['속강종류']=='명': ele_type='明'; ele_color='yellow'
        elif setting_dict['속강종류']=='암': ele_type='暗'; ele_color='purple'
        elif setting_dict['속강종류']=='모': ele_type='某'; ele_color='white'
        canvas.create_text(254+83,22+50,text=ele_type,font=self.mid_font,fill=ele_color,anchor='w')
        canvas.create_text(254+83+23,22+50,text='+'+setting_dict['속강상세'],font=self.mid_font,fill='white',anchor='w')

        canvas.create_text(254,22+100,text='딜플티\n',font=self.guide_font,fill='white',anchor='w')
        canvas.create_text(254,6+125,text=setting_dict['플티'],font=self.mid_font,fill='white',anchor='w')
        plt_img=[0,0]
        for i in [0,1]:
            if setting_dict['플티상세'][i]=='S': plt_img[i]=self.get_photo_image('ext_img/plt_best.png')
            elif setting_dict['플티상세'][i]=='A': plt_img[i]=self.get_photo_image('ext_img/plt_good.png')
            elif setting_dict['플티상세'][i]=='B': plt_img[i]=self.get_photo_image('ext_img/plt_active.png')
            elif setting_dict['플티상세'][i]=='C': plt_img[i]=self.get_photo_image('ext_img/plt_common.png')
            else: plt_img[i]=self.get_photo_image('ext_img/plt_nope.png')
        img_plt1=canvas.create_image(350,124,image=plt_img[0]);img_plt2=canvas.create_image(390,124,image=plt_img[1])

        canvas.create_text(254,22+150,text='룬/탈리\n',font=self.guide_font,fill='white',anchor='w')
        canvas.create_text(254,6+175,text=setting_dict['탈리'],font=self.mid_font,fill='white',anchor='w')
        tal_img=[0,0]
        for i in [0,1]:
            if setting_dict['탈리상세'][i]=='S': tal_img[i]=self.get_photo_image('ext_img/talisman_unique.png')
            elif setting_dict['탈리상세'][i]=='A': tal_img[i]=self.get_photo_image('ext_img/talisman_rare.png')
            elif setting_dict['탈리상세'][i]=='B': tal_img[i]=self.get_photo_image('ext_img/talisman_common.png')
            else: tal_img[i]=self.get_photo_image('ext_img/talisman_nope.png')
        img_tal1=canvas.create_image(350,170,image=tal_img[0]);img_tal2=canvas.create_image(390,170,image=tal_img[1])
        rune_img=[0,0,0,0,0,0]
        for i in range(0,6):
            if i>=3: shift=8
            else: shift=0
            if setting_dict['룬상세'][i]=='S': rune_color='purple'
            elif setting_dict['룬상세'][i]=='A': rune_color='deep sky blue'
            elif setting_dict['룬상세'][i]=='B': rune_color='silver'
            else: rune_color='black'
            rune_img[i]=canvas.create_rectangle(336+i*11+shift,184,336+6+i*11+shift,184+7,fill=rune_color,width=0)

        canvas.create_text(254,22+200,text='스위칭\n',font=self.guide_font,fill='white',anchor='w')
        canvas.create_text(254,6+225,text=setting_dict['스위칭'],font=self.mid_font,fill='white',anchor='w')
        canvas.create_text(254+83,22+191,text=setting_dict['스위칭상세'],font=self.guide_font,fill='white',anchor='w')
        canvas.create_text(254+83+30,22+210,text=setting_dict['스위칭최대'],font=self.guide_font,fill='white',anchor='w')


        capture_but=tkinter.Button(profile_window,command=lambda:calc_profile.make_profile_image(name,server,def_result),image=self.capture_img,bg=self.dark_sub,borderwidth=0,activebackground=self.dark_sub,anchor='nw')
        capture_but.place(x=378,y=248)
        def profile_detail():
            tkinter.messagebox.showinfo('세부보기',setting_str,parent=profile_window)
        show_detail_img=self.get_photo_image('ext_img/show_detail2.png')
        show_detail=tkinter.Button(profile_window,command=profile_detail,image=show_detail_img,bg=self.dark_sub,borderwidth=0,activebackground=self.dark_sub,anchor='nw')
        show_detail.place(x=179,y=239)

        canvas.create_text(255,262,text='[ESC키로 닫기 가능]',font=self.guide_font,fill='white',anchor='w')

        show_detail.image=show_detail_img
        canvas.image=cha_bg,cha_img,plt_img[0],plt_img[1],tal_img[0],tal_img[1]
        self.place_center(profile_window, 0)
        def exit_p(e):
            profile_window.destroy()
        profile_window.focus_set()
        profile_window.bind("<Escape>", exit_p)

    def show_profile(self, name, server):
        threading.Thread(target=self.show_profile2, args=(name, server), daemon=True).start()

    def wep_img_list_refresh(self, weapon_img_list: List[PhotoImage]):
        for i in range(0, 10):
            try:
                self.change_selected_weapon_img(i, weapon_img_list[i])
            except:
                self.change_selected_weapon_img(i, self.void_weapon_img)

    def sync_wep_list(self):
        wep_name_list = self.wep_name_list

        self.selected_weapon_img_list.clear()
        for wep_name in wep_name_list:
            self.selected_weapon_img_list.append(self.image_list_wep[wep_name])

        self.change_weapon_list_num_text(len(wep_name_list))
        self.wep_img_list_refresh(self.selected_weapon_img_list)

    @staticmethod
    def place_center(toplevel, move_x):
        toplevel.update_idletasks()
        w = toplevel.winfo_screenwidth()
        h = toplevel.winfo_screenheight()
        size = tuple(int(_) for _ in toplevel.geometry().split('+')[0].split('x'))
        x = w/2 - size[0]/2
        y = h/2 - size[1]/2
        toplevel.geometry("%dx%d+%d+%d" % (size + (x+move_x, y)))

    @staticmethod
    def _from_rgb(rgb):
        return "#%02x%02x%02x" % rgb

    @staticmethod
    def get_api_key():
        apikey = ""

        try:
            import calc_api_key
            apikey = calc_api_key.get_api_key()
        except:
            try:
                api_txt_file = open("API_key.txt", "r")
                apikey = api_txt_file.readline()
                if apikey == "":
                    tkinter.messagebox.showerror("에러", "API 접근 권한 획득 실패.")
                api_txt_file.close()
            except:
                tkinter.messagebox.showerror("에러", "API 접근 권한 획득 실패.")

        return apikey


if __name__ == "__main__":
    calculator = Calculator()

    if calculator.auto_custom == 1:
        calculator.create_custom_window(1)

    calculator.update_thread()
    calculator.update_thread2()

    calculator.window.mainloop()
